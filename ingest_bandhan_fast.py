"""
ingest_bandhan_fast.py
----------------------
Parallel Bandhan backfill — downloads scheme files concurrently with
ThreadPoolExecutor to avoid the 44-second shell timeout.

Usage:
  python3 ingest_bandhan_fast.py --period 2024-06
  python3 ingest_bandhan_fast.py --period 2024-06 --period 2024-07  (multi)
  python3 ingest_bandhan_fast.py --from-period 2024-06 --to-period 2024-08

Requires bandhan_monthly_urls.json (built by ingest_new5amcs.py's live API enumerator,
or fetched here if absent).
"""

import os, sys, re, json, io, time, warnings, argparse
from pathlib import Path
from concurrent.futures import ThreadPoolExecutor, as_completed
import requests
import pandas as pd
import numpy as np

warnings.filterwarnings('ignore')

BASE  = Path('/sessions/admiring-nifty-dijkstra/mnt/outputs/mf_data')
HOLD  = BASE / 'holdings'
CACHE = BASE / 'raw_cache'
CACHE.mkdir(parents=True, exist_ok=True)

HEADERS = {
    'User-Agent': 'Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36',
    'Accept': 'text/html,application/xhtml+xml,*/*;q=0.8',
    'Accept-Language': 'en-IN,en;q=0.9',
    'Referer': 'https://bandhanmutual.com/',
}

MAX_WORKERS = 10   # parallel download threads
TIMEOUT     = 30   # per-file request timeout


# ── helpers ──────────────────────────────────────────────────────────────────

def safe_float(v):
    try:
        return float(v)
    except:
        return None

def is_valid_isin(s):
    return isinstance(s, str) and re.match(r'^[A-Z]{2}[A-Z0-9]{10}$', s.strip()) is not None


# ── AMFI lookup (cached to disk as JSON) ──────────────────────────────────────

AMFI_CACHE_PATH = BASE / 'amfi_lookup_cache.json'

def build_amfi_lookup():
    """Fetch NAVAll.txt, return (name_lookup, isin_lookup) for Bandhan."""
    if AMFI_CACHE_PATH.exists():
        age = time.time() - AMFI_CACHE_PATH.stat().st_mtime
        if age < 86400:   # < 1 day old — reuse
            with open(AMFI_CACHE_PATH) as f:
                d = json.load(f)
            name_lk = {tuple(k.split('|||')): v for k, v in d['name'].items()}
            isin_lk = d['isin']
            print(f"  AMFI lookup (cached): {len(name_lk)} name, {len(isin_lk)} ISIN entries")
            return name_lk, isin_lk

    url = "https://portal.amfiindia.com/spages/NAVAll.txt"
    try:
        resp = requests.get(url, headers=HEADERS, timeout=45)
        resp.raise_for_status()
        data = resp.content.decode('utf-8', errors='replace')
    except Exception as e:
        print(f"  WARN: AMFI fetch failed: {e}")
        return {}, {}

    lines = data.splitlines()
    amc_map = {'bandhan mutual fund': 'Bandhan'}
    lookup      = {}
    isin_lookup = {}
    current_amc = None
    for line in lines:
        line = line.strip()
        if not line:
            continue
        if ';' not in line:
            low = line.lower()
            matched = None
            for k, v in amc_map.items():
                if k in low:
                    matched = v
                    break
            current_amc = matched
        elif current_amc:
            parts = line.split(';')
            if len(parts) >= 4:
                code = parts[0].strip()
                isin = parts[1].strip()
                name = parts[3].strip()
                norm = re.sub(r'\s+', ' ', name.lower())
                lookup[(current_amc, norm)] = code
                if isin and isin != '-':
                    isin_lookup[isin] = code

    print(f"  AMFI lookup: {len(lookup)} name, {len(isin_lookup)} ISIN entries for Bandhan")

    # Save cache
    serialisable = {'name': {'|||'.join(k): v for k, v in lookup.items()}, 'isin': isin_lookup}
    with open(AMFI_CACHE_PATH, 'w') as f:
        json.dump(serialisable, f)

    return lookup, isin_lookup


# ── URL catalog ───────────────────────────────────────────────────────────────

BANDHAN_SKIP_KWS = [
    'liquid', 'overnight', 'gilt', 'g-sec', 'debt', 'bond', 'income',
    'ultra short', 'low duration', 'short duration', 'medium duration',
    'credit risk', 'money market', 'floater', 'dynamic bond',
    'fixed maturity', 'fmp', 'interval', 'arbitrage', 'fixed term plan',
    'banking & psu', 'banking and psu', 'ibx', 'crisil',
    'constant duration', 'corporate bond',
]

def is_bandhan_equity_scheme(title):
    t = title.lower()
    return not any(k in t for k in BANDHAN_SKIP_KWS)

def load_url_catalog(periods_wanted):
    """Load bandhan_monthly_urls.json, filtered to periods_wanted set.
    Catalog lives at outputs/bandhan_monthly_urls.json (not inside mf_data/).
    Format: { "YYYY-MM": [ {"url": ..., "title": ..., "filename": ...}, ... ] }
    """
    catalog_path = BASE.parent / 'bandhan_monthly_urls.json'
    if not catalog_path.exists():
        print(f"  ERROR: Catalog not found at {catalog_path}")
        print("  Run: python3 ingest_new5amcs.py --amc meta")
        sys.exit(1)

    with open(catalog_path) as f:
        raw = json.load(f)

    result = {}   # period → list of (url, title)
    for period, entries in raw.items():
        if period not in periods_wanted:
            continue
        for entry in entries:
            url   = entry.get('url', '')
            title = entry.get('title', '')
            if not url or not is_bandhan_equity_scheme(title):
                continue
            result.setdefault(period, []).append((url, title))

    return result


# ── parallel downloader ───────────────────────────────────────────────────────

def _download_one(args):
    """Worker: (url, cache_path) → (url, title, bytes_or_None)"""
    url, title, cache_path = args
    if cache_path.exists() and cache_path.stat().st_size > 1000:
        return url, title, cache_path.read_bytes()
    try:
        resp = requests.get(url, headers=HEADERS, timeout=TIMEOUT, stream=True)
        if resp.status_code == 200:
            data = resp.content
            if len(data) < 1000:
                return url, title, None
            cache_path.write_bytes(data)
            return url, title, data
        return url, title, None
    except Exception as e:
        print(f"    ERROR {url[:60]}: {e}")
        return url, title, None

def download_parallel(scheme_files, period):
    """Download all scheme files for a period in parallel. Returns list of (title, bytes)."""
    tasks = []
    for url, title in scheme_files:
        fname  = re.sub(r'[^\w.]', '_', url.split('/')[-1].split('?')[0])[:60]
        cpath  = CACHE / f"bandhan_{period}_{fname}"
        tasks.append((url, title, cpath))

    results = []
    with ThreadPoolExecutor(max_workers=MAX_WORKERS) as ex:
        futs = {ex.submit(_download_one, t): t for t in tasks}
        for fut in as_completed(futs):
            url, title, data = fut.result()
            if data:
                results.append((title, data))
    return results


# ── parse one scheme XLSX ─────────────────────────────────────────────────────

def parse_bandhan(data_bytes, period, title=''):
    """
    Confirmed Bandhan XLSX layout (2024-06 through 2026-04):
      Row 0: (scheme_code_idf, None, ...)
      Row 1: ('INDEX', 'Portfolio Statement as on {Month} {DD},{YYYY}', ...)
      Row 2: (None, fund_name, ...)
      Row 3: blank
      Row 4: (None, 'Name of the Instrument', 'ISIN', 'Industry/Rating',
                     'Quantity', 'Market/Fair Value', '% to NAV', ...)
      Row 5+: data — col0=idf_code, col1=stock_name, col2=ISIN, col6=pct (decimal 0-1)
    """
    import openpyxl
    as_of = (pd.to_datetime(period + '-01') + pd.offsets.MonthEnd(0)).date()

    try:
        wb = openpyxl.load_workbook(io.BytesIO(data_bytes), read_only=True, data_only=True)
    except Exception as e:
        return []

    ws = wb[wb.sheetnames[0]]
    all_rows = list(ws.iter_rows(values_only=True))
    if len(all_rows) < 5:
        return []

    # Row 0 col 0 = internal scheme code (e.g. 'IDF028')
    scheme_code_file = str(all_rows[0][0]).strip() if all_rows[0][0] else ''

    # Row 1 col 1 = "Portfolio Statement as on {Month} {DD},{YYYY}"
    for ri in range(min(3, len(all_rows))):
        for ci in range(min(4, len(all_rows[ri]))):
            cell = all_rows[ri][ci]
            if cell and isinstance(cell, str) and 'portfolio statement as on' in cell.lower():
                m = re.search(r'(\w+ \d{1,2},?\s*\d{4})', cell)
                if m:
                    try:
                        date_str = re.sub(r',\s*', ' ', m.group(1))
                        as_of = pd.to_datetime(date_str, format='%B %d %Y').date()
                    except:
                        pass
                break

    # Fund name: row AFTER the "Portfolio Statement" row, col 1
    # Layout: row0=(idf_code,None...), row1=(INDEX, "Portfolio Statement as on..."),
    #         row2=(None, fund_name), row3=blank, row4=headers
    fund_name = ''
    portfolio_row = None
    for ri, row in enumerate(all_rows[:6]):
        for cell in row:
            if cell and isinstance(cell, str) and 'portfolio statement as on' in cell.lower():
                portfolio_row = ri
                break
        if portfolio_row is not None:
            break

    if portfolio_row is not None:
        # Fund name is in the next non-empty row, column 1
        for ri in range(portfolio_row + 1, min(portfolio_row + 4, len(all_rows))):
            row = all_rows[ri]
            # Try col 1 first, then col 0
            for ci in [1, 0]:
                if len(row) > ci and row[ci]:
                    cell = str(row[ci]).strip()
                    if (len(cell) > 5
                            and 'name of the instrument' not in cell.lower()
                            and not re.match(r'^IDF\d+', cell)
                            and cell.upper() != 'INDEX'):
                        fund_name = cell
                        break
            if fund_name:
                break

    if not fund_name:
        # Fallback: extract from catalog title "Monthly and Half-Yearly - {FundName} - DD Mon YYYY"
        m = re.match(r'^(?:Monthly[^-]*-\s*)(.+?)(?:\s*-\s*\d{1,2}\s)', title)
        fund_name = m.group(1).strip() if m else title.split(' - ')[-2] if ' - ' in title else title

    # Find header row
    hdr_idx = None
    for i, row in enumerate(all_rows[:10]):
        vals = [str(v).strip() if v else '' for v in row]
        if any('ISIN' in v or 'isin' in v.lower() for v in vals):
            hdr_idx = i
            break
    if hdr_idx is None:
        return []

    hdr      = [str(v).lower().strip() if v else '' for v in all_rows[hdr_idx]]
    isin_col = next((i for i, h in enumerate(hdr) if h == 'isin' or 'isin' in h), None)
    name_col = next((i for i, h in enumerate(hdr) if 'name' in h or 'instrument' in h or 'security' in h), None)
    pct_col  = next((i for i, h in enumerate(hdr) if '%' in h and 'nav' in h), None)

    if isin_col is None or pct_col is None:
        return []

    # Detect decimal vs percentage
    sample_pcts = []
    for row in all_rows[hdr_idx+1:hdr_idx+10]:
        if row and len(row) > pct_col and row[pct_col] is not None:
            v = safe_float(row[pct_col])
            if v and v > 0:
                sample_pcts.append(v)
    pct_multiplier = 100.0 if sample_pcts and max(sample_pcts) < 2.0 else 1.0

    rows = []
    for row in all_rows[hdr_idx+1:]:
        if not row or not any(row):
            continue
        isin = str(row[isin_col]).strip() if len(row) > isin_col and row[isin_col] else ''
        if not is_valid_isin(isin):
            continue
        pct = safe_float(row[pct_col]) if len(row) > pct_col else None
        if pct is None:
            continue
        pct *= pct_multiplier
        sname = str(row[name_col]).strip() if name_col is not None and len(row) > name_col and row[name_col] else ''
        rows.append({
            'isin':        isin,
            'stock_name':  sname,
            'pct_nav':     pct,
            'scheme_name': fund_name,
            '_sheet':      scheme_code_file,
            'amc':         'Bandhan',
            'as_of_date':  as_of,
        })
    return rows


# ── scheme code mapping ───────────────────────────────────────────────────────

def build_scheme_code_map(rows, name_lk, isin_lk):
    names   = set(r['scheme_name'] for r in rows)
    mapping = {}
    for name in names:
        if not name:
            continue
        norm = re.sub(r'\s+', ' ', name.lower().strip())
        if ('Bandhan', norm) in name_lk:
            mapping[name] = name_lk[('Bandhan', norm)]
            continue
        best, best_score = None, 0.0
        for (a, n), code in name_lk.items():
            if a != 'Bandhan':
                continue
            if 'direct' not in n or ('growth' not in n and 'gr' not in n):
                continue
            name_words = set(re.findall(r'\w+', norm)) - {'fund', 'the', 'of', 'and', 'plan', 'option', 'direct', 'growth', 'regular'}
            cand_words = set(re.findall(r'\w+', n))
            if not name_words:
                continue
            score = len(name_words & cand_words) / len(name_words)
            if score > best_score:
                best_score = score
                best = code
        if best and best_score >= 0.5:
            mapping[name] = best
    found = sum(1 for v in mapping.values() if v)
    print(f"  Scheme code mapping: {found}/{len(names)} resolved")
    return mapping


# ── write to holdings parquets ────────────────────────────────────────────────

WRITE_COLS = ['isin', 'stock_name', 'pct_nav', 'scheme_name', '_sheet',
              'amc', 'scheme_code', 'as_of_date']

def append_to_holdings(rows, scheme_map):
    if not rows:
        return 0
    df = pd.DataFrame(rows)
    raw_code = df['scheme_name'].map(scheme_map)
    df['scheme_code'] = pd.to_numeric(raw_code, errors='coerce').astype('Float64')
    df['as_of_date']  = pd.to_datetime(df['as_of_date'])
    df['year']        = df['as_of_date'].dt.year
    df['month']       = df['as_of_date'].dt.month

    total = 0
    for (yr, mo), grp in df.groupby(['year', 'month']):
        period = f"{yr:04d}-{mo:02d}"
        fpath  = HOLD / f"{period}.parquet"
        out_cols = [c for c in WRITE_COLS if c in grp.columns]
        grp_out  = grp[out_cols].copy()
        if fpath.exists():
            existing = pd.read_parquet(fpath)
            existing = existing[existing['amc'] != 'Bandhan']
            combined = pd.concat([existing, grp_out], ignore_index=True)
        else:
            combined = grp_out
        combined.to_parquet(fpath, index=False)
        total += len(grp_out)
        print(f"    {period}: wrote {len(grp_out)} Bandhan rows  (file total: {len(combined)})")
    return total


# ── main ──────────────────────────────────────────────────────────────────────

def parse_period_range(from_p, to_p):
    """Generate YYYY-MM strings between from_p and to_p inclusive."""
    from datetime import date
    y0, m0 = int(from_p[:4]), int(from_p[5:7])
    y1, m1 = int(to_p[:4]),   int(to_p[5:7])
    periods = []
    y, m = y0, m0
    while (y, m) <= (y1, m1):
        periods.append(f"{y:04d}-{m:02d}")
        m += 1
        if m > 12:
            m = 1
            y += 1
    return periods

if __name__ == '__main__':
    parser = argparse.ArgumentParser()
    parser.add_argument('--period',      nargs='+', help='Specific period(s) YYYY-MM')
    parser.add_argument('--from-period', help='Start of range YYYY-MM')
    parser.add_argument('--to-period',   help='End of range YYYY-MM')
    parser.add_argument('--dry-run',     action='store_true')
    args = parser.parse_args()

    if args.period:
        periods_wanted = set(args.period)
    elif args.from_period and args.to_period:
        periods_wanted = set(parse_period_range(args.from_period, args.to_period))
    elif args.from_period:
        periods_wanted = {args.from_period}
    else:
        parser.error("Supply --period, or --from-period / --to-period")

    print(f"\n── BANDHAN FAST (parallel, {MAX_WORKERS} workers) ──")
    print(f"  Periods: {sorted(periods_wanted)}")

    # 1. AMFI lookup
    name_lk, isin_lk = build_amfi_lookup()

    # 2. Load URL catalog
    by_period = load_url_catalog(periods_wanted)
    missing   = periods_wanted - set(by_period.keys())
    if missing:
        print(f"  WARN: No catalog entries for: {sorted(missing)}")

    # 3. Process each period
    all_rows = []
    t0 = time.time()
    for period in sorted(by_period.keys()):
        scheme_files = by_period[period]
        print(f"\n  {period}: downloading {len(scheme_files)} files (parallel)...", flush=True)
        t1 = time.time()
        downloaded = download_parallel(scheme_files, period)
        print(f"    downloaded {len(downloaded)}/{len(scheme_files)} files in {time.time()-t1:.1f}s")

        period_rows = []
        for title, data in downloaded:
            rows = parse_bandhan(data, period, title)
            period_rows.extend(rows)

        # Deduplicate
        if period_rows:
            df_p = pd.DataFrame(period_rows)
            before = len(df_p)
            df_p   = df_p.drop_duplicates(subset=['isin', 'scheme_name', 'as_of_date'])
            after  = len(df_p)
            if before != after:
                print(f"    deduped {before} → {after} rows")
            period_rows = df_p.to_dict('records')

        print(f"  {period}: {len(period_rows)} holding rows", flush=True)
        all_rows.extend(period_rows)

    print(f"\n  Total: {len(all_rows)} rows across {len(by_period)} periods")
    print(f"  Total time: {time.time()-t0:.1f}s")

    if not args.dry_run and all_rows:
        sm = build_scheme_code_map(all_rows, name_lk, isin_lk)
        n  = append_to_holdings(all_rows, sm)
        print(f"\n  Bandhan rows written: {n}")
    elif args.dry_run:
        print("  [dry-run] skipping write")

    print("\nDone.")
