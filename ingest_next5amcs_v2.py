"""
ingest_next5amcs_v2.py
----------------------
Ingest 36+ months of monthly portfolio holdings for 5 new AMCs:
  Baroda BNP Paribas, HSBC, WhiteOak Capital, Sundaram, Edelweiss

Each AMC has its own URL catalog + parser.

Usage:
  python3 ingest_next5amcs_v2.py --amc all
  python3 ingest_next5amcs_v2.py --amc sundaram edelweiss
  python3 ingest_next5amcs_v2.py --amc baroda_bnp hsbc whiteoak
  python3 ingest_next5amcs_v2.py --dry-run --amc whiteoak
"""

import os, sys, re, json, io, time, warnings
from pathlib import Path
from datetime import date, datetime
import requests
import pandas as pd
import numpy as np

warnings.filterwarnings('ignore')

BASE  = Path('/sessions/admiring-nifty-dijkstra/mnt/outputs/mf_data')
HOLD  = BASE / 'holdings'
CACHE = BASE / 'raw_cache'
CACHE.mkdir(parents=True, exist_ok=True)
HOLD.mkdir(parents=True, exist_ok=True)

HEADERS = {
    'User-Agent': 'Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 Chrome/124',
    'Accept': 'text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8',
    'Accept-Language': 'en-IN,en;q=0.9',
}

# ─────────────────────────────────────────────────────────────────────────────
# URL CATALOGS
# ─────────────────────────────────────────────────────────────────────────────

BARODA_BNP_URLS = [
    ("2026-04", "https://www.barodabnpparibasmf.in/assets/download_documents/BOBBNPMF_Monthly_Portfolio_30-04-2026_18070.xls"),
    ("2026-03", "https://www.barodabnpparibasmf.in/assets/download_documents/BOBBNPMF_Monthly_Portfolio_31-03-2026_17568.xls"),
    ("2026-01", "https://www.barodabnpparibasmf.in/assets/download_documents/BOBBNPMF_Monthly_Portfolio_31-01-2026_16625.xls"),
    ("2025-12", "https://www.barodabnpparibasmf.in/assets/download_documents/BOBBNPMF_Monthly_Portfolio_31-12-2025_16137.xls"),
    ("2025-11", "https://www.barodabnpparibasmf.in/assets/download_documents/BOBBNPMF_Monthly_Portfolio_30-11-2025_15662.xls"),
    ("2025-10", "https://www.barodabnpparibasmf.in/assets/download_documents/BOBBNPMF_Monthly_Portfolio_31-10-2025_15324.xls"),
    ("2025-09", "https://www.barodabnpparibasmf.in/assets/download_documents/BOBBNPMF_Monthly_Portfolio_30-09-2025_14667.xls"),
    ("2025-08", "https://www.barodabnpparibasmf.in/assets/download_documents/BOBBNPMF_Monthly_Portfolio_31-08-2025_14178.xls"),
    ("2025-06", "https://www.barodabnpparibasmf.in/assets/download_documents/BOBBNPMF_Monthly_Portfolio_30-06-2025_13252.xls"),
    ("2025-04", "https://www.barodabnpparibasmf.in/assets/download_documents/BOBBNPMF_Monthly_Portfolio_30-04-2025_12282.xls"),
    ("2025-03", "https://www.barodabnpparibasmf.in/assets/download_documents/BOBBNPMF_Monthly_Portfolio_31-03-2025_11814.xls"),
    ("2025-02", "https://www.barodabnpparibasmf.in/assets/download_documents/BOBBNPMF_Monthly_Portfolio_28-02-2025_11393.xls"),
    ("2025-01", "https://www.barodabnpparibasmf.in/assets/download_documents/BOBBNPMF_Monthly_Portfolio_31-01-2025_10928.xls"),
    ("2024-11", "https://www.barodabnpparibasmf.in/assets/download_documents/BOBBNPMF_Monthly-Portfolio_30-11-2024_10005.xls"),
    ("2023-11", "https://www.barodabnpparibasmf.in/assets/download_documents/BOBBNPMF_Monthly_portfolio_30-11-2023_4893.xls"),
]

# Sundaram: 80 URLs scraped from AdvisorKhoj; 2 files per month, take first
# Filename format: monthlyportfolio_{DDMMYY}{HHMMSS}.xlsx
# Portfolio period = publication_month − 1 month
SUNDARAM_URLS_RAW = [
    "https://www.sundarammutual.com/uploaddir/MonthlyPortfolio/monthlyportfolio_090526110429.xlsx",
    "https://www.sundarammutual.com/uploaddir/MonthlyPortfolio/monthlyportfolio_090526110515.xlsx",
    "https://www.sundarammutual.com/uploaddir/MonthlyPortfolio/monthlyportfolio_090426191346.xlsx",
    "https://www.sundarammutual.com/uploaddir/MonthlyPortfolio/monthlyportfolio_090426194008.xlsx",
    "https://www.sundarammutual.com/uploaddir/MonthlyPortfolio/monthlyportfolio_100326101758.xlsx",
    "https://www.sundarammutual.com/uploaddir/MonthlyPortfolio/monthlyportfolio_100326102345.xlsx",
    "https://www.sundarammutual.com/uploaddir/MonthlyPortfolio/monthlyportfolio_100226114424.xlsx",
    "https://www.sundarammutual.com/uploaddir/MonthlyPortfolio/monthlyportfolio_090226181440.xlsx",
    "https://www.sundarammutual.com/uploaddir/MonthlyPortfolio/monthlyportfolio_090126162746.xlsx",
    "https://www.sundarammutual.com/uploaddir/MonthlyPortfolio/monthlyportfolio_090126163010.xlsx",
    "https://www.sundarammutual.com/uploaddir/MonthlyPortfolio/monthlyportfolio_091225181339.xlsx",
    "https://www.sundarammutual.com/uploaddir/MonthlyPortfolio/monthlyportfolio_091225181438.xlsx",
    "https://www.sundarammutual.com/uploaddir/MonthlyPortfolio/monthlyportfolio_101125121829.xlsx",
    "https://www.sundarammutual.com/uploaddir/MonthlyPortfolio/monthlyportfolio_101125121922.xlsx",
    "https://www.sundarammutual.com/uploaddir/MonthlyPortfolio/monthlyportfolio_101025105519.xlsx",
    "https://www.sundarammutual.com/uploaddir/MonthlyPortfolio/monthlyportfolio_101025105606.xlsx",
    "https://www.sundarammutual.com/uploaddir/MonthlyPortfolio/monthlyportfolio_090925172539.xlsx",
    "https://www.sundarammutual.com/uploaddir/MonthlyPortfolio/monthlyportfolio_090925172605.xlsx",
    "https://www.sundarammutual.com/uploaddir/MonthlyPortfolio/monthlyportfolio_080825185044.xlsx",
    "https://www.sundarammutual.com/uploaddir/MonthlyPortfolio/monthlyportfolio_080825185205.xlsx",
    "https://www.sundarammutual.com/uploaddir/MonthlyPortfolio/monthlyportfolio_090725181404.xlsx",
    "https://www.sundarammutual.com/uploaddir/MonthlyPortfolio/monthlyportfolio_090725181423.xlsx",
    "https://www.sundarammutual.com/uploaddir/MonthlyPortfolio/monthlyportfolio_090625181255.xlsx",
    "https://www.sundarammutual.com/uploaddir/MonthlyPortfolio/monthlyportfolio_090625181321.xlsx",
    "https://www.sundarammutual.com/uploaddir/MonthlyPortfolio/monthlyportfolio_120525155907.xlsx",
    "https://www.sundarammutual.com/uploaddir/MonthlyPortfolio/monthlyportfolio_090525170406.xlsx",
    "https://www.sundarammutual.com/uploaddir/MonthlyPortfolio/monthlyportfolio_100425172905.xlsx",
    "https://www.sundarammutual.com/uploaddir/MonthlyPortfolio/monthlyportfolio_100425172942.xlsx",
    "https://www.sundarammutual.com/uploaddir/MonthlyPortfolio/monthlyportfolio_100225103152.xlsx",
    "https://www.sundarammutual.com/uploaddir/MonthlyPortfolio/monthlyportfolio_100225103238.xlsx",
    "https://www.sundarammutual.com/uploaddir/MonthlyPortfolio/monthlyportfolio_100125092938.xlsx",
    "https://www.sundarammutual.com/uploaddir/MonthlyPortfolio/monthlyportfolio_100125093034.xlsx",
    "https://www.sundarammutual.com/uploaddir/MonthlyPortfolio/monthlyportfolio_091224182339.xlsx",
    "https://www.sundarammutual.com/uploaddir/MonthlyPortfolio/monthlyportfolio_091224182412.xlsx",
    "https://www.sundarammutual.com/uploaddir/MonthlyPortfolio/monthlyportfolio_081124194300.xlsx",
    "https://www.sundarammutual.com/uploaddir/MonthlyPortfolio/monthlyportfolio_081124194406.xlsx",
    "https://www.sundarammutual.com/uploaddir/MonthlyPortfolio/monthlyportfolio_101024085541.xlsx",
    "https://www.sundarammutual.com/uploaddir/MonthlyPortfolio/monthlyportfolio_101024085446.xlsx",
    "https://www.sundarammutual.com/uploaddir/MonthlyPortfolio/monthlyportfolio_100924101617.xlsx",
    "https://www.sundarammutual.com/uploaddir/MonthlyPortfolio/monthlyportfolio_100924101517.xlsx",
    "https://www.sundarammutual.com/uploaddir/MonthlyPortfolio/monthlyportfolio_090824183302.xlsx",
    "https://www.sundarammutual.com/uploaddir/MonthlyPortfolio/monthlyportfolio_090824183216.xlsx",
    "https://www.sundarammutual.com/uploaddir/MonthlyPortfolio/monthlyportfolio_100724223007.xlsx",
    "https://www.sundarammutual.com/uploaddir/MonthlyPortfolio/monthlyportfolio_100724222900.xlsx",
    "https://www.sundarammutual.com/uploaddir/MonthlyPortfolio/monthlyportfolio_100624185231.xlsx",
    "https://www.sundarammutual.com/uploaddir/MonthlyPortfolio/monthlyportfolio_100624185133.xlsx",
    "https://www.sundarammutual.com/uploaddir/MonthlyPortfolio/monthlyportfolio_100524152746.xlsx",
    "https://www.sundarammutual.com/uploaddir/MonthlyPortfolio/monthlyportfolio_100524153222.xlsx",
    "https://www.sundarammutual.com/uploaddir/MonthlyPortfolio/monthlyportfolio_100424105711.xlsx",
    "https://www.sundarammutual.com/uploaddir/MonthlyPortfolio/monthlyportfolio_100424094841.xlsx",
    "https://www.sundarammutual.com/uploaddir/MonthlyPortfolio/monthlyportfolio_120324123636.xlsx",
    "https://www.sundarammutual.com/uploaddir/MonthlyPortfolio/monthlyportfolio_080324100952.xlsx",
    "https://www.sundarammutual.com/uploaddir/MonthlyPortfolio/monthlyportfolio_090224100446.xlsx",
    "https://www.sundarammutual.com/uploaddir/MonthlyPortfolio/monthlyportfolio_090224100518.xlsx",
    "https://www.sundarammutual.com/uploaddir/MonthlyPortfolio/monthlyportfolio_100124092648.xlsx",
    "https://www.sundarammutual.com/uploaddir/MonthlyPortfolio/monthlyportfolio_100124092727.xlsx",
    "https://www.sundarammutual.com/uploaddir/MonthlyPortfolio/monthlyportfolio_081223174537.xlsx",
    "https://www.sundarammutual.com/uploaddir/MonthlyPortfolio/monthlyportfolio_081223174649.xlsx",
    "https://www.sundarammutual.com/uploaddir/MonthlyPortfolio/monthlyportfolio_091123145642.xlsx",
    "https://www.sundarammutual.com/uploaddir/MonthlyPortfolio/monthlyportfolio_091123145719.xlsx",
    "https://www.sundarammutual.com/uploaddir/MonthlyPortfolio/monthlyportfolio_101023105804.xlsx",
    "https://www.sundarammutual.com/uploaddir/MonthlyPortfolio/monthlyportfolio_101023105906.xlsx",
    "https://www.sundarammutual.com/uploaddir/MonthlyPortfolio/monthlyportfolio_080923150311.xlsx",
    "https://www.sundarammutual.com/uploaddir/MonthlyPortfolio/monthlyportfolio_080923150449.xlsx",
    "https://www.sundarammutual.com/uploaddir/MonthlyPortfolio/monthlyportfolio_110823120359.xlsx",
    "https://www.sundarammutual.com/uploaddir/MonthlyPortfolio/monthlyportfolio_110823120300.xlsx",
    "https://www.sundarammutual.com/uploaddir/MonthlyPortfolio/monthlyportfolio_080723160658.xlsx",
    "https://www.sundarammutual.com/uploaddir/MonthlyPortfolio/monthlyportfolio_080723160754.xlsx",
    "https://www.sundarammutual.com/uploaddir/MonthlyPortfolio/monthlyportfolio_080623180206.xlsx",
    "https://www.sundarammutual.com/uploaddir/MonthlyPortfolio/monthlyportfolio_080623180229.xlsx",
    "https://www.sundarammutual.com/uploaddir/MonthlyPortfolio/monthlyportfolio_100523114353.xlsx",
    "https://www.sundarammutual.com/uploaddir/MonthlyPortfolio/monthlyportfolio_100523114427.xlsx",
    "https://www.sundarammutual.com/uploaddir/MonthlyPortfolio/monthlyportfolio_100423102336.xlsx",
    "https://www.sundarammutual.com/uploaddir/MonthlyPortfolio/monthlyportfolio_100423102514.xlsx",
    "https://www.sundarammutual.com/uploaddir/MonthlyPortfolio/monthlyportfolio_090323181912.xlsx",
    "https://www.sundarammutual.com/uploaddir/MonthlyPortfolio/monthlyportfolio_090323182005.xlsx",
    "https://www.sundarammutual.com/uploaddir/MonthlyPortfolio/monthlyportfolio_090223151704.xlsx",
    "https://www.sundarammutual.com/uploaddir/MonthlyPortfolio/monthlyportfolio_090223151828.xlsx",
]

EDELWEISS_URLS_RAW = [
    "https://www.edelweissmf.com/Files/MONTHLY%20PORTFOLIO%20OF%20SCHEME/Monthly%20Portfolio%20and%20Risk-o-Meter/2026/Apr/published/EDEL_Portfolio%20Monthly%20Notes%2030-Apr-2026_09052026_041124_PM.xlsx",
    "https://www.edelweissmf.com/Files/MONTHLY%20PORTFOLIO%20OF%20SCHEME/Monthly%20Portfolio%20and%20Risk-o-Meter/2026/Mar/Published/EDEL_Portfolio%20Monthly%20Notes%2031-Mar-2026_10042026_104718_AM.xlsx",
    "https://www.edelweissmf.com/Files/MONTHLY%20PORTFOLIO%20OF%20SCHEME/Monthly%20Portfolio%20and%20Risk-o-Meter/2026/Feb/published/EDEL_Portfolio%20Monthly%20Notes%2028-Feb-2026_10032026_102057_AM.xlsx",
    "https://www.edelweissmf.com/Files/MONTHLY%20PORTFOLIO%20OF%20SCHEME/Monthly%20Portfolio%20and%20Risk-o-Meter/2026/Jan/Published/EDEL_Portfolio%20Monthly%20Notes%2031-Jan-2026_10022026_102622_AM.xlsx",
    "https://www.edelweissmf.com/Files/MONTHLY%20PORTFOLIO%20OF%20SCHEME/Monthly%20Portfolio%20and%20Risk-o-Meter/2025/Dec/published/EDEL_Portfolio%20Monthly%20Notes%2031-Dec-2025_09012026_050035_PM.xlsx",
    "https://www.edelweissmf.com/Files/MONTHLY%20PORTFOLIO%20OF%20SCHEME/Monthly%20Portfolio%20and%20Risk-o-Meter/2025/Dec/published/EDEL_Portfolio%20Monthly%20Notes%2030-Nov-2025_10122025_114312_AM.xlsx",
    "https://www.edelweissmf.com/Files/MONTHLY%20PORTFOLIO%20OF%20SCHEME/Monthly%20Portfolio%20and%20Risk-o-Meter/2025/Oct/published/EDEL_Portfolio%20Monthly%20Notes%2031-Oct-2025_10112025_022751_PM.xlsx",
    "https://www.edelweissmf.com/Files/MONTHLY%20PORTFOLIO%20OF%20SCHEME/Monthly%20Portfolio%20and%20Risk-o-Meter/2025/Sep/published/EDEL_Portfolio%20Monthly%20Notes%2030-Sep-2025_10102025_032334_PM.xlsx",
    "https://www.edelweissmf.com/Files/MONTHLY%20PORTFOLIO%20OF%20SCHEME/Monthly%20Portfolio%20and%20Risk-o-Meter/2025/Aug/published/EDEL_Portfolio%20Monthly%20Notes%2031-Aug-2025_10092025_035211_PM.xlsx",
    "https://www.edelweissmf.com/Files/MONTHLY%20PORTFOLIO%20OF%20SCHEME/Monthly%20Portfolio%20and%20Risk-o-Meter/2025/Jul/published/EDEL_Portfolio%20Monthly%20Notes%2031-Jul-2025_08082025_104611_AM.xlsx",
    "https://www.edelweissmf.com/Files/MONTHLY%20PORTFOLIO%20OF%20SCHEME/Monthly%20Portfolio%20and%20Risk-o-Meter/2025/Jun/published/EDEL_Portfolio%20Monthly%20Notes%2030-Jun-2025_10072025_014026_PM.xlsx",
    "https://www.edelweissmf.com/Files/MONTHLY%20PORTFOLIO%20OF%20SCHEME/Monthly%20Portfolio%20and%20Risk-o-Meter/2025/May/published/EDEL_Portfolio%20Monthly%20Notes%2031-May-2025_10062025_122454_PM.xlsx",
    "https://www.edelweissmf.com/Files/MONTHLY%20PORTFOLIO%20OF%20SCHEME/Monthly%20Portfolio%20and%20Risk-o-Meter/2025/Apr/Published/EDEL_Portfolio%20Monthly%20Notes%2030-Apr-2025_09052025_033533_PM.xlsx",
    "https://www.edelweissmf.com/Files/MONTHLY%20PORTFOLIO%20OF%20SCHEME/Monthly%20Portfolio%20and%20Risk-o-Meter/2025/Feb/published/EDEL_Portfolio%20Monthly%20Notes%2031-Mar-2025_09042025_030408_PM.xlsx",
    "https://www.edelweissmf.com/Files/MONTHLY%20PORTFOLIO%20OF%20SCHEME/Monthly%20Portfolio%20and%20Risk-o-Meter/2025/Feb/published/EDEL_Portfolio%20Monthly%20Notes%2028-Feb-2025_10032025_101524_AM.xlsx",
    "https://www.edelweissmf.com/Files/MONTHLY%20PORTFOLIO%20OF%20SCHEME/Monthly%20Portfolio%20and%20Risk-o-Meter/2025/Jan/published/EDEL_Portfolio%20Monthly%20Notes%2031-Jan-2025_10022025_104332_AM.xlsx",
    "https://www.edelweissmf.com/Files/MONTHLY%20PORTFOLIO%20OF%20SCHEME/Monthly%20Portfolio%20and%20Risk-o-Meter/published/EDEL_Portfolio%20Monthly%20Notes%2031-Dec-2024_10012025_102521_AM.xlsx",
    "https://www.edelweissmf.com/Files/MONTHLY%20PORTFOLIO%20OF%20SCHEME/Monthly%20Portfolio%20and%20Risk-o-Meter/2024/Nov/published/EDEL_Portfolio%20Monthly%20Notes%2030-Nov-2024_09122024_055141_PM.xlsx",
    "https://www.edelweissmf.com/Files/MONTHLY%20PORTFOLIO%20OF%20SCHEME/Monthly%20Portfolio%20and%20Risk-o-Meter/2024/Oct/published/EDEL_Portfolio%20Monthly%20Notes%2031-Oct-2024_08112024_104200_PM.xlsx",
    "https://www.edelweissmf.com/Files/MONTHLY%20PORTFOLIO%20OF%20SCHEME/Monthly%20Portfolio%20and%20Risk-o-Meter/2024/Sep/published/EDEL_Portfolio%20Monthly%20Notes%2030-Sep-2024_09102024_070217_PM.xlsx",
    "https://www.edelweissmf.com/Files/MONTHLY%20PORTFOLIO%20OF%20SCHEME/Monthly%20Portfolio%20and%20Risk-o-Meter/2024/Aug/published/EDEL_Portfolio%20Monthly%20Notes%2031-Aug-2024_10092024_092747_AM.xlsx",
    "https://www.edelweissmf.com/Files/MONTHLY%20PORTFOLIO%20OF%20SCHEME/Monthly%20Portfolio%20and%20Risk-o-Meter/2024/Jul/published/EDEL_Portfolio%20Monthly%20Notes%2031-Jul-2024_08082024_062139_PM.xlsx",
    "https://www.edelweissmf.com/Files/MONTHLY%20PORTFOLIO%20OF%20SCHEME/Monthly%20Portfolio%20and%20Risk-o-Meter/2024/Jun/published/EDEL_Portfolio%20Monthly%20Notes%2030-Jun-2024_10072024_113051_AM.xlsx",
    "https://www.edelweissmf.com/Files/MONTHLY%20PORTFOLIO%20OF%20SCHEME/Monthly%20Portfolio%20and%20Risk-o-Meter/2024/May/Published/EDEL_Portfolio%20Monthly%20Notes%2031-May-2024_10062024_042828_PM.xlsx",
    "https://www.edelweissmf.com/Files/MONTHLY%20PORTFOLIO%20OF%20SCHEME/Monthly%20Portfolio%20and%20Risk-o-Meter/2024/Apr/published/EDEL_Portfolio%20Monthly%20Notes%2030-Apr-2024_09052024_045717_PM.xlsx",
    "https://www.edelweissmf.com/Files/MONTHLY%20PORTFOLIO%20OF%20SCHEME/Monthly%20Portfolio%20and%20Risk-o-Meter/2024/Mar/published/EDEL_Portfolio%20Monthly%20Notes%2031-Mar-2024_09042024_101124_AM.xlsx",
    "https://www.edelweissmf.com/Files/MONTHLY%20PORTFOLIO%20OF%20SCHEME/Monthly%20Portfolio%20and%20Risk-o-Meter/2024/Feb/published/EDEL_Portfolio%20Monthly%20Notes%2029-Feb-2024_07032024_062312_PM.xlsx",
    "https://www.edelweissmf.com/Files/MONTHLY%20PORTFOLIO%20OF%20SCHEME/Monthly%20Portfolio%20and%20Risk-o-Meter/2024/Jan/published/EDEL_Portfolio%20Monthly%20Notes%2031-Jan-2024_09022024_010639_PM.xlsx",
    "https://www.edelweissmf.com/Files/MONTHLY%20PORTFOLIO%20OF%20SCHEME/Monthly%20Portfolio%20and%20Risk-o-Meter/2023/Dec/published/EDEL_Portfolio%20Monthly%20Notes%2031-Dec-2023_09012024_022314_PM.xlsx",
    "https://www.edelweissmf.com/Files/MONTHLY%20PORTFOLIO%20OF%20SCHEME/Monthly%20Portfolio%20and%20Risk-o-Meter/2023/Nov/published/EDEL_Portfolio%20Monthly%20Notes%2030-Nov-2023_07122023_045850_PM.xlsx",
    "https://www.edelweissmf.com/Files/MONTHLY%20PORTFOLIO%20OF%20SCHEME/Monthly%20Portfolio%20and%20Risk-o-Meter/2023/Oct/published/EDEL_Portfolio%20Monthly%20Notes%2031-Oct-2023_10112023_092042_AM.xlsx",
    "https://www.edelweissmf.com/Files/MONTHLY%20PORTFOLIO%20OF%20SCHEME/Monthly%20Portfolio%20and%20Risk-o-Meter/2023/Sep/published/EDEL_Portfolio%20Monthly%20Notes%2030-Sep-2023_09102023_122622_PM.xlsx",
    "https://www.edelweissmf.com/Files/MONTHLY%20PORTFOLIO%20OF%20SCHEME/Monthly%20Portfolio%20and%20Risk-o-Meter/2023/Aug/published/EDEL_Portfolio%20Monthly%20Notes%2031-Aug-2023_revised_08092023_110325_AM.xlsx",
    "https://www.edelweissmf.com/Files/MONTHLY%20PORTFOLIO%20OF%20SCHEME/Monthly%20Portfolio%20and%20Risk-o-Meter/2023/Jul/published/EDEL_Portfolio%20Monthly%20Notes%2031-Jul-2023_08082023_011733_PM.xlsx",
    "https://www.edelweissmf.com/Files/MONTHLY%20PORTFOLIO%20OF%20SCHEME/Monthly%20Portfolio%20and%20Risk-o-Meter/2023/Jun/published/EDEL_Portfolio%20Monthly%20Notes%2030-Jun-2023_10072023_105142_AM.xlsx",
    "https://www.edelweissmf.com/Files/MONTHLY%20PORTFOLIO%20OF%20SCHEME/Monthly%20Portfolio%20and%20Risk-o-Meter/2023/May/published/EDEL_Portfolio%20Monthly%20Notes%2031-May-2023_08062023_080122_PM.xlsx",
    "https://www.edelweissmf.com/Files/MONTHLY%20PORTFOLIO%20OF%20SCHEME/Monthly%20Portfolio%20and%20Risk-o-Meter/2023/Apr/published/EDEL_Portfolio%20Monthly%20Notes%2030-Apr-2023_10052023_111331_AM.xlsx",
    "https://www.edelweissmf.com/Files/MONTHLY%20PORTFOLIO%20OF%20SCHEME/Monthly%20Portfolio%20and%20Risk-o-Meter/2023/Mar/published/EDEL_Portfolio%20Monthly%20Notes%2031-Mar-2023_08042023_110130_PM.xlsx",
    "https://www.edelweissmf.com/Files/MONTHLY%20PORTFOLIO%20OF%20SCHEME/Monthly%20Portfolio%20and%20Risk-o-Meter/2023/Feb/published/EDEL_Portfolio%20Monthly%20Notes%2028-Feb-2023_10032023_024846_PM.xlsx",
    "https://www.edelweissmf.com/Files/MONTHLY%20PORTFOLIO%20OF%20SCHEME/Monthly%20Portfolio%20and%20Risk-o-Meter/2023/Jan/published/EDEL_Portfolio%20Monthly%20Notes%2031-Jan-2023_09022023_040621_PM.xlsx",
]

HSBC_SLUGS = [
    "hsbc-large-cap-fund", "hsbc-flexi-cap-fund", "hsbc-large-mid-cap-fund",
    "hsbc-small-cap-fund", "hsbc-value-fund", "hsbc-multi-cap-fund",
    "hsbc-focused-fund", "hsbc-india-export-opportunities-fund",
    "hsbc-financial-services-fund", "hsbc-corporate-bond-fund",
    "hsbc-dynamic-bond-fund", "hsbc-short-duration-fund",
    "hsbc-ultra-short-duration-fund", "hsbc-overnight-fund", "hsbc-liquid-fund",
    "hsbc-elss-tax-saver-fund", "hsbc-balanced-advantage-fund",
    "hsbc-arbitrage-fund", "hsbc-banking-and-psu-debt-fund",
    "hsbc-nifty-50-index-fund", "hsbc-nifty-next-50-index-fund",
    "hsbc-medium-to-long-duration-fund", "hsbc-credit-risk-fund",
    "hsbc-infrastructure-fund", "hsbc-midcap-fund",
    "hsbc-aggressive-hybrid-fund", "hsbc-money-market-fund", "hsbc-gilt-fund",
    "hsbc-conservative-hybrid-fund", "hsbc-low-duration-fund",
    "hsbc-brazil-fund", "hsbc-medium-duration-fund", "hsbc-consumption-fund",
    "hsbc-equity-savings-fund", "hsbc-business-cycles-fund",
    "hsbc-multi-asset-allocation-fund",
]

# ─────────────────────────────────────────────────────────────────────────────
# HELPERS
# ─────────────────────────────────────────────────────────────────────────────

def is_valid_isin(s):
    if not isinstance(s, str):
        return False
    return bool(re.match(r'^IN[A-Z0-9]{10}$', s.strip()))


def safe_float(x):
    try:
        v = float(str(x).replace(',', '').replace('%', '').strip())
        return v if -200 < v < 200 else None
    except:
        return None


def download_file(url, cache_path, amc_label=""):
    """Download url → cache_path, returning bytes. Skips if already cached."""
    if cache_path.exists() and cache_path.stat().st_size > 1000:
        return cache_path.read_bytes()

    hdrs = dict(HEADERS)
    if 'barodabnp' in url or 'baroda' in url.lower():
        hdrs['Referer'] = 'https://www.barodabnpparibasmf.in/'
    elif 'sundaram' in url:
        hdrs['Referer'] = 'https://www.sundarammutual.com/'
    elif 'edelweiss' in url:
        hdrs['Referer'] = 'https://www.edelweissmf.com/'
    elif 'hsbc' in url:
        hdrs['Referer'] = 'https://www.assetmanagement.hsbc.co.in/'
    elif 'whiteoak' in url or 'content.whiteoakamc' in url:
        hdrs['Referer'] = 'https://www.whiteoakamc.com/'

    try:
        r = requests.get(url, headers=hdrs, timeout=60)
        if r.status_code == 200:
            ct = r.headers.get('content-type', '').lower()
            data = r.content
            if len(data) < 1000:
                print(f"    WARN tiny {len(data)}b from {url[:70]}")
                return None
            if 'html' in ct and len(data) < 500000:
                print(f"    WARN HTML response ({len(data)}b) for {url[:70]}")
                return None
            cache_path.write_bytes(data)
            return data
        else:
            print(f"    HTTP {r.status_code}: {url[:80]}")
            return None
    except Exception as e:
        print(f"    ERR downloading {url[:60]}: {e}")
        return None


def load_workbook_any(data_bytes, ext='.xlsx'):
    """Load xlsx or xls from bytes. Always try openpyxl first (handles .xls renamed to xlsx)."""
    import openpyxl, xlrd
    # Always try openpyxl first — many .xls files from Indian AMCs are actually xlsx
    try:
        bio = io.BytesIO(data_bytes)
        wb = openpyxl.load_workbook(bio, read_only=True, data_only=True)
        return ('openpyxl', wb)
    except Exception:
        pass
    # Fallback to xlrd for genuine legacy .xls
    try:
        book = xlrd.open_workbook(file_contents=data_bytes)
        return ('xlrd', book)
    except Exception as e:
        print(f"    load error (tried openpyxl+xlrd): {e}")
        return None


def wb_sheetnames(wb_tuple):
    kind, wb = wb_tuple
    if kind == 'xlrd':
        return wb.sheet_names()
    return wb.sheetnames


def wb_sheet_to_rows(wb_tuple, sheet_name):
    """Return list of tuples (row values) for a sheet."""
    kind, wb = wb_tuple
    if kind == 'xlrd':
        try:
            sh = wb.sheet_by_name(sheet_name)
        except Exception:
            return []
        return [sh.row_values(i) for i in range(sh.nrows)]
    else:
        if sheet_name not in wb.sheetnames:
            return []
        ws = wb[sheet_name]
        return list(ws.iter_rows(values_only=True))


def find_header_row(rows, required=('isin',), min_cols=3):
    """Find index of first row containing all required keywords."""
    for i, r in enumerate(rows):
        vals = [str(v).lower().strip() if v is not None else '' for v in r]
        if len([v for v in vals if v]) >= min_cols:
            if all(any(req in v for v in vals) for req in required):
                return i
    return None


def parse_date_from_text(text):
    """Try to parse a date from free-text like '31 March 2025' or '31-03-2025'."""
    if not isinstance(text, str):
        return None
    patterns = [
        r'(\d{1,2})[/-](\d{1,2})[/-](\d{4})',
        r'(\d{1,2})\s+(\w+)\s+(\d{4})',
    ]
    for pat in patterns:
        m = re.search(pat, text, re.I)
        if m:
            try:
                return pd.to_datetime(m.group(), dayfirst=True).date()
            except:
                pass
    return None


def build_amfi_lookup_for_amcs(amc_map):
    """
    Fetch AMFI NAVAll.txt and build (amc, norm_name) → scheme_code lookup.
    amc_map: { 'amfi keyword string': 'AMC label', ... }
    Returns: (name_lookup, isin_lookup)
    """
    url = "https://portal.amfiindia.com/spages/NAVAll.txt"
    try:
        r = requests.get(url, headers={'User-Agent': 'Mozilla/5.0'}, timeout=45)
        r.raise_for_status()
        data = r.content.decode('utf-8', errors='replace')
    except Exception as e:
        print(f"  AMFI fetch failed: {e}")
        return {}, {}

    name_lk = {}
    isin_lk = {}
    current_amc = None

    for line in data.splitlines():
        line = line.strip()
        if not line:
            continue
        if ';' not in line:
            low = line.lower()
            current_amc = None
            for k, v in amc_map.items():
                if k in low:
                    current_amc = v
                    break
        elif current_amc:
            parts = line.split(';')
            if len(parts) >= 4:
                code = parts[0].strip()
                isin = parts[1].strip()
                name = parts[3].strip()
                norm = re.sub(r'\s+', ' ', name.lower().strip())
                name_lk[(current_amc, norm)] = code
                if isin and isin != '-':
                    isin_lk[isin] = code

    print(f"  AMFI: {len(name_lk)} name entries, {len(isin_lk)} ISIN entries")
    return name_lk, isin_lk


def build_scheme_code_map(amc, rows, name_lk, isin_lk):
    """Map scheme_name → scheme_code using AMFI lookups."""
    names = set(r['scheme_name'] for r in rows if r.get('scheme_name'))
    mapping = {}

    for name in names:
        norm = re.sub(r'\s+', ' ', name.lower().strip())
        if (amc, norm) in name_lk:
            mapping[name] = name_lk[(amc, norm)]
            continue

        # Try ISIN from rows
        for r in rows:
            if r.get('scheme_name') == name and r.get('isin') and r['isin'] in isin_lk:
                mapping[name] = isin_lk[r['isin']]
                break
        if name in mapping:
            continue

        # Fuzzy: word overlap for Direct Growth schemes
        best, best_score = None, 0
        name_words = set(re.findall(r'\w+', norm)) - {
            'fund', 'the', 'of', 'and', 'plan', 'option', 'direct', 'growth',
            'regular', 'mutual', 'paribas', 'baroda', 'bnp'
        }
        if not name_words:
            continue
        for (a, n), code in name_lk.items():
            if a != amc:
                continue
            if 'direct' not in n or 'growth' not in n:
                continue
            cand_words = set(re.findall(r'\w+', n))
            score = len(name_words & cand_words) / len(name_words)
            if score > best_score:
                best_score, best = score, code
        if best and best_score >= 0.5:
            mapping[name] = best

    found = sum(1 for v in mapping.values() if v)
    print(f"  Scheme code mapping: {found}/{len(names)} resolved for {amc}")
    return mapping


def append_to_holdings(rows, scheme_map, amc):
    """Write rows to per-month holdings parquets, replacing existing AMC rows."""
    if not rows:
        return 0

    df = pd.DataFrame(rows)
    df['scheme_code'] = pd.to_numeric(
        df['scheme_name'].map(scheme_map), errors='coerce'
    ).astype('Float64')
    df['as_of_date'] = pd.to_datetime(df['as_of_date'])
    df['year']  = df['as_of_date'].dt.year
    df['month'] = df['as_of_date'].dt.month

    WRITE_COLS = ['isin', 'stock_name', 'pct_nav', 'scheme_name', '_sheet',
                  'amc', 'scheme_code', 'as_of_date']
    total = 0

    for (yr, mo), grp in df.groupby(['year', 'month']):
        period = f"{yr:04d}-{mo:02d}"
        fpath  = HOLD / f"{period}.parquet"
        out_cols = [c for c in WRITE_COLS if c in grp.columns]
        grp_out = grp[out_cols].copy()

        if fpath.exists():
            existing = pd.read_parquet(fpath)
            existing = existing[existing['amc'] != amc]
            combined = pd.concat([existing, grp_out], ignore_index=True)
        else:
            combined = grp_out

        combined.to_parquet(fpath, index=False)
        total += len(grp_out)
        print(f"    {period}: {len(grp_out)} rows for {amc}  (file total: {len(combined)})")

    return total


# ─────────────────────────────────────────────────────────────────────────────
# BARODA BNP PARIBAS PARSER
# ─────────────────────────────────────────────────────────────────────────────

def parse_baroda_bnp(data_bytes, period, ext='.xls'):
    """
    Multi-sheet .xls file.
    Index sheet: [Sr No., Short Name, Scheme Name]
    Per-scheme sheet (sheet_id = short name):
      Row 0: [None, 'Baroda BNP Paribas <scheme name>', ...]
      Row 2: [whitespace, 'Monthly Portfolio Statement as on <date>', ...]
      Row 3: headers [None, 'Name of the Instrument', 'ISIN', 'Rating',
                      'Quantity', 'Market/Fair Value', '% to Net Assets', 'YTM']
      Row 4+: data
    % to NAV is at column index 6 (0-indexed).
    """
    wb = load_workbook_any(data_bytes, ext)
    if wb is None:
        return []

    sheets = wb_sheetnames(wb)
    default_as_of = (pd.to_datetime(period + '-01') + pd.offsets.MonthEnd(0)).date()

    # Build index: short_name → full scheme name
    idx_sheet = None
    for s in sheets:
        if s.strip().lower() in ('index', 'ind', 'sheet1'):
            idx_sheet = s
            break
    if idx_sheet is None and sheets:
        idx_sheet = sheets[0]

    index_rows = wb_sheet_to_rows(wb, idx_sheet) if idx_sheet else []
    index_map = {}  # sheet_code → full scheme name
    for r in index_rows:
        if not r or len(r) < 3:
            continue
        code = str(r[1]).strip() if r[1] else ''
        name = str(r[2]).strip() if r[2] else ''
        if code and name and code.lower() not in ('short name', 'acronym', 'scheme id'):
            index_map[code] = name

    all_rows = []
    skipped_sheets = {'Index', 'index', 'INDEX', idx_sheet}

    for sname in sheets:
        if sname in skipped_sheets:
            continue
        sheet_rows = wb_sheet_to_rows(wb, sname)
        if len(sheet_rows) < 5:
            continue

        # Scheme name: from index_map or row 0 / row 1
        full_name = index_map.get(sname, '')
        if not full_name:
            for i in range(min(4, len(sheet_rows))):
                for v in sheet_rows[i]:
                    if isinstance(v, str) and len(v.strip()) > 10 and 'baroda' in v.lower():
                        full_name = v.strip()
                        break
                if full_name:
                    break
        if not full_name:
            full_name = sname

        # As-of date: scan rows 0-4 for 'as on'
        as_of = default_as_of
        for i in range(min(5, len(sheet_rows))):
            for v in sheet_rows[i]:
                if isinstance(v, str) and 'as on' in v.lower():
                    d = parse_date_from_text(v)
                    if d:
                        as_of = d
                        break

        # Find header row (must have 'isin')
        hdr_idx = find_header_row(sheet_rows, required=('isin',))
        if hdr_idx is None:
            continue

        hdr = [str(v).lower().strip() if v is not None else '' for v in sheet_rows[hdr_idx]]
        isin_col = next((i for i, h in enumerate(hdr) if 'isin' in h), None)
        name_col = next((i for i, h in enumerate(hdr) if 'name' in h or 'instrument' in h), None)
        pct_col  = next((i for i, h in enumerate(hdr) if '%' in h and ('nav' in h or 'net' in h or 'asset' in h)), None)
        if pct_col is None:
            # Fallback: Baroda BNP confirms col 6
            pct_col = 6

        if isin_col is None:
            continue

        for r in sheet_rows[hdr_idx + 1:]:
            if not r:
                continue
            isin = str(r[isin_col]).strip() if len(r) > isin_col and r[isin_col] else ''
            if not is_valid_isin(isin):
                continue
            pct = safe_float(r[pct_col]) if len(r) > pct_col else None
            if pct is None:
                continue
            sn = str(r[name_col]).strip() if name_col is not None and len(r) > name_col and r[name_col] else ''
            all_rows.append({
                'isin':        isin,
                'stock_name':  sn,
                'pct_nav':     pct,
                'scheme_name': full_name,
                '_sheet':      sname,
                'amc':         'Baroda_BNP',
                'as_of_date':  as_of,
            })

    return all_rows


# ─────────────────────────────────────────────────────────────────────────────
# SUNDARAM PARSER
# ─────────────────────────────────────────────────────────────────────────────

def sundaram_url_to_period(url):
    """Extract portfolio period (YYYY-MM) from Sundaram filename timestamp DDMMYY."""
    m = re.search(r'monthlyportfolio_(\d{2})(\d{2})(\d{2})\d+\.xlsx', url)
    if not m:
        return None
    dd, mm, yy = int(m.group(1)), int(m.group(2)), int(m.group(3))
    year = 2000 + yy
    # Publication date is ~9th of the month; portfolio = previous month
    from datetime import date
    pub_date = date(year, mm, dd)
    # Subtract one month
    if mm == 1:
        port_month = 12; port_year = year - 1
    else:
        port_month = mm - 1; port_year = year
    return f"{port_year:04d}-{port_month:02d}"


def dedup_sundaram_urls(url_list):
    """Take one URL per portfolio period (first encountered)."""
    seen = {}
    for url in url_list:
        period = sundaram_url_to_period(url)
        if period and period not in seen:
            seen[period] = url
    # Return sorted by period
    return sorted(seen.items(), key=lambda x: x[0], reverse=True)


def parse_sundaram(data_bytes, period):
    """
    Multi-sheet XLSX.
    Index sheet: [S.NO., ACRONYM, SCHEME NAME]
    Per-scheme sheets: look for ISIN and % to NAV columns.
    """
    wb = load_workbook_any(data_bytes, '.xlsx')
    if wb is None:
        return []

    sheets = wb_sheetnames(wb)
    default_as_of = (pd.to_datetime(period + '-01') + pd.offsets.MonthEnd(0)).date()

    # Build index from 'Index' sheet
    idx_sheet = next((s for s in sheets if s.strip().lower() == 'index'), None)
    if idx_sheet is None and sheets:
        idx_sheet = sheets[0]

    index_map = {}
    if idx_sheet:
        for r in wb_sheet_to_rows(wb, idx_sheet):
            if not r or len(r) < 3:
                continue
            code = str(r[1]).strip() if r[1] else ''
            name = str(r[2]).strip() if r[2] else ''
            if code and name and not re.match(r'(?i)^(s\.?no|acronym|scheme)', code):
                index_map[code] = name

    all_rows = []
    skip = {idx_sheet} if idx_sheet else set()

    for sname in sheets:
        if sname in skip:
            continue
        sheet_rows = wb_sheet_to_rows(wb, sname)
        if len(sheet_rows) < 5:
            continue

        full_name = index_map.get(sname, '')
        if not full_name:
            # Try to find scheme name in first few rows
            for i in range(min(5, len(sheet_rows))):
                for v in sheet_rows[i]:
                    if isinstance(v, str) and len(v.strip()) > 15 and 'sundaram' in v.lower():
                        full_name = v.strip()
                        break
                if full_name:
                    break
        if not full_name:
            full_name = sname

        # As-of date
        as_of = default_as_of
        for i in range(min(6, len(sheet_rows))):
            for v in sheet_rows[i]:
                if isinstance(v, str) and ('as on' in v.lower() or 'as at' in v.lower()):
                    d = parse_date_from_text(v)
                    if d:
                        as_of = d
                        break

        hdr_idx = find_header_row(sheet_rows, required=('isin',))
        if hdr_idx is None:
            continue

        hdr = [str(v).lower().strip() if v is not None else '' for v in sheet_rows[hdr_idx]]
        isin_col = next((i for i, h in enumerate(hdr) if 'isin' in h), None)
        name_col = next((i for i, h in enumerate(hdr) if 'name' in h or 'instrument' in h or 'security' in h), None)
        pct_col  = next((i for i, h in enumerate(hdr) if '%' in h and ('nav' in h or 'net' in h or 'asset' in h)), None)
        if pct_col is None:
            pct_col = next((i for i, h in enumerate(hdr) if '% to' in h or 'to nav' in h), None)

        if isin_col is None or pct_col is None:
            continue

        for r in sheet_rows[hdr_idx + 1:]:
            if not r:
                continue
            isin = str(r[isin_col]).strip() if len(r) > isin_col and r[isin_col] else ''
            if not is_valid_isin(isin):
                continue
            pct = safe_float(r[pct_col]) if len(r) > pct_col else None
            if pct is None:
                continue
            sn = str(r[name_col]).strip() if name_col is not None and len(r) > name_col and r[name_col] else ''
            all_rows.append({
                'isin':        isin,
                'stock_name':  sn,
                'pct_nav':     pct,
                'scheme_name': full_name,
                '_sheet':      sname,
                'amc':         'Sundaram',
                'as_of_date':  as_of,
            })

    return all_rows


# ─────────────────────────────────────────────────────────────────────────────
# EDELWEISS PARSER
# ─────────────────────────────────────────────────────────────────────────────

def edelweiss_url_to_period(url):
    """Extract portfolio period from Edelweiss filename: 'DD-Mon-YYYY'."""
    fname = url.split('/')[-1]
    # e.g. 'EDEL_Portfolio Monthly Notes 30-Apr-2026_09052026_041124_PM.xlsx'
    m = re.search(r'(\d{1,2}-\w{3}-\d{4})', fname)
    if not m:
        return None
    try:
        d = pd.to_datetime(m.group(1), format='%d-%b-%Y', dayfirst=True)
        return f"{d.year:04d}-{d.month:02d}"
    except:
        return None


def parse_edelweiss(data_bytes, period):
    """
    Multi-sheet XLSX.
    Index sheet: Row 0-1 = header, Row 2: ['Fund Id', 'Fund Desc', ...]
    Per-scheme sheets: look for ISIN + % NAV columns.
    """
    wb = load_workbook_any(data_bytes, '.xlsx')
    if wb is None:
        return []

    sheets = wb_sheetnames(wb)
    default_as_of = (pd.to_datetime(period + '-01') + pd.offsets.MonthEnd(0)).date()

    # Build index from 'Index' sheet
    idx_sheet = next((s for s in sheets if s.strip().lower() == 'index'), None)
    if idx_sheet is None and sheets:
        idx_sheet = sheets[0]

    index_map = {}
    if idx_sheet:
        idx_rows = wb_sheet_to_rows(wb, idx_sheet)
        hdr_i = None
        for i, r in enumerate(idx_rows):
            vals = [str(v).lower().strip() if v else '' for v in r]
            if any('fund id' in v or 'fund desc' in v for v in vals):
                hdr_i = i
                break
        if hdr_i is not None:
            idx_hdr = [str(v).lower().strip() if v else '' for v in idx_rows[hdr_i]]
            id_col   = next((i for i, h in enumerate(idx_hdr) if 'fund id' in h or 'id' == h), None)
            desc_col = next((i for i, h in enumerate(idx_hdr) if 'fund desc' in h or 'desc' in h or 'name' in h), None)
            if id_col is not None and desc_col is not None:
                for r in idx_rows[hdr_i + 1:]:
                    if not r or len(r) <= max(id_col, desc_col):
                        continue
                    code = str(r[id_col]).strip() if r[id_col] else ''
                    name = str(r[desc_col]).strip() if r[desc_col] else ''
                    if code and name:
                        index_map[code] = name

    all_rows = []
    skip = {idx_sheet} if idx_sheet else set()

    for sname in sheets:
        if sname in skip:
            continue
        sheet_rows = wb_sheet_to_rows(wb, sname)
        if len(sheet_rows) < 5:
            continue

        full_name = index_map.get(sname, '')
        if not full_name:
            for i in range(min(5, len(sheet_rows))):
                for v in sheet_rows[i]:
                    if isinstance(v, str) and len(v.strip()) > 15 and 'edelweiss' in v.lower():
                        full_name = v.strip()
                        break
                if full_name:
                    break
        if not full_name:
            full_name = sname

        as_of = default_as_of
        for i in range(min(6, len(sheet_rows))):
            for v in sheet_rows[i]:
                if isinstance(v, str) and ('as on' in v.lower() or 'statement' in v.lower()):
                    d = parse_date_from_text(v)
                    if d:
                        as_of = d
                        break

        hdr_idx = find_header_row(sheet_rows, required=('isin',))
        if hdr_idx is None:
            continue

        hdr = [str(v).lower().strip() if v is not None else '' for v in sheet_rows[hdr_idx]]
        isin_col = next((i for i, h in enumerate(hdr) if 'isin' in h), None)
        name_col = next((i for i, h in enumerate(hdr) if 'name' in h or 'instrument' in h or 'security' in h), None)
        pct_col  = next((i for i, h in enumerate(hdr) if '%' in h and ('nav' in h or 'net' in h or 'asset' in h)), None)
        if pct_col is None:
            pct_col = next((i for i, h in enumerate(hdr) if '% to' in h), None)

        if isin_col is None or pct_col is None:
            continue

        for r in sheet_rows[hdr_idx + 1:]:
            if not r:
                continue
            isin = str(r[isin_col]).strip() if len(r) > isin_col and r[isin_col] else ''
            if not is_valid_isin(isin):
                continue
            pct = safe_float(r[pct_col]) if len(r) > pct_col else None
            if pct is None:
                continue
            sn = str(r[name_col]).strip() if name_col is not None and len(r) > name_col and r[name_col] else ''
            all_rows.append({
                'isin':        isin,
                'stock_name':  sn,
                'pct_nav':     pct,
                'scheme_name': full_name,
                '_sheet':      sname,
                'amc':         'Edelweiss',
                'as_of_date':  as_of,
            })

    return all_rows


# ─────────────────────────────────────────────────────────────────────────────
# WHITEOAK PARSER (per-scheme xlsx from Strapi catalog)
# ─────────────────────────────────────────────────────────────────────────────

def load_whiteoak_catalog():
    path = Path('/sessions/admiring-nifty-dijkstra/mnt/outputs/whiteoak_catalog.json')
    if not path.exists():
        print("  ERROR: whiteoak_catalog.json not found")
        return {}
    with open(path) as f:
        return json.load(f)


def whiteoak_equity_filter(scheme_name):
    """Keep only equity/hybrid schemes, exclude pure debt/liquid."""
    n = name_l = scheme_name.lower()
    exclude = ['liquid', 'overnight', 'ultra short', 'low duration', 'short duration',
               'medium duration', 'dynamic bond', 'corporate bond', 'credit risk',
               'gilt', 'money market', 'conservative hybrid', 'arbitrage',
               'banking and psu', 'debt', 'bond', 'fixed income']
    return not any(e in name_l for e in exclude)


def parse_whiteoak_perfile(data_bytes, scheme_name, period):
    """
    Single-scheme per-scheme xlsx from WhiteOak.
    Look for ISIN column and % NAV column.
    """
    try:
        import openpyxl
        wb = openpyxl.load_workbook(io.BytesIO(data_bytes), read_only=True, data_only=True)
    except Exception as e:
        print(f"    openpyxl error: {e}")
        return []

    default_as_of = (pd.to_datetime(period + '-01') + pd.offsets.MonthEnd(0)).date()
    all_rows = []

    for wsname in wb.sheetnames:
        ws = wb[wsname]
        sheet_rows = list(ws.iter_rows(values_only=True))
        if len(sheet_rows) < 5:
            continue

        as_of = default_as_of
        for i in range(min(8, len(sheet_rows))):
            for v in sheet_rows[i]:
                if isinstance(v, str) and ('as on' in v.lower() or 'as at' in v.lower()):
                    d = parse_date_from_text(v)
                    if d:
                        as_of = d
                        break

        hdr_idx = find_header_row(sheet_rows, required=('isin',))
        if hdr_idx is None:
            continue

        hdr = [str(v).lower().strip() if v is not None else '' for v in sheet_rows[hdr_idx]]
        isin_col = next((i for i, h in enumerate(hdr) if 'isin' in h), None)
        name_col = next((i for i, h in enumerate(hdr) if 'name' in h or 'instrument' in h or 'security' in h), None)
        pct_col  = next((i for i, h in enumerate(hdr) if '%' in h and ('nav' in h or 'net' in h or 'asset' in h)), None)
        if pct_col is None:
            pct_col = next((i for i, h in enumerate(hdr) if '% to' in h or 'to nav' in h), None)

        if isin_col is None or pct_col is None:
            continue

        for r in sheet_rows[hdr_idx + 1:]:
            if not r:
                continue
            isin = str(r[isin_col]).strip() if len(r) > isin_col and r[isin_col] else ''
            if not is_valid_isin(isin):
                continue
            pct = safe_float(r[pct_col]) if len(r) > pct_col else None
            if pct is None:
                continue
            sn = str(r[name_col]).strip() if name_col is not None and len(r) > name_col and r[name_col] else ''
            all_rows.append({
                'isin':        isin,
                'stock_name':  sn,
                'pct_nav':     pct,
                'scheme_name': scheme_name,
                '_sheet':      wsname,
                'amc':         'WhiteOak',
                'as_of_date':  as_of,
            })
        break  # Only parse first sheet per file

    return all_rows


# ─────────────────────────────────────────────────────────────────────────────
# HSBC PARSER (per-scheme xlsx)
# ─────────────────────────────────────────────────────────────────────────────

def generate_hsbc_urls(from_period='2023-01'):
    """
    Generate (period, slug, url) tuples for all HSBC slugs × months.
    URL: /-/media/files/attachments/india/mutual-funds/portfolios/
         document-{DDMMYYYY}/{slug}-{DD-mon-YYYY}.xlsx
    """
    import calendar
    from datetime import date

    MONTH_ABB = {1:'jan',2:'feb',3:'mar',4:'apr',5:'may',6:'jun',
                 7:'jul',8:'aug',9:'sep',10:'oct',11:'nov',12:'dec'}
    BASE_URL = "https://www.assetmanagement.hsbc.co.in/-/media/files/attachments/india/mutual-funds/portfolios"

    fp = pd.to_datetime(from_period + '-01')
    today = pd.Timestamp.now()
    cur = fp
    combos = []

    while cur <= today:
        yr, mo = cur.year, cur.month
        last_day = calendar.monthrange(yr, mo)[1]
        dd = f"{last_day:02d}"
        mm = f"{mo:02d}"
        yyyy = str(yr)
        mon = MONTH_ABB[mo]
        folder = f"document-{dd}{mm}{yyyy}"
        for slug in HSBC_SLUGS:
            fname = f"{slug}-{dd}-{mon}-{yyyy}.xlsx"
            url = f"{BASE_URL}/{folder}/{fname}"
            combos.append((f"{yr:04d}-{mo:02d}", slug, url))
        cur += pd.DateOffset(months=1)

    return combos


def parse_hsbc_perfile(data_bytes, scheme_slug, period):
    """
    Single-scheme HSBC xlsx. Find ISIN + % to NAV columns.
    """
    try:
        import openpyxl
        wb = openpyxl.load_workbook(io.BytesIO(data_bytes), read_only=True, data_only=True)
    except Exception as e:
        print(f"    openpyxl error: {e}")
        return []

    default_as_of = (pd.to_datetime(period + '-01') + pd.offsets.MonthEnd(0)).date()
    all_rows = []

    # Derive scheme name from slug
    scheme_name = ' '.join(w.capitalize() for w in scheme_slug.split('-'))

    for wsname in wb.sheetnames:
        ws = wb[wsname]
        sheet_rows = list(ws.iter_rows(values_only=True))
        if len(sheet_rows) < 5:
            continue

        # Extract actual scheme name from first few rows
        actual_name = scheme_name
        as_of = default_as_of
        for i in range(min(8, len(sheet_rows))):
            for v in sheet_rows[i]:
                if isinstance(v, str):
                    if 'hsbc' in v.lower() and len(v.strip()) > 10:
                        # Strip parenthetical description "(An open ended...)"
                        clean = v.strip().split('(')[0].strip().rstrip('.')
                        actual_name = clean if clean else v.strip()
                    if 'as on' in v.lower() or 'as at' in v.lower():
                        d = parse_date_from_text(v)
                        if d:
                            as_of = d

        hdr_idx = find_header_row(sheet_rows, required=('isin',))
        if hdr_idx is None:
            continue

        hdr = [str(v).lower().strip() if v is not None else '' for v in sheet_rows[hdr_idx]]
        isin_col = next((i for i, h in enumerate(hdr) if 'isin' in h), None)
        name_col = next((i for i, h in enumerate(hdr) if 'name' in h or 'instrument' in h or 'security' in h), None)
        # HSBC header uses "Percentage to Net Assets" (word, not '%' symbol)
        pct_col  = next((i for i, h in enumerate(hdr) if
                         ('%' in h or 'percent' in h) and
                         ('nav' in h or 'net' in h or 'asset' in h)), None)
        if pct_col is None:
            pct_col = next((i for i, h in enumerate(hdr) if '% to' in h or 'percent' in h), None)

        if isin_col is None or pct_col is None:
            continue

        sheet_data_rows = []
        for r in sheet_rows[hdr_idx + 1:]:
            if not r:
                continue
            isin = str(r[isin_col]).strip() if len(r) > isin_col and r[isin_col] else ''
            if not is_valid_isin(isin):
                continue
            pct = safe_float(r[pct_col]) if len(r) > pct_col else None
            if pct is None:
                continue
            sn = str(r[name_col]).strip() if name_col is not None and len(r) > name_col and r[name_col] else ''
            sheet_data_rows.append((isin, sn, pct))

        # HSBC stores pct as decimal fraction (0.0974 = 9.74%); convert to percentage
        if sheet_data_rows and all(abs(p) < 2 for _, _, p in sheet_data_rows):
            sheet_data_rows = [(i, s, p * 100) for i, s, p in sheet_data_rows]

        for isin, sn, pct in sheet_data_rows:
            all_rows.append({
                'isin':        isin,
                'stock_name':  sn,
                'pct_nav':     pct,
                'scheme_name': actual_name,
                '_sheet':      wsname,
                'amc':         'HSBC',
                'as_of_date':  as_of,
            })
        break  # First sheet only

    return all_rows


# ─────────────────────────────────────────────────────────────────────────────
# AMC INGEST RUNNERS
# ─────────────────────────────────────────────────────────────────────────────

def ingest_baroda_bnp(name_lk, isin_lk, dry_run=False):
    print("\n── BARODA BNP PARIBAS ──")
    all_rows = []
    for period, url in BARODA_BNP_URLS:
        ext = '.xls' if url.endswith('.xls') else '.xlsx'
        fname = re.sub(r'[^\w.]', '_', url.split('/')[-1])[:80]
        cpath = CACHE / f"baroda_{period}_{fname}"
        data = download_file(url, cpath, 'Baroda_BNP')
        if data is None:
            print(f"  {period}: SKIP")
            continue
        rows = parse_baroda_bnp(data, period, ext)
        print(f"  {period}: {len(rows)} rows from {len(set(r['_sheet'] for r in rows))} sheets")
        all_rows.extend(rows)
        time.sleep(0.4)

    if not dry_run and all_rows:
        sm = build_scheme_code_map('Baroda_BNP', all_rows, name_lk, isin_lk)
        n = append_to_holdings(all_rows, sm, 'Baroda_BNP')
        print(f"  Baroda_BNP total: {n} rows written")
    return all_rows


def ingest_sundaram(name_lk, isin_lk, dry_run=False, chunk=0):
    print("\n── SUNDARAM ──")
    url_list = dedup_sundaram_urls(SUNDARAM_URLS_RAW)
    if chunk > 0:
        url_list = url_list[:chunk]
    print(f"  {len(url_list)} unique periods: {[p for p, _ in url_list[:5]]} ...")
    all_rows = []
    for period, url in url_list:
        fname = url.split('/')[-1]
        cpath = CACHE / f"sundaram_{period}_{fname}"
        data = download_file(url, cpath, 'Sundaram')
        if data is None:
            print(f"  {period}: SKIP")
            continue
        rows = parse_sundaram(data, period)
        n_sheets = len(set(r['_sheet'] for r in rows))
        print(f"  {period}: {len(rows)} rows from {n_sheets} sheets")
        all_rows.extend(rows)
        time.sleep(0.3)

    if not dry_run and all_rows:
        sm = build_scheme_code_map('Sundaram', all_rows, name_lk, isin_lk)
        n = append_to_holdings(all_rows, sm, 'Sundaram')
        print(f"  Sundaram total: {n} rows written")
    return all_rows


def ingest_edelweiss(name_lk, isin_lk, dry_run=False, chunk=0):
    print("\n── EDELWEISS ──")
    # Deduplicate by period
    seen = {}
    for url in EDELWEISS_URLS_RAW:
        period = edelweiss_url_to_period(url)
        if period and period not in seen:
            seen[period] = url
    url_list = sorted(seen.items(), key=lambda x: x[0], reverse=True)
    if chunk > 0:
        url_list = url_list[:chunk]
    print(f"  {len(url_list)} unique periods: {[p for p, _ in url_list[:5]]} ...")

    # Edelweiss: simple user-agent only (Referer triggers 403, no Referer works)
    edel_headers = {'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36'}

    all_rows = []
    for period, url in url_list:
        fname = re.sub(r'[^\w.]', '_', url.split('/')[-1])[:80]
        cpath = CACHE / f"edelweiss_{period}_{fname}"

        if cpath.exists() and cpath.stat().st_size > 1000:
            data = cpath.read_bytes()
        else:
            try:
                r = requests.get(url, headers=edel_headers, timeout=60)
                if r.status_code == 200 and len(r.content) > 1000:
                    data = r.content
                    cpath.write_bytes(data)
                else:
                    print(f"  {period}: SKIP (HTTP {r.status_code}, {len(r.content)}b)")
                    time.sleep(0.3)
                    continue
            except Exception as e:
                print(f"  {period}: SKIP ({e})")
                continue

        rows = parse_edelweiss(data, period)
        n_sheets = len(set(r['_sheet'] for r in rows))
        print(f"  {period}: {len(rows)} rows from {n_sheets} sheets")
        all_rows.extend(rows)
        time.sleep(0.3)

    if not dry_run and all_rows:
        sm = build_scheme_code_map('Edelweiss', all_rows, name_lk, isin_lk)
        n = append_to_holdings(all_rows, sm, 'Edelweiss')
        print(f"  Edelweiss total: {n} rows written")
    return all_rows


def ingest_whiteoak(name_lk, isin_lk, dry_run=False, chunk=0, from_period='2021-01'):
    print("\n── WHITEOAK ──")
    catalog = load_whiteoak_catalog()
    if not catalog:
        return []

    # Filter equity schemes from catalog
    # Key format: "scheme_name|YYYY-MM"
    equity_entries = {}
    for key, val in catalog.items():
        parts = key.split('|', 1)
        if len(parts) != 2:
            continue
        scheme_name, cat_period = parts
        if not whiteoak_equity_filter(scheme_name):
            continue
        if cat_period < from_period:
            continue
        equity_entries[key] = val

    print(f"  Equity entries in catalog: {len(equity_entries)}")
    periods = sorted(set(k.split('|')[1] for k in equity_entries))
    if chunk > 0:
        periods = periods[:chunk]
    print(f"  Periods to process ({len(periods)}): {periods[:5]} ...")

    all_rows = []
    # Process chronologically, per period
    for period in periods:
        period_rows = []
        for key, val in equity_entries.items():
            if not key.endswith(f'|{period}'):
                continue
            scheme_name = key.split('|')[0]
            url = val['url']
            ext = val.get('ext', '.xlsx')
            safe_scheme = re.sub(r'[^\w]', '_', scheme_name)[:40]
            fname = f"whiteoak_{period}_{safe_scheme}{ext}"
            cpath = CACHE / fname
            data = download_file(url, cpath, 'WhiteOak')
            if data is None:
                print(f"  {period} {scheme_name[:40]}: SKIP")
                continue
            rows = parse_whiteoak_perfile(data, scheme_name, period)
            period_rows.extend(rows)
            time.sleep(0.2)

        if period_rows:
            print(f"  {period}: {len(period_rows)} rows from {len(set(r['scheme_name'] for r in period_rows))} schemes")
        all_rows.extend(period_rows)

    if not dry_run and all_rows:
        sm = build_scheme_code_map('WhiteOak', all_rows, name_lk, isin_lk)
        n = append_to_holdings(all_rows, sm, 'WhiteOak')
        print(f"  WhiteOak total: {n} rows written")
    return all_rows


def ingest_hsbc(name_lk, isin_lk, dry_run=False, from_period='2023-01', chunk=0):
    print("\n── HSBC ──")
    combos = generate_hsbc_urls(from_period)

    # Chunk: limit to first N months
    if chunk > 0:
        all_periods = sorted(set(p for p, _, _ in combos))
        keep_periods = set(all_periods[:chunk])
        combos = [(p, s, u) for p, s, u in combos if p in keep_periods]
    print(f"  {len(combos)} slug×month combos to probe")

    all_rows = []
    period_hits = {}
    failed_slugs = set()

    for period, slug, url in combos:
        if slug in failed_slugs:
            continue
        fname = f"hsbc_{period}_{slug}.xlsx"
        cpath = CACHE / fname

        # Skip if already cached successfully
        if cpath.exists() and cpath.stat().st_size > 1000:
            data = cpath.read_bytes()
        else:
            data = download_file(url, cpath, 'HSBC')
            if data is None:
                continue
            time.sleep(0.15)

        rows = parse_hsbc_perfile(data, slug, period)
        if rows:
            if period not in period_hits:
                period_hits[period] = 0
            period_hits[period] += len(rows)
            all_rows.extend(rows)

    # Print summary per period
    for p in sorted(period_hits.keys()):
        print(f"  {p}: {period_hits[p]} rows")

    if not dry_run and all_rows:
        sm = build_scheme_code_map('HSBC', all_rows, name_lk, isin_lk)
        n = append_to_holdings(all_rows, sm, 'HSBC')
        print(f"  HSBC total: {n} rows written")
    return all_rows


# ─────────────────────────────────────────────────────────────────────────────
# ENTRY POINT
# ─────────────────────────────────────────────────────────────────────────────

if __name__ == '__main__':
    import argparse
    parser = argparse.ArgumentParser()
    parser.add_argument('--amc', nargs='+',
                        choices=['baroda_bnp', 'sundaram', 'edelweiss', 'whiteoak', 'hsbc', 'all'],
                        default=['all'])
    parser.add_argument('--dry-run', action='store_true')
    parser.add_argument('--from-period', default='2023-01')
    parser.add_argument('--chunk', type=int, default=0, help='0=all, N=first N periods only')
    args = parser.parse_args()

    do_all = 'all' in args.amc

    AMC_MAP = {
        'baroda bnp paribas mutual fund': 'Baroda_BNP',
        'hsbc mutual fund': 'HSBC',
        'whiteoak capital mutual fund': 'WhiteOak',
        'sundaram mutual fund': 'Sundaram',
        'edelweiss mutual fund': 'Edelweiss',
    }

    print("Fetching AMFI lookups...")
    name_lk, isin_lk = build_amfi_lookup_for_amcs(AMC_MAP)

    if 'baroda_bnp' in args.amc or do_all:
        ingest_baroda_bnp(name_lk, isin_lk, dry_run=args.dry_run)

    if 'sundaram' in args.amc or do_all:
        ingest_sundaram(name_lk, isin_lk, dry_run=args.dry_run, chunk=args.chunk)

    if 'edelweiss' in args.amc or do_all:
        ingest_edelweiss(name_lk, isin_lk, dry_run=args.dry_run, chunk=args.chunk)

    if 'whiteoak' in args.amc or do_all:
        ingest_whiteoak(name_lk, isin_lk, dry_run=args.dry_run,
                        chunk=args.chunk, from_period=args.from_period or '2021-01')

    if 'hsbc' in args.amc or do_all:
        ingest_hsbc(name_lk, isin_lk, dry_run=args.dry_run,
                    from_period=args.from_period or '2023-01', chunk=args.chunk)

    print("\n✓ Done.")
