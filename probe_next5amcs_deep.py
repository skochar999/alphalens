"""
probe_next5amcs_deep.py
-----------------------
Deeper probe of AMC portfolio disclosure pages — looking specifically for
SEBI-mandated monthly portfolio statements (not factsheets).
Also checks HSBC file format, and tries direct portfolio disclosure URLs.
"""
import requests, re, json
from concurrent.futures import ThreadPoolExecutor, as_completed
from pathlib import Path
import io

HEADERS = {
    'User-Agent': 'Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36',
    'Accept': 'text/html,application/xhtml+xml,*/*;q=0.8',
}

def fetch(url, referer=''):
    hdrs = dict(HEADERS)
    if referer:
        hdrs['Referer'] = referer
    try:
        r = requests.get(url, headers=hdrs, timeout=20)
        return r.status_code, r.text, r.content
    except Exception as e:
        return 0, str(e), b''

def find_portfolio_links(html, base_domain=''):
    """Find links that look like portfolio/holdings/sebi disclosure files."""
    # Look for xlsx/xls/zip that suggest portfolio (not factsheet/KIM/application)
    all_links = re.findall(
        r'(?:href|src)[=:\s]+["\']?(https?://[^\s"\'<>]+)["\']?',
        html, re.I
    )
    all_links += re.findall(r'https?://[^\s"\'<>]+', html, re.I)

    portfolio_kws = ['portfolio', 'sebi', 'holdings', 'disclosure', 'monthlyport',
                     'monthly-port', 'monthly_port', 'mp_', '_mp_', 'ports-']
    factsheet_kws = ['factsheet', 'fact_sheet', 'fact-sheet', 'kim', 'application',
                     'reckoner', 'addenda', 'notice', 'form', 'brochure']

    results = []
    for lnk in all_links:
        low = lnk.lower()
        if not any(ext in low for ext in ['.xlsx', '.xls', '.zip', '.pdf']):
            continue
        if any(k in low for k in factsheet_kws):
            continue
        if any(k in low for k in portfolio_kws):
            results.append(lnk)
    return list(dict.fromkeys(results))[:10]

# ── 1. Baroda BNP Paribas ──────────────────────────────────────────────────
print("\n" + "="*65)
print("BARODA BNP PARIBAS")
print("="*65)

# Check their actual SEBI disclosures page
baroda_urls = [
    'https://www.barodabnpparibasmf.in/investor-services/sebi-disclosures',
    'https://www.barodabnpparibasmf.in/investor-services/monthly-portfolio-disclosure',
    'https://www.barodabnpparibasmf.in/monthly-portfolio-statement',
    'https://www.barodabnpparibasmf.in/investor-services/portfolio',
]
for url in baroda_urls:
    code, html, _ = fetch(url)
    links = find_portfolio_links(html)
    print(f"  [{code}] {url[-60:]}")
    for l in links[:4]: print(f"    {l[:85]}")

# Try API endpoint pattern
# Baroda BNP uses React/Next.js — check if there's a data API
baroda_api = 'https://www.barodabnpparibasmf.in/_next/data/'
code, html, _ = fetch('https://www.barodabnpparibasmf.in/investor-services/portfolio-disclosure')
print(f"\n  [portfolio-disclosure page] {code}")
# Look for ANY xlsx/zip links
xlsx_links = re.findall(r'https?://[^\s"\'<>]+\.(?:xlsx|xls|zip)', html, re.I)
print(f"  xlsx/zip links found: {len(xlsx_links)}")
for l in xlsx_links[:5]: print(f"    {l[:85]}")
# Look for PDF links with portfolio keywords
pdf_port = [l for l in re.findall(r'https?://[^\s"\'<>]+\.pdf', html, re.I)
            if any(k in l.lower() for k in ['portfolio', 'sebi', 'disclosure', 'monthly'])]
for l in pdf_port[:3]: print(f"    PDF: {l[:85]}")


# ── 2. HSBC ────────────────────────────────────────────────────────────────
print("\n" + "="*65)
print("HSBC")
print("="*65)

# Sample one of the HSBC UUID files to check format
hsbc_sample = 'https://www.assetmanagement.hsbc.co.in/assets/documents/mutual-funds/en/0ad32e13-42ac-4832-86f4-e0cd18fe8fa7.xlsx'
code, _, content = fetch(hsbc_sample, referer='https://www.assetmanagement.hsbc.co.in/')
print(f"  Sample HSBC file: HTTP {code}, size={len(content)} bytes")
if code == 200 and len(content) > 1000:
    # Try to read as Excel
    try:
        import openpyxl
        wb = openpyxl.load_workbook(io.BytesIO(content), read_only=True, data_only=True)
        ws = wb.active
        rows = list(ws.iter_rows(values_only=True))
        print(f"  Sheets: {wb.sheetnames}")
        print(f"  Rows: {len(rows)}")
        print(f"  First 5 rows:")
        for r in rows[:5]:
            print(f"    {[str(v)[:25] if v else None for v in r[:6]]}")
    except Exception as e:
        print(f"  openpyxl error: {e}")
        # Maybe it's a ZIP?
        if content[:2] == b'PK':
            import zipfile
            try:
                zf = zipfile.ZipFile(io.BytesIO(content))
                print(f"  ZIP contents: {zf.namelist()[:10]}")
            except: pass

# Get more HSBC links from AdvisorKhoj
code, html, _ = fetch('https://www.advisorkhoj.com/form-download-centre/Mutual/HSBC-Mutual-Fund/Monthly-Portfolio-Disclosures')
hsbc_all = re.findall(r'https?://[^\s"\'<>]+\.xlsx', html, re.I)
print(f"\n  All HSBC xlsx links from AdvisorKhoj: {len(hsbc_all)}")
for l in hsbc_all[:5]: print(f"    {l[:90]}")


# ── 3. Invesco ────────────────────────────────────────────────────────────
print("\n" + "="*65)
print("INVESCO")
print("="*65)

# Invesco direct website
invesco_urls = [
    'https://www.invescomutualfund.com/portfolio-disclosure',
    'https://www.invescomutualfund.com/monthly-portfolio',
    'https://invescomutualfund.com/en-in/investor-education/portfolio-disclosure',
]
for url in invesco_urls:
    code, html, _ = fetch(url)
    xlsx_links = re.findall(r'https?://[^\s"\'<>]+\.(?:xlsx|xls|zip)', html, re.I)
    print(f"  [{code}] {url[-60:]} — {len(xlsx_links)} xlsx/zip links")
    for l in xlsx_links[:3]: print(f"    {l[:85]}")

# Check their docs hosting pattern
# Invesco AdvisorKhoj showed: invescomutualfund.com/docs/default-source/
code, html, _ = fetch('https://www.advisorkhoj.com/form-download-centre/Mutual/Invesco-Mutual-Fund/Monthly-Portfolio-Disclosures')
inv_all = re.findall(r'https?://[^\s"\'<>]+\.(?:xlsx|xls|zip)', html, re.I)
inv_pdfs_port = [l for l in re.findall(r'https?://[^\s"\'<>]+\.pdf', html, re.I)
                  if any(k in l.lower() for k in ['portfolio', 'sebi', 'disclosure', 'monthly-port'])]
print(f"\n  Invesco xlsx/zip from AdvisorKhoj: {inv_all}")
print(f"  Invesco portfolio PDFs: {inv_pdfs_port[:3]}")

# Check if there's an Invesco API
inv_api = 'https://www.invescomutualfund.com/api/documents'
code, html, _ = fetch(inv_api)
print(f"  Invesco docs API: {code}")


# ── 4. Canara Robeco ─────────────────────────────────────────────────────
print("\n" + "="*65)
print("CANARA ROBECO")
print("="*65)

# Their wp-content structure already found — check actual portfolio files
canara_urls = [
    'https://www.canararobeco.com/portfolio-disclosure',
    'https://www.canararobeco.com/investor-service/portfolio-disclosure',
    'https://www.canararobeco.com/investor-service/monthly-portfolio',
]
for url in canara_urls:
    code, html, _ = fetch(url)
    xlsx_links = re.findall(r'https?://[^\s"\'<>]+\.(?:xlsx|xls|zip)', html, re.I)
    port_pdfs  = [l for l in re.findall(r'https?://[^\s"\'<>]+\.pdf', html, re.I)
                   if any(k in l.lower() for k in ['portfolio', 'sebi', 'monthly', 'disclosure'])]
    print(f"  [{code}] {url[-55:]} — {len(xlsx_links)} xlsx, {len(port_pdfs)} port PDFs")
    for l in xlsx_links[:3]: print(f"    {l[:85]}")
    for l in port_pdfs[:3]: print(f"    PDF: {l[:85]}")


# ── 5. WhiteOak ───────────────────────────────────────────────────────────
print("\n" + "="*65)
print("WHITEOAK CAPITAL")
print("="*65)

whiteoak_urls = [
    'https://www.whiteoakmf.com/portfolio-disclosure',
    'https://www.whiteoakmf.com/monthly-portfolio-disclosure',
    'https://www.whiteoakmf.com/sebi-disclosure',
    'https://whiteoakamc.com/monthly-portfolio',
]
for url in whiteoak_urls:
    code, html, _ = fetch(url)
    xlsx_links = re.findall(r'https?://[^\s"\'<>]+\.(?:xlsx|xls|zip)', html, re.I)
    port_pdfs  = [l for l in re.findall(r'https?://[^\s"\'<>]+\.pdf', html, re.I)
                   if any(k in l.lower() for k in ['portfolio', 'sebi', 'monthly', 'disclosure'])]
    print(f"  [{code}] {url[-55:]} — {len(xlsx_links)} xlsx, {len(port_pdfs)} port PDFs")
    for l in xlsx_links[:3]: print(f"    {l[:85]}")

# Check WO content CDN for pattern
code, html, _ = fetch('https://www.advisorkhoj.com/form-download-centre/Mutual/WhiteOak-Capital-Mutual-Fund/Monthly-Portfolio-Disclosures')
wo_xlsx = re.findall(r'https?://[^\s"\'<>]+\.(?:xlsx|xls|zip)', html, re.I)
wo_pdfs = [l for l in re.findall(r'https?://[^\s"\'<>]+\.pdf', html, re.I)
            if any(k in l.lower() for k in ['portfolio', 'monthly', 'disclosure', 'sebi'])]
print(f"\n  WO xlsx/zip from AdvisorKhoj: {wo_xlsx[:3]}")
print(f"  WO portfolio PDFs: {wo_pdfs[:5]}")

print("\n\nDone.")
