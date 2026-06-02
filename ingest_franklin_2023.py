#!/usr/bin/env python3
"""
ingest_franklin_2023.py
=======================
Backfill Franklin Templeton 2023 monthly holdings (Jan–Dec 2023)
directly from AdvisorKhoj links into the holdings parquet cache.

Franklin serves a single multi-sheet XLSX per month:
  Row 1:  Scheme full name
  Row 4:  Headers — ISIN | Name | Industry | Qty | Mkt Value | % to NAV | YTM
  Row 5+: Section headers + equity/debt rows

We keep only equity rows: ISIN starts with 'IN' and has 12 chars,
and pct_nav (col 5) is a float in [0.01, 100].

Usage:
    python3 ingest_franklin_2023.py --mf-data ./mf_data
    python3 ingest_franklin_2023.py --mf-data ./mf_data --force
    python3 ingest_franklin_2023.py --mf-data ./mf_data --month 2023-06
"""
from __future__ import annotations

import argparse
import io
import logging
import subprocess
import sys
from pathlib import Path

import openpyxl
import pandas as pd

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s  %(levelname)-8s  %(message)s",
    datefmt="%H:%M:%S",
)
log = logging.getLogger("franklin2023")

UA = (
    "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) "
    "AppleWebKit/537.36 (KHTML, like Gecko) Chrome/124.0.0.0 Safari/537.36"
)

# ── AdvisorKhoj URLs for Franklin 2023 ──────────────────────────────────────
# Scraped from https://www.advisorkhoj.com/form-download-centre/Mutual/
#              Franklin-Templeton-Mutual-Fund/Monthly-Portfolio-Disclosures
# Verified: all return HTTP 200 + valid XLSX (~2 MB each)
FRANKLIN_2023_URLS: dict[str, str] = {
    "2023-01": "https://www.franklintempletonindia.com/download/en-in/monthly-portfolio-dsclr/474214f2-7061-456f-bd01-84c22569a4a1/Monthly-Portfolio-ISIN-31-Jan-2023.xlsx",
    "2023-02": "https://www.franklintempletonindia.com/download/en-in/monthly-portfolio-dsclr/e25bbd90-cdf9-4f3b-a544-4030d7827200/Monthly-Portfolio-ISIN-28-Feb-2023.xlsx",
    "2023-03": "https://www.franklintempletonindia.com/download/en-in/monthly-portfolio-dsclr/f419d7e3-dd4f-4bbc-80f7-86ca6292c93e/Monthly-Portfolio-ISIN-31-Mar-2023.xlsx",
    "2023-04": "https://www.franklintempletonindia.com/download/en-in/monthly-portfolio-dsclr/889ba2b0-b0c5-47a4-8076-a798ffc3658d/Monthly-Portfolio-ISIN-28-Apr-2023.xlsx",
    "2023-05": "https://www.franklintempletonindia.com/download/en-in/monthly-portfolio-dsclr/119e0e25-c3a4-43a9-ba3a-178a65bb0c44/Monthly-Portfolio-ISIN-31-May-2023.xlsx",
    "2023-06": "https://www.franklintempletonindia.com/download/en-in/monthly-portfolio-dsclr/d6dbbfa8-7a32-4067-a1c8-f9c98a28de1d/Monthly-Portfolio-ISIN-30-Jun-2023.xlsx",
    "2023-07": "https://www.franklintempletonindia.com/download/en-in/monthly-portfolio-dsclr/284f75f9-6ef9-40e1-921e-f670eafc1e4d/Monthly-Portfolio-ISIN-31-Jul-2023.xlsx",
    "2023-08": "https://www.franklintempletonindia.com/download/en-in/monthly-portfolio-dsclr/f28b1a1b-583f-4c4b-a2df-ed5a8fe2d94f/Monthly-Portfolio-ISIN-31-Aug-2023.xlsx",
    "2023-09": "https://www.franklintempletonindia.com/download/en-in/monthly-portfolio-dsclr/c50fd1e3-aabc-45fd-ab64-54e2bf45346f/Monthly-Portfolio-ISIN-30-Sep-2023.xlsx",
    "2023-10": "https://www.franklintempletonindia.com/download/en-in/monthly-portfolio-dsclr/c8fd7954-d7e4-4280-a5d6-8a78d1137a6e/Monthly-Portfolio-ISIN-31-Oct-2023.xlsx",
    "2023-11": "https://www.franklintempletonindia.com/download/en-in/monthly-portfolio-dsclr/bb3a4227-ef4d-48a0-a322-4a5ce12d4f5d/Monthly-Portfolio-ISIN-30-Nov-2023.xlsx",
    "2023-12": "https://www.franklintempletonindia.com/download/en-in/monthly-portfolio-dsclr/88389532-9d16-4bcd-aaa6-29b50224b6d1/Monthly-Portfolio-ISIN-29-Dec-2023.xlsx",
}

COLS = ["isin", "stock_name", "pct_nav", "scheme_name", "_sheet", "amc", "scheme_code", "as_of_date"]


# ── Download ─────────────────────────────────────────────────────────────────

def _download(url: str) -> bytes | None:
    cmd = [
        "curl", "-s", "-L", "--max-time", "60",
        "-A", UA,
        "-H", "Accept-Language: en-IN,en;q=0.9",
        "--compressed",
        url,
    ]
    r = subprocess.run(cmd, capture_output=True, timeout=65)
    if r.returncode != 0 or len(r.stdout) < 1000:
        return None
    return r.stdout


# ── Parse one Franklin XLSX ───────────────────────────────────────────────────

def _parse_franklin_xlsx(data: bytes, month_str: str) -> pd.DataFrame:
    """
    Parse a Franklin monthly portfolio XLSX.

    Sheet layout:
      Row 1: full scheme name
      Row 2: blank
      Row 3: "Portfolio Statement as on …"
      Row 4: column headers
      Row 5+: section headers or data rows

    Data row has:
      col 0: ISIN  (12-char, starts with 'IN')
      col 1: Name of Instrument
      col 2: Industry / Rating
      col 3: Quantity
      col 4: Market Value
      col 5: % to Net Assets   ← this is pct_nav (already percentage)

    We keep only equity rows: ISIN starts with 'INE'/'INF' and pct_nav valid.
    """
    try:
        wb = openpyxl.load_workbook(io.BytesIO(data), data_only=True, read_only=True)
    except Exception as e:
        log.warning(f"  openpyxl failed: {e}")
        return pd.DataFrame()

    records = []
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            continue

        # Row 1 = scheme name
        scheme_name = str(rows[0][0]).strip() if rows[0] and rows[0][0] else sheet_name

        # Clean up known footnote markers
        for marker in [" ^", " *", " #", "(", "formerly known as"]:
            if marker.lower() in scheme_name.lower():
                scheme_name = scheme_name.split(marker)[0].strip()

        # Only keep equity-focused schemes (skip debt/liquid/gilt)
        scheme_lower = scheme_name.lower()
        skip_keywords = ["liquid", "gilt", "overnight", "money market", "ultra short",
                         "low duration", "short term", "credit risk", "corporate bond",
                         "floating rate", "dynamic bond", "banking and psu",
                         "conservative", "arbitrage"]
        if any(kw in scheme_lower for kw in skip_keywords):
            continue

        # Parse data rows (skip header rows 0–3)
        for row in rows[4:]:
            if not row or row[0] is None:
                continue
            isin = str(row[0]).strip()
            # Equity ISINs: 12-char starting with INE, INF, etc.
            if len(isin) != 12 or not isin.startswith("IN"):
                continue
            # Skip debt ISINs (bonds typically have 2-digit series numbers)
            # Equity ISINs end in digits with no rating-style patterns
            try:
                pct_raw = row[5]
                pct_nav = float(pct_raw)
            except (TypeError, ValueError, IndexError):
                continue
            # pct_nav sanity: Franklin reports as actual % (e.g. 7.59 means 7.59%)
            if not (0.001 < pct_nav < 50):
                continue

            stock_name = str(row[1]).strip() if row[1] else ""
            if not stock_name or stock_name.lower().startswith("sub total"):
                continue

            records.append({
                "isin": isin,
                "stock_name": stock_name,
                "pct_nav": round(pct_nav, 4),
                "scheme_name": scheme_name,
                "_sheet": sheet_name,
                "amc": "Franklin",
                "scheme_code": pd.NA,
                "as_of_date": month_str,
            })

    return pd.DataFrame(records) if records else pd.DataFrame()


# ── Scheme code matching ─────────────────────────────────────────────────────

def _attach_codes(df: pd.DataFrame, scheme_df: pd.DataFrame) -> pd.DataFrame:
    """Fuzzy-match scheme names → scheme_codes using rapidfuzz."""
    try:
        from rapidfuzz import process, fuzz
    except ImportError:
        log.warning("rapidfuzz not installed — scheme codes will be NA")
        return df

    known_names = scheme_df["scheme_name"].tolist()
    code_map = dict(zip(scheme_df["scheme_name"], scheme_df["scheme_code"]))

    results = []
    for sname in df["scheme_name"].unique():
        match, score, _ = process.extractOne(
            sname, known_names, scorer=fuzz.token_sort_ratio
        )
        if score >= 75:
            results.append((sname, code_map.get(match, pd.NA)))
        else:
            results.append((sname, pd.NA))

    name_to_code = dict(results)
    df["scheme_code"] = df["scheme_name"].map(name_to_code)
    return df


# ── Merge into cache ──────────────────────────────────────────────────────────

def _merge_month(month_str: str, new_df: pd.DataFrame, hold_dir: Path) -> None:
    out_path = hold_dir / f"{month_str}.parquet"
    if out_path.exists():
        existing = pd.read_parquet(out_path)
        existing = existing[existing["amc"] != "Franklin"]  # remove old Franklin rows
    else:
        existing = pd.DataFrame(columns=COLS)

    for c in COLS:
        if c not in existing.columns:
            existing[c] = pd.NA
        if c not in new_df.columns:
            new_df[c] = pd.NA

    combined = pd.concat([existing[COLS], new_df[COLS]], ignore_index=True)
    combined.to_parquet(out_path, index=False)
    n_matched = combined["scheme_code"].notna().sum()
    log.info(f"  {month_str}: saved {out_path.name}  "
             f"{len(combined):,} total rows  "
             f"{n_matched/len(combined)*100:.0f}% matched")


# ── Main ──────────────────────────────────────────────────────────────────────

def main():
    p = argparse.ArgumentParser()
    p.add_argument("--mf-data", default="./mf_data")
    p.add_argument("--force", action="store_true",
                   help="Re-download even if Franklin already present in cache")
    p.add_argument("--month", default=None,
                   help="Process only this month (e.g. 2023-06)")
    args = p.parse_args()

    mf_data = Path(args.mf_data)
    hold_dir = mf_data / "holdings"
    hold_dir.mkdir(parents=True, exist_ok=True)

    scheme_path = mf_data / "scheme_list.parquet"
    scheme_df = pd.read_parquet(scheme_path) if scheme_path.exists() else None

    months = list(FRANKLIN_2023_URLS.keys())
    if args.month:
        months = [m for m in months if m == args.month]

    total_ok = 0
    for month_str in sorted(months):
        out_path = hold_dir / f"{month_str}.parquet"

        # Skip if already ingested (unless --force)
        if not args.force and out_path.exists():
            existing = pd.read_parquet(out_path)
            if "Franklin" in existing.get("amc", pd.Series([])).values:
                existing_franklin = existing[existing["amc"] == "Franklin"]
                log.info(f"  {month_str}: Franklin already present "
                         f"({len(existing_franklin)} rows) — skipping")
                continue

        url = FRANKLIN_2023_URLS[month_str]
        log.info(f"\n{month_str}: downloading …")
        data = _download(url)
        if not data:
            log.warning(f"  {month_str}: download failed ({url})")
            continue

        if data[:4] != b'PK\x03\x04':
            log.warning(f"  {month_str}: unexpected magic bytes — not an XLSX")
            continue

        df = _parse_franklin_xlsx(data, month_str)
        if df.empty:
            log.warning(f"  {month_str}: no equity records parsed")
            continue

        log.info(f"  {month_str}: {len(df)} equity rows, "
                 f"{df['scheme_name'].nunique()} schemes")

        if scheme_df is not None:
            df = _attach_codes(df, scheme_df)
            df["scheme_code"] = pd.to_numeric(
                df["scheme_code"], errors="coerce"
            ).astype("Int64")

        _merge_month(month_str, df, hold_dir)
        total_ok += 1

    log.info(f"\nDone: {total_ok}/{len(months)} months ingested")

    # Summary
    if total_ok > 0:
        import os
        files = sorted(f for f in os.listdir(hold_dir) if f.endswith(".parquet"))
        all_rows = []
        for f in files:
            df_f = pd.read_parquet(hold_dir / f)
            month = f.replace(".parquet", "")
            if "amc" in df_f.columns:
                for amc in df_f["amc"].dropna().unique():
                    all_rows.append({"month": month, "amc": amc})
        cov = pd.DataFrame(all_rows)
        print("\n=== Holdings coverage ===")
        print(cov.groupby("amc")["month"].agg(["min", "max", "count"])
              .sort_values("count", ascending=False).to_string())


if __name__ == "__main__":
    main()
