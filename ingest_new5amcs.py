"""
ingest_new5amcs.py
------------------
Download and parse 36-40 months of monthly portfolio disclosures for 5 new AMCs:
  Quant, UTI, Bandhan, Tata, Motilal Oswal

URLs sourced from AdvisorKhoj's download center (static SSR catalog).
All files are stored in a local raw_cache to avoid re-downloads.

Output: appends rows to mf_data/holdings/YYYY-MM.parquet
Schema: isin, stock_name, pct_nav, scheme_name, _sheet, amc, scheme_code, as_of_date, year, month
"""

import os, sys, re, json, time, zipfile, io, glob, warnings
from pathlib import Path
from datetime import datetime, date
import requests
import pandas as pd
import numpy as np

warnings.filterwarnings('ignore')

BASE   = Path('/sessions/admiring-nifty-dijkstra/mnt/outputs/mf_data')
HOLD   = BASE / 'holdings'
CACHE  = BASE / 'raw_cache'
CACHE.mkdir(parents=True, exist_ok=True)

HEADERS = {
    'User-Agent': 'Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36',
    'Accept': 'text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8',
    'Accept-Language': 'en-IN,en;q=0.9',
}

# ─────────────────────────────────────────────────────────────────────────────
# URL CATALOGS  (from AdvisorKhoj SSR pages, Jan 2023 – Apr 2026)
# ─────────────────────────────────────────────────────────────────────────────

QUANT_URLS = [
    ("2026-04", "https://quantmutual.com/Admin/disclouser/Monthly_Portfolio_30042026.xlsx"),
    ("2026-03", "https://quantmutual.com/Admin/disclouser/Monthly_Portfolio_March2026.xlsx"),
    ("2026-02", "https://quantmutual.com/Admin/disclouser/Monthly_Portfolio_Feb26.xlsx"),
    ("2026-01", "https://www.quantmutual.com/Admin/disclouser/quant_Mutual_Fund_Monthly_Portfolio_Jan2026.xlsx"),
    ("2025-12", "https://www.quantmutual.com/Admin/disclouser/Monthly_Portfolio_Dec25_31122025.xlsx"),
    ("2025-11", "https://www.quantmutual.com/Admin/disclouser/quant_MF_Monthly_Portfolio_Nov_2025.xlsx"),
    ("2025-10", "https://quantmutual.com/Admin/disclouser/Monthly_Portfolio_oct25_31102025.xlsx"),
    ("2025-09", "https://quantmutual.com/Admin/disclouser/Monthly_Portfolio_sept-2025.xlsx"),
    ("2025-08", "https://quantmutual.com/Admin/disclouser/Monthly_Portfolio_Aug_2025.xlsx"),
    ("2025-07", "https://quantmutual.com/Admin/disclouser/portfolio_July.xlsx"),
    ("2025-06", "https://quantmutual.com/Admin/disclouser/Monthly_30062025.xlsx"),
    ("2025-05", "https://quantmutual.com/Admin/disclouser/Monthly_Portfolio_May_2025.xlsx"),
    ("2025-04", "https://quantmutual.com/Admin/disclouser/Monthly_Portfolio_April_2025.xlsx"),
    ("2025-03", "https://quantmutual.com/Admin/disclouser/Monthly-Portfolio_March25.xlsx"),
    ("2025-02", "https://quantmutual.com/Admin/disclouser/PORTFOLIO_Feb_28022025.xlsx"),
    ("2025-01", "https://quantmutual.com/Admin/disclouser/Monthly_Portfolio_31012025.xlsx"),
    ("2024-12", "https://quantmutual.com/Admin/disclouser/portfolio-dec_31122024.xlsx"),
    ("2024-11", "https://quantmutual.com/Admin/disclouser/Portfolio_Novx.xlsx"),
    ("2024-10", "https://quantmutual.com/Admin/disclouser/Monthly-Portfolio_Oct.xlsx"),
    ("2024-09", "https://quantmutual.com/Admin/disclouser/Monthly_Portfolio_Sept.xlsx"),
    ("2024-08", "https://quantmutual.com/Admin/disclouser/Monthly_Portfolio_AUG.xlsx"),
    ("2024-07", "https://quantmutual.com/Admin/disclouser/Monthly_Portfolio_July24.xlsx"),
    ("2024-06", "https://quantmutual.com/Admin/disclouser/Portfolio_June24.xlsx"),
    ("2024-05", "https://quantmutual.com/Admin/disclouser/Monthly_Portfolio_May_2024.xlsx"),
    ("2024-04", "https://quantmutual.com/Admin/disclouser/Monthly_Portfolio_Apr_24.xlsx"),
    ("2024-03", "https://quantmutual.com/Admin/disclouser/Monthly_Portfolio_Mar_24.xlsx"),
    ("2024-02", "https://quantmutual.com/Admin/disclouser/Portfolio_Monthly_Feb_2024.xlsx"),
    ("2024-01", "https://quantmutual.com/Admin/disclouser/Monthly_Portfolio_January_2024.xlsx"),
    ("2023-12", "https://quantmutual.com/Admin/disclouser/quant_Monthly_Portfolio_Dec23.xlsx"),
    ("2023-11", "https://quantmutual.com/Admin/disclouser/Monthly_Portfolio_Nov_2023.xlsx"),
    ("2023-10", "https://quantmutual.com/Admin/disclouser/Monthly_Portfolio_Oct_2023.xlsx"),
    ("2023-09", "https://quantmutual.com/Admin/disclouser/Monthly_Portfolio_Sept_2023.xlsx"),
    ("2023-08", "https://quantmutual.com/Admin/disclouser/Monthly%20Portfolio_August_2023.xlsx"),
    ("2023-07", "https://quantmutual.com/Admin/disclouser/Monthly_Portfolio_July_2023.xlsx"),
    ("2023-06", "https://quantmutual.com/Admin/disclouser/Monthly_Portfolio_June_2023.xlsx"),
    ("2023-05", "https://quantmutual.com/Admin/disclouser/Monthly_Portfolio_May_2023.xlsx"),
    ("2023-04", "https://quantmutual.com/Admin/disclouser/Portfolio_Monthly_April_2023.xlsx"),
    ("2023-03", "https://quantmutual.com/Admin/disclouser/Portfolio_Monthly_mARCH_2023.xlsx"),
    ("2023-02", "https://quantmutual.com/Admin/disclouser/Portfolio_Monthly_Feb_2023.xlsx"),
    ("2023-01", "https://quantmutual.com/Admin/disclouser/Monthly_Portfolio_January_2023.xlsx"),
]

UTI_URLS = [
    ("2026-04", "https://d3ce1o48hc5oli.cloudfront.net/s3fs-public/2026-05/uti_mf_scheme_portfolios_30.04.2026.zip?VersionId=gCevjGjcZstHQr_XwNN3YuMVO8W7fXyW"),
    ("2026-03", "https://d3ce1o48hc5oli.cloudfront.net/s3fs-public/2026-04/uti_mf_scheme_portfolios_31.03.2026_0.zip?VersionId=.6d_JxxXNQnLPXQfZvacOBhNOriLwd3u"),
    ("2026-02", "https://d3ce1o48hc5oli.cloudfront.net/static/generic-zip/March-26/FW_%20UTI_MF_Scheme_portfolios-28.02.2026.zip"),
    ("2026-01", "https://d3ce1o48hc5oli.cloudfront.net/s3fs-public/2026-02/fw_uti_mf_scheme_portfolios_31.01.2026_1.zip?VersionId=xkJ3oDaGr6oBLeRnpY_d.ZaVg5vLenyU"),
    ("2025-12", "https://d3ce1o48hc5oli.cloudfront.net/s3fs-public/2026-01/fw_uti_mf_scheme_portfolios_31.12.2025_1.zip?VersionId=SyZ8fB4ffNzsvZOMp.cIq6qUxlSiE_Lz"),
    ("2025-11", "https://d3ce1o48hc5oli.cloudfront.net/s3fs-public/2025-12/fw_uti_mf_portfolios_30.11.2025_1.zip?VersionId=3taQvdY9d1LM1.0Fxfv9ijJW7Bear2pa"),
    ("2025-10", "https://d3ce1o48hc5oli.cloudfront.net/s3fs-public/2025-11/fw_uti_mf_portfolios_31.10.2025_0.zip?VersionId=hfieXngwtE.oEMOSVpDYw2rzeWoQBrcP"),
    ("2025-09", "https://d3ce1o48hc5oli.cloudfront.net/s3fs-public/2025-10/fw_uti_mf_scheme_portfoliios_as_of_30.09.2025_1.zip?VersionId=x.7niwpOdkZHApvnB6BB7Ud2pDAN0mNP"),
    ("2025-08", "https://d3ce1o48hc5oli.cloudfront.net/s3fs-public/2025-09/fw_uti_mf_scheme_portfolios_31.08.2025_1.zip?VersionId=9akMDZsEnDTpfv8t6K05d0dgCeKoBOL6"),
    ("2025-07", "https://d3ce1o48hc5oli.cloudfront.net/s3fs-public/2025-08/fw_uti_mf_scheme_portfolios_31.07.2025_1.zip?VersionId=sGQ0XvBvGQDKh7R7n4jyp7tqETsy7NHg"),
    ("2025-06", "https://d3ce1o48hc5oli.cloudfront.net/s3fs-public/2025-07/fw_uti_mf_scheme_portfolios_as_of_30.06.2025_1.zip?VersionId=tToatHRmpL_5865KhbtqO1.lPbNYJFJG"),
    ("2025-05", "https://d3ce1o48hc5oli.cloudfront.net/s3fs-public/2025-06/fw_uti_mf_scheme_portfolios_31.05.2025_1_0.zip?VersionId=RU7397LvDoOTzj8KeWLq94MrtzwbENmc"),
    ("2025-04", "https://d3ce1o48hc5oli.cloudfront.net/s3fs-public/2025-05/fw_uti_mf_scheme_portfolios_30.04.205.zip?VersionId=VAqg3IeLWoDSAtH7E5X2W6eOT6gbu19P"),
    ("2025-03", "https://d3ce1o48hc5oli.cloudfront.net/s3fs-public/2025-04/fw_uti_mf_scheme_portfolios_31.03.2025_0.zip?VersionId=Ui7u8hA41X1AF3wq2hWixJP8TDpCzWW2"),
    ("2025-02", "https://d3ce1o48hc5oli.cloudfront.net/s3fs-public/2025-03/FW_%20UTI%20MF%20SCHEME%20PORTFOLIOS%20-%2028.02.2025.zip?VersionId=JbFR1a6v36_2z0v21VyNNk3NT9fPm8pg"),
    ("2025-01", "https://d3ce1o48hc5oli.cloudfront.net/s3fs-public/2025-02/Fw_%20UTI%20MF%20PORTFOLIOS-%2031.01.2025.zip?VersionId=zdMWBb4UnOKSLlWulJCcKjUV6EdFYAZE"),
    ("2024-12", "https://d3ce1o48hc5oli.cloudfront.net/s3fs-public/2025-01/Fw_%20UTI%20MF%20SCHEME%20PORTFOLIOS%20-%2031.12.2024.zip?VersionId=7QxqDSvJLMPcC1y0h5.XWO2naAEeHhan"),
    ("2024-11", "https://d3ce1o48hc5oli.cloudfront.net/s3fs-public/2024-12/RE_%20UTI%20MF%20Scheme%20portfolios%20as%20of%2030.11.2024.zip?VersionId=GzDengR7Ry9463yjRuzQcqUKIA7WZfZL"),
    ("2024-10", "https://d3ce1o48hc5oli.cloudfront.net/s3fs-public/2024-11/FW_%20UTI%20MF%20scheme%20portfolios%20as%20of%2031.10.2024.zip?VersionId=9xHb_cqn7NoeS3cTw4EK5dYk8BwhRlow"),
    ("2024-09", "https://d3ce1o48hc5oli.cloudfront.net/s3fs-public/2024-10/Fw_%20UTI%20MF%20SCHEME%20PORTFOLIOS%20-%2030.09.2024.zip?VersionId=pXfNprqE2FdgIKxUe4Csy3qsqVnVHWiW"),
    ("2024-08", "https://d3ce1o48hc5oli.cloudfront.net/s3fs-public/2024-09/FW_%20UTI%20MF%20SCHEME%20PORTFOLIOS%20-%2031.08.2024.zip?VersionId=ZWUm79DlpYz_iVcuiL9Svzpi5EY9DZMd"),
    ("2024-07", "https://d3ce1o48hc5oli.cloudfront.net/s3fs-public/2024-07/Fw_%20UTIMF%20SCHEME%20PORTFOLIOS%20-%2030.06.2024.zip?VersionId=uICVzgd8ssyjDp9N0FwN5JMxOVbVXYoY"),
    ("2024-06", "https://d3ce1o48hc5oli.cloudfront.net/s3fs-public/2024-07/Fw_%20UTIMF%20SCHEME%20PORTFOLIOS%20-%2030.06.2024.zip?VersionId=uICVzgd8ssyjDp9N0FwN5JMxOVbVXYoY"),
    ("2024-05", "https://d3ce1o48hc5oli.cloudfront.net/s3fs-public/2024-06/UTI%20MF%20SCHEME%20PORTFOLIOS%20-%2031.05.2024.zip?VersionId=KRRDlpK.xdZYUyQKaAP_7LRK3iLYPqZM"),
    ("2024-04", "https://d3ce1o48hc5oli.cloudfront.net/s3fs-public/2024-05/UTI%20MF%20SCHEME%20PORTFOLIOS%20-%2030.04.2024_1.zip?VersionId=KZGdjv3EV.BEQ3c327xlKowXcuzKLz1l"),
    ("2024-03", "https://d3ce1o48hc5oli.cloudfront.net/s3fs-public/2024-04/UTI%20MF%20Scheme%20portfolios%20-%2031.03.2024.zip?VersionId=w_zp7qExyhqxSwYMIFJ36Q5Gh8ciolMS"),
    ("2024-02", "https://d3ce1o48hc5oli.cloudfront.net/s3fs-public/2024-03/UTI%20MF%20SCHEMES%20PORTFOLIOS%20AS%20OF%2029.02.2024.zip?VersionId=CBTARvBy2ds5HS2.zdtKu2Ci2p4PGPQA"),
    ("2024-01", "https://d3ce1o48hc5oli.cloudfront.net/s3fs-public/2024-02/Consolidated%20Portfolio%20January%20-%2020240131.zip?VersionId=24LoC_lgH25IGNJVWAdizb7ughh1Nfbi"),
    ("2023-12", "https://d3ce1o48hc5oli.cloudfront.net/s3fs-public/2024-01/UTI%20MF%20SCHEME%20PORTFOLIOS%20-%2031.12.2023.zip?VersionId=O6smRoHshh3L27oenmhoxvihTr16B.Xf"),
    ("2023-11", "https://d3ce1o48hc5oli.cloudfront.net/s3fs-public/2023-12/UTI%20MF%20SCHEME%20PORTFOLIOS%20-%2030.11.2023.zip?VersionId=YZSK2YArm6IgMJC4DKKcf6D8yWMP7xvH"),
    ("2023-10", "https://d3ce1o48hc5oli.cloudfront.net/s3fs-public/2023-11/Consolidated%20Portfolio%20October%202023.zip?VersionId=okRAMmXYZBAwOMMQcwDXd0s5E338au5x"),
    ("2023-09", "https://doc.utimf.com/uticontainer/Consolidated%20Portfolio%20September%20202320231009-102406.zip"),
    ("2023-08", "https://doc.utimf.com/uticontainer/Consolidated%20Portfolio%20August%20202320230908-112320.zip"),
    ("2023-07", "https://doc.utimf.com/uticontainer/Consolidated%20Portfolio%20July%20202320230808-135045.zip"),
    ("2023-06", "https://doc.utimf.com/uticontainer/Consolidated%20Portfolio%20June%20202320230710-055943.zip"),
    ("2023-05", "https://doc.utimf.com/uticontainer/Consolidated%20Portfolio%20May%20202320230609-060340.zip"),
    ("2023-04", "https://doc.utimf.com/uticontainer/Consolidated%20Portfolio%20April%20202320230509-095714.zip"),
    ("2023-03", "https://doc.utimf.com/uticontainer/Consolidated%20Portfolio%20March%20202320230409-151841.zip"),
    ("2023-02", "https://doc.utimf.com/uticontainer/Consolidated%20Portfolio%20February%20202320230309-072644.zip"),
    ("2023-01", "https://doc.utimf.com/uticontainer/Consolidated%20Portfolio%20January%20202320230208-103320.zip"),
]

TATA_URLS = [
    ("2026-04", "https://betacms.tatamutualfund.com/system/files/2026-05/Monthly%20Portfolio%20as%20on%2030th%20April%202026%20%281%29.xlsx"),
    ("2026-03", "https://betacms.tatamutualfund.com/system/files/2026-04/Monthly%20Portfolio%20as%20on%2031st%20March%202026.xlsx"),
    ("2026-02", "https://betacms.tatamutualfund.com/system/files/2026-03/Monthly%20Portfolio%20as%20on%2028th%20February%202026.xlsx"),
    ("2026-01", "https://betacms.tatamutualfund.com/system/files/2026-02/Monthly%20Portfolio%20as%20on%2031st%20January%202026.xlsx"),
    ("2025-12", "https://betacms.tatamutualfund.com/system/files/2026-01/Monthly%20Portfolio%20as%20on%2031st%20December%202025.xlsx"),
    ("2025-11", "https://betacms.tatamutualfund.com/system/files/2025-12/Monthly%20Portfolio%20as%20on%2030th%20November%202025.xlsx"),
    ("2025-10", "https://betacms.tatamutualfund.com/system/files/2025-11/Monthly%20Portfolio%20as%2031st%20October%202025.xlsx"),
    ("2025-09", "https://betacms.tatamutualfund.com/system/files/2025-10/Monthly%20Portfolio%20as%20on%2030th%20September%202025.xlsx"),
    ("2025-08", "https://betacms.tatamutualfund.com/system/files/2025-09/Monthly%20Portfolio%20as%20on%2031st%20August%202025.xlsx"),
    ("2025-07", "https://betacms.tatamutualfund.com/system/files/2025-08/Monthly%20Portfolio%20as%20on%2031st%20July%202025.xlsx"),
    ("2025-06", "https://betacms.tatamutualfund.com/system/files/2025-07/Monthly%20Portfolio%20as%20on%2030th%20June%202025.xlsx"),
    ("2025-05", "https://betacms.tatamutualfund.com/system/files/2025-06/Monthly%20Portfolio%20as%20on%2031st%20May%202025.xls"),
    ("2025-04", "https://betacms.tatamutualfund.com/system/files/2025-05/Monthly%20Portfolio%20as%20on%2030th%20April%202025.xlsx"),
    ("2025-03", "https://betacms.tatamutualfund.com/system/files/2025-04/Monthly%20Portfolio%20as%20on%2031st%20March%202025.xlsx"),
    ("2025-02", "https://betacms.tatamutualfund.com/system/files/2025-03/Monthly%20Portfolio%20as%20on%2028th%20February%202025.xls"),
    ("2025-01", "https://betacms.tatamutualfund.com/system/files/2025-02/Monthly%20Portfolio%20as%20on%2031st%20January%202025.xls"),
    ("2024-12", "https://betacms.tatamutualfund.com/system/files/2025-01/Monthly%20Portfolio%20as%20on%2031st%20December%202024.xls"),
    ("2024-11", "https://betacms.tatamutualfund.com/system/files/2024-12/Monthly%20Portfolio%20as%20on%2030th%20November%202024%20%282%29.xlsx"),
    ("2024-10", "https://betacms.tatamutualfund.com/system/files/2024-11/Monthly%20Portfolio%20as%20on%2031st%20October%202024.xlsx"),
    ("2024-09", "https://betacms.tatamutualfund.com/system/files/2024-10/Monthly%20Portfolio%20as%20on%2030th%20September%202024_0.xlsx"),
    ("2024-08", "https://betacms.tatamutualfund.com/system/files/2024-09/Monthly%20Portfolio%20as%20on%2030th%20August%202024.xlsx"),
    ("2024-07", "https://betacms.tatamutualfund.com/system/files/2024-08/Monthly%20Portfolio%20as%20on%2031st%20July%202024.xlsx"),
    ("2024-06", "https://betacms.tatamutualfund.com/system/files/2024-07/Monthly%20Portfolio%20as%20on%2030th%20June%202024.xlsx"),
    ("2024-05", "https://betacms.tatamutualfund.com/system/files/2024-06/Monthly%20Portfolio%20as%20on%2031st%20May%202024.xlsx"),
    ("2024-04", "https://betacms.tatamutualfund.com/system/files/2024-05/Monthly%20Portfolio%20as%20on%2030th%20April%202024.xlsx"),
    ("2024-03", "https://betacms.tatamutualfund.com/system/files/2024-04/Monthly%20Portfolio%20as%20on%2031st%20March%202024.xlsx"),
    ("2024-02", "https://betacms.tatamutualfund.com/system/files/2024-03/Monthly%20Portfolio%20as%20on%2029th%20February%202024%20%281%29.xlsx"),
    ("2024-01", "https://betacms.tatamutualfund.com/system/files/2024-02/Monthly%20Portfolio%20as%20on%2031st%20January%202024_1.xlsx"),
    ("2023-12", "https://betacms.tatamutualfund.com/system/files/2024-01/Monthly%20Portfolio%20as%20on%2031st%20December%202023_REVISED.xlsx"),
    ("2023-11", "https://betacms.tatamutualfund.com/system/files/2023-12/Monthly%20Portfolio%20as%20on%2030th%20November%202023.xlsx"),
    ("2023-10", "https://betacms.tatamutualfund.com/system/files/2023-11/Monthly%20Portfolio%20as%20on%2031st%20October%202023.xlsx"),
    ("2023-09", "https://betacms.tatamutualfund.com/system/files/2023-11/monthly-portfolio-as-on-30th-september-2023.xlsx"),
    ("2023-08", "https://betacms.tatamutualfund.com/system/files/2023-09/Portfolio%20as%20on%2031st%20August%2C%202023.xls"),
    ("2023-07", "https://betacms.tatamutualfund.com/system/files/2023-09/Portfolio%20as%20on%2031st%20July%2C%202023.xls"),
    ("2023-06", "https://betacms.tatamutualfund.com/system/files/2023-09/Portfolio%20as%20on%2030th%20June%2C%202023.xls"),
    ("2023-05", "https://betacms.tatamutualfund.com/system/files/2023-09/Portfolio%20as%20on%2031st%20May%2C%202023.xlsx"),
    ("2023-04", "https://betacms.tatamutualfund.com/system/files/2023-06/monthly-portfolio-as-on-30th-april-2023bab0d4a7981e44e2841fd351409d0c3c.xlsx"),
    ("2023-03", "https://betacms.tatamutualfund.com/system/files/2023-06/monthly-portfolio-as-on-31st-march-2023.xlsx"),
    ("2023-02", "https://betacms.tatamutualfund.com/system/files/2023-06/monthly-portfolio-as-on-28th-february-2023.xlsx"),
    ("2023-01", "https://betacms.tatamutualfund.com/system/files/2023-06/monthly-portfolio-as-on-31st-january-2023.xlsx"),
]

MOTILAL_URLS = [
    ("2026-04", "https://www.motilaloswalmf.com/content/dam/motilal-mf/downloads/mf/month-end-portfolio/2026/may/Motilal%20Portfolio%2030%20April%202026%20-%20Final.xlsx"),
    ("2026-03", "https://www.motilaloswalmf.com/content/dam/motilal-mf/downloads/mf/month-end-portfolio/2026/apr/IN_MF_MOTILAL_FACTSHEET_31.03.2026_Final.xlsx"),
    ("2026-02", "https://www.motilaloswalmf.com/CMS/assets/uploads/Documents/8c1a9-scheme-portfolio-details-february-26.xlsx"),
    ("2026-01", "https://www.motilaloswalmf.com/CMS/assets/uploads/Documents/b5209-scheme-portfolio-details-january-2026-2-.xlsx"),
    ("2025-12", "https://www.motilaloswalmf.com/CMS/assets/uploads/Documents/db566-scheme-portfolio-details-december-2025.xlsx"),
    ("2025-11", "https://www.motilaloswalmf.com/CMS/assets/uploads/Documents/ee45d-scheme-portfolio-details-november-2025.xlsx"),
    ("2025-10", "https://www.motilaloswalmf.com/CMS/assets/uploads/Documents/6abd7-scheme-portfolio-details-october-2025.xlsx"),
    ("2025-09", "https://www.motilaloswalmf.com/CMS/assets/uploads/Documents/9ec4e-scheme-portfolio-details-october-2025.xlsx"),
    ("2025-08", "https://www.motilaloswalmf.com/CMS/assets/uploads/Documents/deebc-scheme-portfolio-details-aug-2025.xlsx"),
    ("2025-07", "https://www.motilaloswalmf.com/CMS/assets/uploads/Documents/09555-scheme-portfolio-details-july-2025.xlsx"),
    ("2025-06", "https://www.motilaloswalmf.com/CMS/assets/uploads/Documents/bc9a7-month-end-portfolio-june-2025.xlsx"),
    ("2025-05", "https://www.motilaloswalmf.com/CMS/assets/uploads/Documents/27945-month-end-portfolio-may-2025.xlsx"),
    ("2025-04", "https://www.motilaloswalmf.com/CMS/assets/uploads/Documents/32d91-month-end-portfolio-april-2025.xlsx"),
    ("2025-03", "https://www.motilaloswalmf.com/CMS/assets/uploads/Documents/46d57-month-end-portfolio-march-2025.xls"),
    ("2025-02", "https://www.motilaloswalmf.com/CMS/assets/uploads/Documents/5a466-month-end-portfolio-february-2025.xlsx"),
    ("2025-01", "https://www.motilaloswalmf.com/CMS/assets/uploads/Documents/b1185-month-end-portfolio-january-2025.xlsx"),
    ("2024-12", "https://www.motilaloswalmf.com/CMS/assets/uploads/Documents/b4801-month-end-portfolio-december-2024.xls"),
    ("2024-11", "https://www.motilaloswalmf.com/CMS/assets/uploads/Documents/c9f2c-month-end-portfolio-november-2024.xls"),
    ("2024-10", "https://www.motilaloswalmf.com/CMS/assets/uploads/Documents/6f698-month-end-portfolio-october-2024.xls"),
    ("2024-09", "https://www.motilaloswalmf.com/CMS/assets/uploads/Documents/c6c4b-month-end-portfolio-september-2024.xls"),
    ("2024-08", "https://www.motilaloswalmf.com/CMS/assets/uploads/Documents/557e2-month-end-portfolio-august-2024.xls"),
    ("2024-07", "https://www.motilaloswalmf.com/CMS/assets/uploads/Documents/3d79f-month-end-portfolio-july-2024.xls"),
    ("2024-06", "https://www.motilaloswalmf.com/CMS/assets/uploads/Documents/b9860-month-end-portfolio-june-2024.xls"),
    ("2024-05", "https://www.motilaloswalmf.com/CMS/assets/uploads/Documents/5a56c-copy-of-in_mf_motilal_factsheet_31.05.2024_final.xls"),
    ("2024-04", "https://www.motilaloswalmf.com/CMS/assets/uploads/Documents/685c5-month-end-portfolio-april-2024.xls"),
    ("2024-03", "https://www.motilaloswalmf.com/CMS/assets/uploads/Documents/8199d-month-end-portfolio-march-2024.xls"),
    ("2024-02", "https://www.motilaloswalmf.com/CMS/assets/uploads/Documents/c09fa-month-end-portfolio-february-2024.xls"),
    ("2024-01", "https://www.motilaloswalmf.com/CMS/assets/uploads/Documents/9e0bc-month-end-portfolio-january-2024-2-.xlsx"),
    ("2023-12", "https://www.motilaloswalmf.com/CMS/assets/uploads/Documents/0b55b-month-end-portfolio-december-2023.xls"),
    ("2023-11", "https://www.motilaloswalmf.com/CMS/assets/uploads/Documents/1d69f-month-end-portfolio-november-2023.xlsx"),
    ("2023-10", "https://www.motilaloswalmf.com/CMS/assets/uploads/Documents/b7266-month-end-portfolio-october-2023.xls"),
    ("2023-09", "https://www.motilaloswalmf.com/CMS/assets/uploads/Documents/6afc0-month-end-portfolio-sept-24.xls"),
    ("2023-08", "https://www.motilaloswalmf.com/CMS/assets/uploads/Documents/f1cf0-month-end-portfolio-august-2023.xls"),
    ("2023-07", "https://www.motilaloswalmf.com/CMS/assets/uploads/Documents/d4eb4-month-end-portfolio-july-2023-2-.xls"),
    ("2023-06", "https://www.motilaloswalmf.com/CMS/assets/uploads/Documents/eb953-month-end-portfolio-for-july-2023.xls"),
    ("2023-05", "https://www.motilaloswalmf.com/CMS/assets/uploads/Documents/8e61a-month-end-portfolio-may-2023.xls"),
    ("2023-04", "https://www.motilaloswalmf.com/CMS/assets/uploads/Documents/a853f-month-end-portfolio-april-2023.xls"),
    ("2023-03", "https://www.motilaloswalmf.com/CMS/assets/uploads/Documents/06d94-21b6d-copy-of-in_mf_motilal_factsheet_31_march_2023.xls"),
    ("2023-02", "https://www.motilaloswalmf.com/CMS/assets/uploads/Documents/4a97f-month-end-portfolio-february-2023.xls"),
    ("2023-01", "https://www.motilaloswalmf.com/CMS/assets/uploads/Documents/4adca-motilal-oswal-month-end-portfolio-january-2023.xls"),
]

# ─────────────────────────────────────────────────────────────────────────────
# AMFI SCHEME LOOKUP
# ─────────────────────────────────────────────────────────────────────────────

def build_amfi_lookup():
    """Fetch AMFI NAVAll.txt and build name→scheme_code lookup for 5 new AMCs."""
    url = "https://portal.amfiindia.com/spages/NAVAll.txt"
    try:
        resp = requests.get(url, headers={'User-Agent': 'Mozilla/5.0'}, timeout=45)
        resp.raise_for_status()
        data = resp.content.decode('utf-8', errors='replace')
    except Exception as e:
        print(f"  WARN: AMFI fetch failed: {e}")
        return {}, {}

    lines = data.splitlines()
    print(f"  AMFI raw lines: {len(lines)}")

    amc_map = {
        'quant mutual fund': 'Quant',
        'uti mutual fund': 'UTI',
        'bandhan mutual fund': 'Bandhan',
        'tata mutual fund': 'Tata',
        'motilal oswal mutual fund': 'Motilal',
    }
    lookup = {}       # (amc, normalised_name) → scheme_code
    isin_lookup = {}  # isin → scheme_code
    current_amc = None

    for line in lines:
        line = line.strip()
        if not line:
            continue  # blank lines are separators only — do NOT reset current_amc
        if ';' not in line:
            # AMC header line (no semicolons)
            low = line.lower()
            matched = None
            for k, v in amc_map.items():
                if k in low:
                    matched = v
                    break
            current_amc = matched  # None if unrecognised AMC
        elif current_amc:
            parts = line.split(';')
            if len(parts) >= 4:
                code = parts[0].strip()
                isin = parts[1].strip()
                name = parts[3].strip()
                norm = re.sub(r'\s+', ' ', name.lower())
                lookup[(current_amc, norm)] = code
                if isin and isin != '-':
                    isin_lookup[isin] = code
    print(f"  AMFI lookup: {len(lookup)} name entries, {len(isin_lookup)} ISIN entries")
    return lookup, isin_lookup


# ─────────────────────────────────────────────────────────────────────────────
# DOWNLOAD HELPER
# ─────────────────────────────────────────────────────────────────────────────

def download_file(url, cache_path, amc_name=""):
    """Download url → cache_path, returning bytes. Skips if cached."""
    if cache_path.exists() and cache_path.stat().st_size > 1000:
        return cache_path.read_bytes()

    hdrs = dict(HEADERS)
    if 'tatamutualfund' in url:
        hdrs['Referer'] = 'https://www.tatamutualfund.com/'
        hdrs['Origin']  = 'https://www.tatamutualfund.com'
    elif 'bandhan' in url or 'storage.googleapis' in url:
        hdrs['Referer'] = 'https://bandhanmutual.com/'
    elif 'utimf' in url or 'cloudfront' in url:
        hdrs['Referer'] = 'https://www.utimf.com/'

    try:
        resp = requests.get(url, headers=hdrs, timeout=60, stream=True)
        if resp.status_code == 200:
            data = resp.content
            if len(data) < 1000:
                print(f"    WARN: tiny response ({len(data)} bytes) from {url[:60]}")
                return None
            cache_path.write_bytes(data)
            return data
        else:
            print(f"    HTTP {resp.status_code}: {url[:80]}")
            return None
    except Exception as e:
        print(f"    ERROR downloading {url[:60]}: {e}")
        return None


def load_excel(data_bytes, ext='.xlsx'):
    """Load Excel from bytes, returning openpyxl workbook."""
    import openpyxl, xlrd
    bio = io.BytesIO(data_bytes)
    if ext in ('.xlsx', '.xlsm'):
        try:
            return openpyxl.load_workbook(bio, read_only=True, data_only=True)
        except Exception as e:
            print(f"    openpyxl error: {e}")
            return None
    else:  # .xls
        try:
            book = xlrd.open_workbook(file_contents=data_bytes)
            return ('xlrd', book)
        except Exception as e:
            print(f"    xlrd error: {e}")
            return None


def wb_to_df(wb_or_tuple, sheet_name=None):
    """Convert workbook sheet to DataFrame."""
    import openpyxl, xlrd as xl
    if isinstance(wb_or_tuple, tuple) and wb_or_tuple[0] == 'xlrd':
        book = wb_or_tuple[1]
        names = book.sheet_names()
        if sheet_name:
            matches = [n for n in names if n.lower() == sheet_name.lower()]
            sh = book.sheet_by_name(matches[0] if matches else names[0])
        else:
            sh = book.sheet_by_index(0)
        data = [sh.row_values(r) for r in range(sh.nrows)]
        return pd.DataFrame(data)
    else:
        ws = wb_or_tuple[sheet_name] if sheet_name and sheet_name in wb_or_tuple.sheetnames else \
             wb_or_tuple[wb_or_tuple.sheetnames[0]]
        data = list(ws.values)
        return pd.DataFrame(data)


# ─────────────────────────────────────────────────────────────────────────────
# PARSERS
# ─────────────────────────────────────────────────────────────────────────────

def is_valid_isin(s):
    if not isinstance(s, str):
        return False
    s = s.strip()
    return bool(re.match(r'^IN[A-Z0-9]{10}$', s))


def safe_float(x):
    try:
        v = float(str(x).replace(',', '').replace('%', '').strip())
        return v if -100 < v < 200 else None
    except:
        return None


# ── QUANT ─────────────────────────────────────────────────────────────────────
def parse_quant(data_bytes, period):
    """
    Multi-sheet XLSX. Each sheet = one fund (no separate index).
    Layout per sheet:
      Row 0 col 2: AMC name  ('quant Mutual Fund')
      Row 1 col 2: Fund name ('quant Multi Cap Fund')
      Row 3 col 2: 'MONTHLY PORTFOLIO STATEMENT AS ON 30 Apr 2026'
      Row 7:       SR | ISIN | NAME OF THE INSTRUMENT | RATING | INDUSTRY | QTY | MKT VAL | % to NAV
      Row 8+:      holdings
    """
    import openpyxl
    try:
        wb = openpyxl.load_workbook(io.BytesIO(data_bytes), read_only=True, data_only=True)
    except Exception as e:
        print(f"    Quant openpyxl error: {e}")
        return []

    rows = []
    default_as_of = (pd.to_datetime(period + '-01') + pd.offsets.MonthEnd(0)).date()

    for sname in wb.sheetnames:
        try:
            ws = wb[sname]
            all_rows = list(ws.iter_rows(values_only=True))
        except:
            continue

        if len(all_rows) < 8:
            continue

        # Extract fund name: scan first 5 rows for the longest non-empty string in col 1 or 2
        fund_name = sname
        as_of = default_as_of
        for i, r in enumerate(all_rows[:6]):
            for col_idx in range(min(4, len(r))):
                cell = r[col_idx]
                if isinstance(cell, str) and len(cell.strip()) > 5:
                    # Date line
                    if re.search(r'as\s+on\s+\d', cell, re.I) or re.search(r'portfolio\s+statement', cell, re.I):
                        m = re.search(r'\d{1,2}\s+\w+\s+\d{4}|\d{1,2}[/-]\d{1,2}[/-]\d{2,4}', cell)
                        if m:
                            try:
                                as_of = pd.to_datetime(m.group(), dayfirst=True).date()
                            except:
                                pass
                    # Fund name: typically at row 1 and longer than 10 chars, not 'quant mutual fund'
                    elif i == 1 and 'mutual fund' not in cell.lower() and len(cell.strip()) > 8:
                        fund_name = cell.strip()

        # Find header row (contains 'isin')
        hdr_idx = None
        for i, r in enumerate(all_rows):
            vals = [str(v).lower().strip() if v else '' for v in r]
            if any('isin' in v for v in vals):
                hdr_idx = i
                break

        if hdr_idx is None:
            continue

        hdr = [str(v).lower().strip() if v else '' for v in all_rows[hdr_idx]]
        isin_col  = next((i for i, h in enumerate(hdr) if 'isin' in h), None)
        name_col  = next((i for i, h in enumerate(hdr) if 'name' in h or 'instrument' in h), None)
        pct_col   = next((i for i, h in enumerate(hdr) if '%' in h and 'nav' in h), None)

        if isin_col is None or pct_col is None:
            continue

        for r in all_rows[hdr_idx+1:]:
            if not r or not any(r):
                continue
            isin = str(r[isin_col]).strip() if len(r) > isin_col and r[isin_col] else ''
            if not is_valid_isin(isin):
                continue
            pct  = safe_float(r[pct_col]) if len(r) > pct_col else None
            if pct is None:
                continue
            sname_val = str(r[name_col]).strip() if name_col is not None and len(r) > name_col and r[name_col] else ''
            rows.append({
                'isin':        isin,
                'stock_name':  sname_val,
                'pct_nav':     pct,
                'scheme_name': fund_name,
                '_sheet':      sname,
                'amc':         'Quant',
                'as_of_date':  as_of,
            })

    return rows


# ── UTI ──────────────────────────────────────────────────────────────────────
def parse_uti(zip_bytes, period):
    """
    ZIP → find Excel file → single 'exposure' sheet.
    SCHEME CODE{N}STARTS markers in col0 → extract scheme blocks.
    Row 5 (0-indexed) = headers after each marker.
    """
    try:
        zf = zipfile.ZipFile(io.BytesIO(zip_bytes))
    except Exception as e:
        print(f"    UTI zip error: {e}")
        return []

    # Find Sebi Exposure Excel (the portfolio holdings file)
    # Accept .xlsx, .xls, and .xlsm (macro-enabled workbook, treated as xlsx)
    def _is_xl(n): return n.lower().endswith(('.xlsx', '.xls', '.xlsm'))
    def _is_xlsx_like(n): return n.lower().endswith(('.xlsx', '.xlsm'))

    # Priority 1: file name contains 'sebi' and 'exposure', not futures
    xls_name = None
    for name in zf.namelist():
        low = name.lower()
        if ('sebi' in low or 'exposure' in low) and _is_xl(name):
            if 'fut' not in low:
                xls_name = name
                break
    # Priority 2: largest xlsx/xlsm file (not risk-o-meter)
    if not xls_name:
        xlsx_files = [(zf.getinfo(n).file_size, n) for n in zf.namelist()
                      if _is_xlsx_like(n) and 'risk' not in n.lower()]
        if xlsx_files:
            xls_name = sorted(xlsx_files, reverse=True)[0][1]
    # Fallback: any xl file
    if not xls_name:
        for name in zf.namelist():
            if _is_xl(name):
                xls_name = name
                break

    if not xls_name:
        print(f"    UTI: no Excel in ZIP. Files: {zf.namelist()}")
        return []

    xls_data = zf.read(xls_name)
    # .xlsm is a macro-enabled xlsx — treat as xlsx for openpyxl
    ext = '.xlsx' if xls_name.lower().endswith(('.xlsx', '.xlsm')) else '.xls'

    as_of = pd.to_datetime(period + '-01') + pd.offsets.MonthEnd(0)
    as_of = as_of.date()

    # Try to parse date from filename
    m = re.search(r'(\d{2})[._\s](\d{2})[._\s](\d{4})', xls_name)
    if m:
        try:
            as_of = datetime(int(m.group(3)), int(m.group(2)), int(m.group(1))).date()
        except:
            pass

    rows = []

    if ext == '.xlsx':
        import openpyxl
        try:
            wb = openpyxl.load_workbook(io.BytesIO(xls_data), read_only=True, data_only=True)
        except Exception as e:
            print(f"    UTI openpyxl: {e}")
            return []
        # Pick exposure sheet or first sheet
        sheet_name = next((s for s in wb.sheetnames if 'exposure' in s.lower()), wb.sheetnames[0])
        ws = wb[sheet_name]
        all_rows = list(ws.iter_rows(values_only=True))
    else:
        import xlrd
        try:
            book = xlrd.open_workbook(file_contents=xls_data)
        except Exception as e:
            print(f"    UTI xlrd: {e}")
            return []
        sh_names = book.sheet_names()
        sh_idx = next((i for i, n in enumerate(sh_names) if 'exposure' in n.lower()), 0)
        sh = book.sheet_by_index(sh_idx)
        all_rows = [sh.row_values(r) for r in range(sh.nrows)]

    # Parse scheme blocks via SCHEME CODE{N}STARTS / ENDS markers.
    # Confirmed layout (Apr 2026):
    #   Row 0: SCHEME CODE002STARTS   (col0)
    #   Row 1: UTI MUTUAL FUND
    #   Row 2: SCHEME: UTI - <Scheme Name>  ← extract scheme name here
    #   Row 3: PROVISIONAL AND UNAUDITED…   (often has date)
    #   Row 4: (blank)
    #   Row 5: NAME OF THE INSTRUMENT | RATING/INDUSTRY | QUANTITY | MARKET-VALUE | % TO NAV | _ | _ | ISIN
    #   Row 6+: data (col0=name, col4=%nav, col7=isin)

    current_scheme = None
    hdr_col_isin = None   # confirmed col idx for ISIN
    hdr_col_pct  = None   # confirmed col idx for % TO NAV
    hdr_col_name = None   # confirmed col idx for instrument name
    in_block = False

    for i, row in enumerate(all_rows):
        if not row:
            continue
        cell0 = str(row[0]).strip() if row[0] else ''

        # Block start marker
        if re.search(r'SCHEME\s*CODE\s*\d*\s*STARTS', cell0, re.I):
            in_block = True
            current_scheme = None
            hdr_col_isin = hdr_col_pct = hdr_col_name = None
            continue

        # Block end marker
        if re.search(r'SCHEME\s*CODE\s*\d*\s*ENDS', cell0, re.I):
            in_block = False
            continue

        if not in_block:
            continue

        # Extract scheme name from 'SCHEME: UTI - ...' line
        if re.match(r'SCHEME\s*:', cell0, re.I):
            current_scheme = re.sub(r'^SCHEME\s*:\s*', '', cell0, flags=re.I).strip()
            continue

        # Look for header row (contains ISIN in any column)
        if hdr_col_isin is None:
            vals = [str(v).lower().strip() if v else '' for v in row]
            if any('isin' in v for v in vals):
                hdr_col_isin = next((j for j, h in enumerate(vals) if 'isin' in h), None)
                hdr_col_pct  = next((j for j, h in enumerate(vals) if '%' in h and 'nav' in h), None)
                hdr_col_name = next((j for j, h in enumerate(vals) if 'instrument' in h or 'name' in h), 0)
                continue

        # Data row
        if hdr_col_isin is not None and current_scheme:
            isin = str(row[hdr_col_isin]).strip() if len(row) > hdr_col_isin and row[hdr_col_isin] else ''
            if not is_valid_isin(isin):
                continue
            pct = safe_float(row[hdr_col_pct]) if hdr_col_pct is not None and len(row) > hdr_col_pct else None
            if pct is None:
                continue
            name_val = str(row[hdr_col_name]).strip() if hdr_col_name is not None and len(row) > hdr_col_name and row[hdr_col_name] else ''
            rows.append({
                'isin':        isin,
                'stock_name':  name_val,
                'pct_nav':     pct,
                'scheme_name': current_scheme,
                '_sheet':      'exposure',
                'amc':         'UTI',
                'as_of_date':  as_of,
            })

    return rows


# ── TATA ─────────────────────────────────────────────────────────────────────
def parse_tata(data_bytes, period, ext='.xlsx'):
    """
    Multi-sheet XLSX/XLS.
    Index sheet: maps code → name.
    Fund sheets: row0 col1=fund_name, row9 col1='as on {date}', row11=headers, data from row14.
    Headers: NAME(1), YIELD(2), INDUSTRY(3), ISIN CODE(4), QTY(5), MKT VAL(6), % to NAV(7)
    """
    as_of = pd.to_datetime(period + '-01') + pd.offsets.MonthEnd(0)
    as_of = as_of.date()

    if ext in ('.xlsx', '.xlsm'):
        import openpyxl
        try:
            wb = openpyxl.load_workbook(io.BytesIO(data_bytes), read_only=True, data_only=True)
        except Exception as e:
            print(f"    Tata openpyxl: {e}")
            return []
        sheets = wb.sheetnames
        def get_sheet_rows(sname):
            return list(wb[sname].iter_rows(values_only=True))
    else:
        import xlrd
        try:
            book = xlrd.open_workbook(file_contents=data_bytes)
        except Exception as e:
            print(f"    Tata xlrd: {e}")
            return []
        sheets = book.sheet_names()
        def get_sheet_rows(sname):
            sh = book.sheet_by_name(sname)
            return [sh.row_values(r) for r in range(sh.nrows)]

    # Build index from 'Index' sheet
    # Tata Index layout: col0=CLASSIFICATION, col1=SCHEME_CODE, col2=SCHEME_NAME
    index = {}
    for sname in sheets:
        if sname.lower().strip() == 'index':
            for row in get_sheet_rows(sname):
                if not row or len(row) < 3:
                    continue
                # Skip header rows
                code = str(row[1]).strip() if row[1] else ''
                name = str(row[2]).strip() if row[2] else ''
                if code and name and code.lower() not in ('scheme code', 'code', 'fund code', 'classification'):
                    index[code] = name
            break

    rows = []

    for sname in sheets:
        if sname.lower().strip() in ('index', 'contents', 'cover page', 'cover',
                                     'dividend history', 'tata scheme risk-o-meter', ''):
            continue

        try:
            all_rows = get_sheet_rows(sname)
        except:
            continue

        if len(all_rows) < 10:
            continue

        fund_name = index.get(sname, sname)

        # Extract date from row 9 or any row containing 'as on'
        for row in all_rows[:15]:
            for cell in row:
                if isinstance(cell, str) and 'as on' in cell.lower():
                    try:
                        d_str = re.sub(r'as on\s*', '', cell, flags=re.I).strip()
                        # remove ordinal suffixes
                        d_str = re.sub(r'(\d+)(st|nd|rd|th)', r'\1', d_str, flags=re.I)
                        as_of_local = pd.to_datetime(d_str, dayfirst=True).date()
                        as_of = as_of_local
                    except:
                        pass
                    break

        # Find header row
        hdr_idx = None
        for i, row in enumerate(all_rows):
            vals = [str(v).lower().strip() if v else '' for v in row]
            if sum(1 for v in vals if 'isin' in v or '% to nav' in v or 'market value' in v) >= 2:
                hdr_idx = i
                break

        if hdr_idx is None:
            # Try to find any row with ISIN
            for i, row in enumerate(all_rows):
                vals = [str(v).lower().strip() if v else '' for v in row]
                if any('isin' in v for v in vals):
                    hdr_idx = i
                    break

        if hdr_idx is None:
            continue

        hdr = [str(v).lower().strip() if v else '' for v in all_rows[hdr_idx]]
        isin_col = next((i for i, h in enumerate(hdr) if 'isin' in h), None)
        name_col = next((i for i, h in enumerate(hdr) if 'name' in h or 'instrument' in h or 'security' in h), None)
        pct_col  = next((i for i, h in enumerate(hdr) if '%' in h and 'nav' in h), None)

        if isin_col is None or pct_col is None:
            continue

        for row in all_rows[hdr_idx+1:]:
            if not row or not any(row):
                continue
            isin = str(row[isin_col]).strip() if len(row) > isin_col and row[isin_col] else ''
            if not is_valid_isin(isin):
                continue
            pct  = safe_float(row[pct_col]) if len(row) > pct_col else None
            if pct is None:
                continue
            sname_val = str(row[name_col]).strip() if name_col is not None and len(row) > name_col and row[name_col] else ''
            rows.append({
                'isin':        isin,
                'stock_name':  sname_val,
                'pct_nav':     pct,
                'scheme_name': fund_name,
                '_sheet':      sname,
                'amc':         'Tata',
                'as_of_date':  as_of,
            })

    return rows


# ── MOTILAL ──────────────────────────────────────────────────────────────────
def parse_motilal(data_bytes, period, ext='.xlsx'):
    """
    Multi-sheet XLSX/XLS.
    INDEX sheet: maps code → name.
    Fund sheets (code like 'YOXX'): row5=date, row9=headers.
    Headers: Sr No(0), Name(1), ISIN Code(2), Industry(3), Qty(4), Mkt Val(5), % to NAV(6)
    """
    as_of = pd.to_datetime(period + '-01') + pd.offsets.MonthEnd(0)
    as_of = as_of.date()

    if ext in ('.xlsx', '.xlsm'):
        import openpyxl
        try:
            wb = openpyxl.load_workbook(io.BytesIO(data_bytes), read_only=True, data_only=True)
        except Exception as e:
            print(f"    Motilal openpyxl: {e}")
            return []
        sheets = wb.sheetnames
        def get_sheet_rows(sname):
            return list(wb[sname].iter_rows(values_only=True))
    else:
        import xlrd
        try:
            book = xlrd.open_workbook(file_contents=data_bytes)
        except Exception as e:
            print(f"    Motilal xlrd: {e}")
            return []
        sheets = book.sheet_names()
        def get_sheet_rows(sname):
            sh = book.sheet_by_name(sname)
            return [sh.row_values(r) for r in range(sh.nrows)]

    # Build index from INDEX sheet
    # Motilal INDEX layout: col2=Sr No (1,2,3...), col3=Fund Name
    # Sheet names: YO01, YO02, ... where suffix = zero-padded Sr No
    index = {}       # sheet_code → fund_name
    srno_map = {}    # Sr No (int) → fund_name
    for sname in sheets:
        if sname.lower().strip() == 'index':
            for row in get_sheet_rows(sname):
                if not row or len(row) < 4:
                    continue
                sr_raw  = row[2]
                nm_raw  = row[3]
                if sr_raw is None or nm_raw is None:
                    continue
                name = str(nm_raw).strip()
                if not name or name.lower() in ('fund name', 'scheme name', 'name'):
                    continue
                try:
                    sr = int(float(str(sr_raw)))
                    srno_map[sr] = name
                except (ValueError, TypeError):
                    pass
            break

    rows = []

    for sname in sheets:
        if sname.lower() in ('index', 'index ', 'contents', 'cover', ''):
            continue

        try:
            all_rows = get_sheet_rows(sname)
        except:
            continue

        if len(all_rows) < 5:
            continue

        # Resolve fund name: YO01 → Sr No 1 → srno_map[1]
        m_sr = re.search(r'(\d+)$', sname)
        if m_sr:
            sr_num = int(m_sr.group(1))
            fund_name = srno_map.get(sr_num, index.get(sname, sname))
        else:
            fund_name = index.get(sname, sname)

        # Extract date: search first 10 rows for a date cell or 'as on' text
        as_of_local = as_of
        for row in all_rows[:10]:
            for cell in row:
                if isinstance(cell, datetime):
                    as_of_local = cell.date()
                    break
                if isinstance(cell, str):
                    m = re.search(r'\d{1,2}[/-]\d{1,2}[/-]\d{2,4}', cell)
                    if m:
                        try:
                            as_of_local = pd.to_datetime(m.group(), dayfirst=True).date()
                        except:
                            pass

        # Find header row containing ISIN
        hdr_idx = None
        for i, row in enumerate(all_rows):
            vals = [str(v).lower().strip() if v else '' for v in row]
            if any('isin' in v for v in vals):
                hdr_idx = i
                break

        if hdr_idx is None:
            continue

        hdr = [str(v).lower().strip() if v else '' for v in all_rows[hdr_idx]]
        isin_col = next((i for i, h in enumerate(hdr) if 'isin' in h), None)
        name_col = next((i for i, h in enumerate(hdr) if 'name' in h or 'instrument' in h or 'security' in h), None)
        pct_col  = next((i for i, h in enumerate(hdr) if '%' in h and 'nav' in h), None)
        if pct_col is None:
            # Fallback: last column
            pct_col = len(hdr) - 1

        if isin_col is None:
            continue

        for row in all_rows[hdr_idx+1:]:
            if not row or not any(row):
                continue
            isin = str(row[isin_col]).strip() if len(row) > isin_col and row[isin_col] else ''
            if not is_valid_isin(isin):
                continue
            pct  = safe_float(row[pct_col]) if len(row) > pct_col else None
            if pct is None:
                continue
            sname_val = str(row[name_col]).strip() if name_col is not None and len(row) > name_col and row[name_col] else ''
            rows.append({
                'isin':        isin,
                'stock_name':  sname_val,
                'pct_nav':     pct,
                'scheme_name': fund_name,
                '_sheet':      sname,
                'amc':         'Motilal',
                'as_of_date':  as_of_local,
            })

    return rows


# ── BANDHAN ──────────────────────────────────────────────────────────────────
def parse_bandhan(data_bytes, period, filename=''):
    """
    Per-scheme XLSX (one file per scheme, from GCS).
    Confirmed layout (Apr 2026):
      Row 0: (scheme_code, 'Portfolio Statement as on April 30,2026', ...)
      Row 1: (None, fund_name, ...)
      Row 2: blank
      Row 3: headers (None, 'Name of the Instrument', 'ISIN', 'Industry/Rating',
                       'Quantity', 'Market/Fair Value', '% to NAV', 'YTM')
      Row 4+: data — ISIN at col 2, name at col 1, pct at col 6 (DECIMAL 0-1, multiply ×100)
    """
    import openpyxl
    as_of = pd.to_datetime(period + '-01') + pd.offsets.MonthEnd(0)
    as_of = as_of.date()

    try:
        wb = openpyxl.load_workbook(io.BytesIO(data_bytes), read_only=True, data_only=True)
    except Exception as e:
        print(f"    Bandhan openpyxl: {e}")
        return []

    ws = wb[wb.sheetnames[0]]
    all_rows = list(ws.iter_rows(values_only=True))

    if len(all_rows) < 5:
        return []

    # Row 0: (scheme_code, 'Portfolio Statement as on {Month} {DD},{YYYY}')
    scheme_code_file = str(all_rows[0][0]).strip() if all_rows[0][0] else ''

    # Extract date from row 0 col 1 string like "Portfolio Statement as on April 30,2026"
    if all_rows[0][1] and isinstance(all_rows[0][1], str):
        cell = all_rows[0][1]
        # Pattern: "Month DD,YYYY" or "Month DD, YYYY"
        m = re.search(r'(\w+ \d{1,2},?\s*\d{4})', cell)
        if m:
            try:
                date_str = re.sub(r',\s*', ' ', m.group(1))
                as_of = pd.to_datetime(date_str, format='%B %d %Y').date()
            except:
                # Try more general parse
                m2 = re.search(r'(\d{1,2})\s+(\w+)\s+(\d{4})', cell)
                if m2:
                    try:
                        as_of = pd.to_datetime(f"{m2.group(1)} {m2.group(2)} {m2.group(3)}", dayfirst=True).date()
                    except:
                        pass

    # Row 1: (None, fund_name, ...)
    fund_name = ''
    if len(all_rows) > 1:
        # Fund name is typically at col 1
        for col_idx in range(min(3, len(all_rows[1]))):
            cell = all_rows[1][col_idx]
            if cell and isinstance(cell, str) and len(cell.strip()) > 5:
                fund_name = cell.strip()
                break

    # Find header row (contains 'ISIN')
    hdr_idx = None
    for i, row in enumerate(all_rows[:8]):
        vals = [str(v).strip() if v else '' for v in row]
        if any('ISIN' in v or 'isin' in v.lower() for v in vals):
            hdr_idx = i
            break

    if hdr_idx is None:
        return []

    hdr = [str(v).lower().strip() if v else '' for v in all_rows[hdr_idx]]
    isin_col = next((i for i, h in enumerate(hdr) if h == 'isin' or 'isin' in h), None)
    name_col = next((i for i, h in enumerate(hdr) if 'name' in h or 'instrument' in h or 'security' in h), None)
    pct_col  = next((i for i, h in enumerate(hdr) if '%' in h and 'nav' in h), None)

    if isin_col is None or pct_col is None:
        return []

    # Detect if pct values are decimal (0-1 range) vs percentage (0-100 range)
    # Sample a few rows to decide
    sample_pcts = []
    for row in all_rows[hdr_idx+1:hdr_idx+10]:
        if row and len(row) > pct_col and row[pct_col] is not None:
            v = safe_float(row[pct_col])
            if v and v > 0:
                sample_pcts.append(v)
    pct_multiplier = 100.0 if sample_pcts and max(sample_pcts) < 2.0 else 1.0

    rows = []
    for row in all_rows[hdr_idx+1:]:
        if not row or not any(row):
            continue
        isin = str(row[isin_col]).strip() if len(row) > isin_col and row[isin_col] else ''
        if not is_valid_isin(isin):
            continue
        pct  = safe_float(row[pct_col]) if len(row) > pct_col else None
        if pct is None:
            continue
        pct *= pct_multiplier
        sname_val = str(row[name_col]).strip() if name_col is not None and len(row) > name_col and row[name_col] else ''
        rows.append({
            'isin':        isin,
            'stock_name':  sname_val,
            'pct_nav':     pct,
            'scheme_name': fund_name,
            '_sheet':      scheme_code_file,
            'amc':         'Bandhan',
            'as_of_date':  as_of,
        })

    return rows


# ─────────────────────────────────────────────────────────────────────────────
# BANDHAN URL DISCOVERY
# ─────────────────────────────────────────────────────────────────────────────

# Equity / hybrid scheme keywords for Bandhan title-based filtering
BANDHAN_EQUITY_KWS = [
    'elss', 'flexi cap', 'small cap', 'mid cap', 'large cap', 'large &',
    'business cycle', 'aggressive hybrid', 'equity', 'infrastructure',
    'multi cap', 'value fund', 'contra', 'momentum', 'focused',
    'dividend yield', 'nifty', 'sensex', 'index fund', 'sector',
    'quant', 'alpha', 'manufacturing', 'consumption', 'pharma',
    'technology', 'financial', 'banking', 'bluechip', 'multicap',
    'balanced advantage', 'multi asset',
]

BANDHAN_SKIP_KWS = [
    'liquid', 'overnight', 'gilt', 'g-sec', 'debt', 'bond', 'income',
    'ultra short', 'low duration', 'short duration', 'medium duration',
    'credit risk', 'money market', 'floater', 'dynamic bond',
    'fixed maturity', 'fmp', 'interval', 'arbitrage', 'fixed term plan',
    'banking & psu', 'banking and psu', 'ibx', 'crisil',
    'constant duration', 'corporate bond',
]


def is_bandhan_equity_scheme(title):
    """Filter Bandhan scheme titles to equity/hybrid only."""
    t = title.lower()
    if any(k in t for k in BANDHAN_SKIP_KWS):
        return False
    return True  # Accept anything not explicitly excluded (conservative—let ISIN filter do final cut)


def get_bandhan_urls_from_catalog(catalog_path=None):
    """
    Load Bandhan monthly URLs from the pre-built JSON catalog.
    Falls back to live API enumeration if catalog is missing.
    Returns list of (period, url, title) tuples for equity-ish schemes only.
    """
    if catalog_path is None:
        catalog_path = CACHE.parent / 'bandhan_monthly_urls.json'

    if not Path(catalog_path).exists():
        print(f"  WARN: Bandhan URL catalog not found at {catalog_path}")
        print("  Run fetch_bandhan_urls.py first to build the catalog.")
        return []

    with open(catalog_path) as f:
        catalog = json.load(f)

    result = []
    for period in sorted(catalog.keys()):
        if period < '2023-01':
            continue
        entries = catalog[period]
        for entry in entries:
            url   = entry.get('url', '')
            title = entry.get('title', '')
            if not url or not is_bandhan_equity_scheme(title):
                continue
            result.append((period, url, title))

    periods = sorted(set(p for p, _, _ in result))
    print(f"  Bandhan catalog: {len(result)} scheme-files across {len(periods)} periods: {periods[0]}–{periods[-1]}")
    return result


def get_bandhan_urls_live(max_pages_per_fy=300):
    """
    Live enumeration of Bandhan WP API for monthly disclosures.
    Uses disclosures_type='Monthly and Half-yearly Disclosures'.
    Returns list of (period, url, title) tuples.
    """
    import re as _re

    api = "https://cmsnew.bandhanmutual.com/wp-json/finance-api/v1/posts/disclosures"
    hdrs = {
        'User-Agent': HEADERS['User-Agent'],
        'Origin': 'https://bandhanmutual.com',
        'Referer': 'https://bandhanmutual.com/',
        'Accept': 'application/json',
    }

    def _extract_period(title):
        m = _re.search(r'(\d{1,2})\s+(\w+)\s+(\d{4})', title)
        if m:
            try:
                dt = pd.to_datetime(f"{m.group(1)} {m.group(2)} {m.group(3)}", dayfirst=True)
                return dt.strftime('%Y-%m')
            except:
                pass
        return None

    period_urls = {}
    seen_ids = set()

    for fy in ['2025', '2024', '2023', '2022']:
        for page in range(1, max_pages_per_fy):
            params = {'financial_year': fy, 'per_page': 50, 'page': page}
            try:
                resp = requests.get(api, headers=hdrs, params=params, timeout=20)
                data = resp.json()
                entries = data.get('data', [])
                if not entries:
                    break
                new_entries = [e for e in entries if e.get('id') not in seen_ids]
                if not new_entries:
                    break
                for e in new_entries:
                    seen_ids.add(e.get('id'))
                    acf = e.get('acf_fields', {}) or {}
                    dtype = (acf.get('disclosures_type', '') or '').lower()
                    title = e.get('title', '') or ''
                    if 'monthly' not in dtype:
                        continue
                    period = _extract_period(title)
                    if not period or period < '2023-01':
                        continue
                    disc_files = acf.get('disclosure_files', []) or []
                    for df in disc_files:
                        if not isinstance(df, dict):
                            continue
                        doc = df.get('document_link', {}) or {}
                        url = doc.get('url', '') if isinstance(doc, dict) else ''
                        if url and url.startswith('http'):
                            if period not in period_urls:
                                period_urls[period] = []
                            if not any(x[0] == url for x in period_urls[period]):
                                period_urls[period].append((url, title))
            except Exception as ex:
                if 'bool' not in str(ex):
                    print(f"    Bandhan API fy={fy} p={page}: {ex}")
                continue
            time.sleep(0.1)

    result = []
    for period in sorted(period_urls.keys()):
        for url, title in period_urls[period]:
            if is_bandhan_equity_scheme(title):
                result.append((period, url, title))
    return result


# ─────────────────────────────────────────────────────────────────────────────
# SCHEME CODE MAPPING
# ─────────────────────────────────────────────────────────────────────────────

def build_scheme_code_map(amc, rows, amfi_name_lookup, amfi_isin_lookup):
    """
    Map scheme_name → scheme_code using AMFI data.
    Returns dict: scheme_name → scheme_code
    """
    names = set(r['scheme_name'] for r in rows)
    mapping = {}

    for name in names:
        if not name:
            continue

        # Try exact normalised match first
        norm = re.sub(r'\s+', ' ', name.lower().strip())
        if (amc, norm) in amfi_name_lookup:
            mapping[name] = amfi_name_lookup[(amc, norm)]
            continue

        # Try partial: find best matching Direct Growth scheme
        best = None
        best_score = 0
        for (a, n), code in amfi_name_lookup.items():
            if a != amc:
                continue
            # Prefer direct plan, growth option
            if 'direct' not in n or ('growth' not in n and 'gr' not in n):
                continue
            # Score by word overlap
            name_words = set(re.findall(r'\w+', norm)) - {'fund', 'the', 'of', 'and', 'plan', 'option', 'direct', 'growth', 'regular'}
            cand_words = set(re.findall(r'\w+', n))
            if not name_words:
                continue
            score = len(name_words & cand_words) / len(name_words)
            if score > best_score:
                best_score = score
                best = code

        if best and best_score >= 0.5:
            mapping[name] = best

    found = sum(1 for v in mapping.values() if v)
    print(f"  Scheme code mapping: {found}/{len(names)} schemes resolved for {amc}")
    return mapping


# ─────────────────────────────────────────────────────────────────────────────
# WRITE TO HOLDINGS PARQUETS
# ─────────────────────────────────────────────────────────────────────────────

def append_to_holdings(rows, scheme_map, amc):
    """Append parsed rows to monthly holdings parquet files."""
    if not rows:
        return 0

    df = pd.DataFrame(rows)
    # scheme_code: map name→code, then cast to Float64 (nullable) to match existing schema
    raw_code = df['scheme_name'].map(scheme_map)
    df['scheme_code'] = pd.to_numeric(raw_code, errors='coerce').astype('Float64')
    df['as_of_date']  = pd.to_datetime(df['as_of_date'])
    df['year']  = df['as_of_date'].dt.year
    df['month'] = df['as_of_date'].dt.month

    # Columns to write — match existing holdings schema exactly
    WRITE_COLS = ['isin', 'stock_name', 'pct_nav', 'scheme_name', '_sheet',
                  'amc', 'scheme_code', 'as_of_date']

    total = 0
    for (yr, mo), grp in df.groupby(['year', 'month']):
        period = f"{yr:04d}-{mo:02d}"
        fpath  = HOLD / f"{period}.parquet"

        # Keep only schema columns that exist in this df
        out_cols = [c for c in WRITE_COLS if c in grp.columns]
        grp_out = grp[out_cols].copy()

        if fpath.exists():
            existing = pd.read_parquet(fpath)
            # Remove old rows for this AMC in this period
            existing = existing[existing['amc'] != amc]
            combined = pd.concat([existing, grp_out], ignore_index=True)
        else:
            combined = grp_out

        combined.to_parquet(fpath, index=False)
        total += len(grp_out)
        print(f"    {period}: wrote {len(grp_out)} rows for {amc}  (file total: {len(combined)})")

    return total


# ─────────────────────────────────────────────────────────────────────────────
# MAIN INGEST ROUTINES
# ─────────────────────────────────────────────────────────────────────────────

def ingest_quant(amfi_name_lookup, amfi_isin_lookup, dry_run=False):
    print("\n── QUANT ──")
    all_rows = []
    for period, url in QUANT_URLS:
        fname  = url.split('/')[-1].split('?')[0]
        cpath  = CACHE / f"quant_{period}_{fname}"
        data   = download_file(url, cpath, 'Quant')
        if data is None:
            print(f"  {period}: SKIP (download failed)")
            continue
        rows = parse_quant(data, period)
        print(f"  {period}: {len(rows)} holding rows")
        all_rows.extend(rows)
        time.sleep(0.3)

    if not dry_run and all_rows:
        sm = build_scheme_code_map('Quant', all_rows, amfi_name_lookup, amfi_isin_lookup)
        n  = append_to_holdings(all_rows, sm, 'Quant')
        print(f"  Quant total: {n} rows written")
    return all_rows


def ingest_uti(amfi_name_lookup, amfi_isin_lookup, dry_run=False):
    print("\n── UTI ──")
    all_rows = []
    for period, url in UTI_URLS:
        fname  = re.sub(r'[^\w.]', '_', url.split('/')[-1].split('?')[0])[:60]
        cpath  = CACHE / f"uti_{period}_{fname}"
        data   = download_file(url, cpath, 'UTI')
        if data is None:
            print(f"  {period}: SKIP (download failed)")
            continue
        rows = parse_uti(data, period)
        print(f"  {period}: {len(rows)} holding rows")
        all_rows.extend(rows)
        time.sleep(0.4)

    if not dry_run and all_rows:
        sm = build_scheme_code_map('UTI', all_rows, amfi_name_lookup, amfi_isin_lookup)
        n  = append_to_holdings(all_rows, sm, 'UTI')
        print(f"  UTI total: {n} rows written")
    return all_rows


def ingest_tata(amfi_name_lookup, amfi_isin_lookup, dry_run=False):
    print("\n── TATA ──")
    all_rows = []
    for period, url in TATA_URLS:
        ext    = '.xls' if url.lower().split('?')[0].endswith('.xls') else '.xlsx'
        fname  = re.sub(r'[^\w.]', '_', url.split('/')[-1].split('?')[0])[:60]
        cpath  = CACHE / f"tata_{period}_{fname}"
        data   = download_file(url, cpath, 'Tata')
        if data is None:
            print(f"  {period}: SKIP (download failed)")
            continue
        rows = parse_tata(data, period, ext)
        print(f"  {period}: {len(rows)} holding rows")
        all_rows.extend(rows)
        time.sleep(0.3)

    if not dry_run and all_rows:
        sm = build_scheme_code_map('Tata', all_rows, amfi_name_lookup, amfi_isin_lookup)
        n  = append_to_holdings(all_rows, sm, 'Tata')
        print(f"  Tata total: {n} rows written")
    return all_rows


def ingest_motilal(amfi_name_lookup, amfi_isin_lookup, dry_run=False):
    print("\n── MOTILAL ──")
    all_rows = []
    for period, url in MOTILAL_URLS:
        ext    = '.xls' if url.lower().split('?')[0].endswith('.xls') else '.xlsx'
        fname  = re.sub(r'[^\w.]', '_', url.split('/')[-1].split('?')[0])[:60]
        cpath  = CACHE / f"motilal_{period}_{fname}"
        data   = download_file(url, cpath, 'Motilal')
        if data is None:
            print(f"  {period}: SKIP (download failed)")
            continue
        rows = parse_motilal(data, period, ext)
        print(f"  {period}: {len(rows)} holding rows")
        all_rows.extend(rows)
        time.sleep(0.3)

    if not dry_run and all_rows:
        sm = build_scheme_code_map('Motilal', all_rows, amfi_name_lookup, amfi_isin_lookup)
        n  = append_to_holdings(all_rows, sm, 'Motilal')
        print(f"  Motilal total: {n} rows written")
    return all_rows


def ingest_bandhan(amfi_name_lookup, amfi_isin_lookup, dry_run=False):
    print("\n── BANDHAN ──")

    # Load URL catalog (pre-built by fetch_bandhan_urls.py)
    catalog_path = BASE / 'bandhan_monthly_urls.json'
    url_list = get_bandhan_urls_from_catalog(catalog_path)

    if not url_list:
        print("  Catalog empty — attempting live API enumeration (slow)...")
        url_list = get_bandhan_urls_live()
        print(f"  Live enumeration: {len(url_list)} scheme-files")

    # Group by period
    from collections import defaultdict
    by_period = defaultdict(list)
    for period, url, title in url_list:
        by_period[period].append((url, title))

    all_rows = []
    for period in sorted(by_period.keys()):
        period_rows = []
        scheme_files = by_period[period]
        print(f"  {period}: {len(scheme_files)} scheme-files", end='', flush=True)
        for url, title in scheme_files:
            fname = re.sub(r'[^\w.]', '_', url.split('/')[-1].split('?')[0])[:60]
            cpath = CACHE / f"bandhan_{period}_{fname}"
            data  = download_file(url, cpath, 'Bandhan')
            if data is None:
                continue
            rows = parse_bandhan(data, period, fname)
            period_rows.extend(rows)
            time.sleep(0.1)

        # Deduplicate within period
        if period_rows:
            df_p = pd.DataFrame(period_rows)
            df_p = df_p.drop_duplicates(subset=['isin', 'scheme_name', 'as_of_date'])
            period_rows = df_p.to_dict('records')

        print(f" → {len(period_rows)} rows")
        all_rows.extend(period_rows)

    print(f"  Bandhan total: {len(all_rows)} holding rows across {len(by_period)} periods")

    if not dry_run and all_rows:
        sm = build_scheme_code_map('Bandhan', all_rows, amfi_name_lookup, amfi_isin_lookup)
        n  = append_to_holdings(all_rows, sm, 'Bandhan')
        print(f"  Bandhan rows written: {n}")
    return all_rows


# ─────────────────────────────────────────────────────────────────────────────
# EXTEND FUND_META
# ─────────────────────────────────────────────────────────────────────────────

# Equity/Hybrid category keywords to filter AMFI schemes
EQUITY_KEYWORDS = [
    'large cap', 'mid cap', 'small cap', 'multi cap', 'flexi cap',
    'large & mid', 'large and mid', 'focused fund', 'value fund', 'contra fund',
    'elss', 'tax saving', 'dividend yield', 'sectoral', 'thematic',
    'hybrid', 'balanced', 'multi asset', 'arbitrage', 'equity savings',
    'international', 'global', 'india', 'infrastructure', 'manufacturing',
    'consumption', 'financial services', 'technology', 'healthcare', 'pharma',
]

EXCLUDE_KEYWORDS = ['liquid', 'overnight', 'gilt', 'g-sec', 'banking & psu', 'banking and psu',
                     'debt', 'bond', 'income', 'ultra short', 'low duration', 'short duration',
                     'medium duration', 'credit risk', 'money market', 'floater', 'dynamic bond',
                     'fixed maturity', 'fmp', 'interval', 'idcw', 'reinvestment', 'fortnightly',
                     'weekly', 'monthly', 'quarterly', 'annual', 'half', 'payout',
                     'unclaimed', 'segregated']


def is_equity_scheme(name):
    n = name.lower()
    if any(ex in n for ex in EXCLUDE_KEYWORDS):
        return False
    if 'direct' not in n:
        return False
    if 'growth' not in n:
        return False
    # Accept broad equity + some hybrids
    return True  # Accept all direct growth for equity AMCs


def extend_fund_meta(amfi_name_lookup):
    """Add new AMC schemes to fund_meta.parquet and scheme_list.parquet."""
    fm_path = BASE / 'fund_meta.parquet'
    sl_path = BASE / 'scheme_list.parquet'

    fm = pd.read_parquet(fm_path)
    sl = pd.read_parquet(sl_path)

    print(f"\n── EXTEND FUND META ──")
    print(f"  Current: {len(fm)} schemes, {sorted(fm['amc'].unique())}")

    # Re-fetch AMFI to get full scheme details
    url = "https://portal.amfiindia.com/spages/NAVAll.txt"
    try:
        resp = requests.get(url, headers={'User-Agent': 'Mozilla/5.0'}, timeout=45)
        resp.raise_for_status()
        data = resp.content.decode('utf-8', errors='replace')
    except Exception as e:
        print(f"  ERROR: AMFI fetch failed: {e}")
        return pd.DataFrame()

    amc_map = {
        'quant mutual fund': 'Quant',
        'uti mutual fund': 'UTI',
        'bandhan mutual fund': 'Bandhan',
        'tata mutual fund': 'Tata',
        'motilal oswal mutual fund': 'Motilal',
    }

    new_rows = []
    current_amc = None

    for line in data.splitlines():
        line = line.strip()
        if not line:
            continue  # blank lines are separators — do NOT reset current_amc
        if ';' not in line:
            low = line.lower()
            matched = None
            for k, v in amc_map.items():
                if k in low:
                    matched = v
                    break
            current_amc = matched
        elif current_amc:
            parts = line.split(';')
            if len(parts) < 5:
                continue
            code  = parts[0].strip()
            isin  = parts[1].strip()
            name  = parts[3].strip()
            if not is_equity_scheme(name):
                continue
            new_rows.append({
                'scheme_code': code,
                'scheme_name': name,
                'amc':         current_amc,
                'isin':        isin if isin != '-' else '',
            })

    df_new = pd.DataFrame(new_rows)
    print(f"  New equity/growth schemes: {len(df_new)}")
    if df_new.empty:
        print("  ERROR: no new schemes found — check AMFI parse logic")
        return df_new
    print(df_new.groupby('amc').size())

    # Cast scheme_code to match existing dtype (int64)
    df_new['scheme_code'] = pd.to_numeric(df_new['scheme_code'], errors='coerce').astype('Int64')

    # TER defaults
    TER_DEFAULT = {
        'Quant': 0.55, 'UTI': 0.65, 'Bandhan': 0.60, 'Tata': 0.60, 'Motilal': 0.55,
    }
    # Proxy defaults — use AMFI Nifty 500 index scheme (code 147625)
    NIFTY500_CODE = 147625

    df_new['category_amfi'] = 'Equity'
    df_new['category']      = 'equity'
    df_new['ter_est']       = df_new['amc'].map(TER_DEFAULT).fillna(0.60)
    df_new['proxy_code']    = pd.array([NIFTY500_CODE] * len(df_new), dtype='Int64')
    df_new['proxy_reason']  = 'Nifty 500 (default)'

    # Remove duplicates with existing
    existing_codes = set(fm['scheme_code'].astype(str))
    df_new = df_new[~df_new['scheme_code'].astype(str).isin(existing_codes)]
    print(f"  After dedup: {len(df_new)} new schemes to add")

    # Append to fund_meta and scheme_list
    fm_new = pd.concat([fm, df_new[fm.columns]], ignore_index=True)
    sl_new_rows = df_new[['scheme_code', 'scheme_name', 'amc', 'isin']].copy()
    sl_new = pd.concat([sl, sl_new_rows], ignore_index=True)

    fm_new.to_parquet(fm_path, index=False)
    sl_new.to_parquet(sl_path, index=False)
    print(f"  fund_meta: {len(fm_new)} schemes ({len(df_new)} added)")
    print(f"  scheme_list: {len(sl_new)} schemes")
    return df_new


# ─────────────────────────────────────────────────────────────────────────────
# ENTRY POINT
# ─────────────────────────────────────────────────────────────────────────────

if __name__ == '__main__':
    import argparse
    parser = argparse.ArgumentParser()
    parser.add_argument('--amc', nargs='+',
                        choices=['quant', 'uti', 'tata', 'motilal', 'bandhan', 'all', 'meta'],
                        default=['all'], help='Which AMCs to ingest')
    parser.add_argument('--dry-run', action='store_true',
                        help='Download and parse but do not write parquets')
    parser.add_argument('--from-period', default='2023-01',
                        help='Only process periods >= this (YYYY-MM)')
    parser.add_argument('--to-period', default='2099-12',
                        help='Only process periods <= this (YYYY-MM)')
    args = parser.parse_args()

    do_all = 'all' in args.amc

    # Apply period filter to URL lists
    def _filter_urls(url_list):
        return [(p, u) for p, u in url_list
                if args.from_period <= p <= args.to_period]

    # Monkey-patch URL lists with period filter
    orig_quant   = list(QUANT_URLS)
    orig_uti     = list(UTI_URLS)
    orig_tata    = list(TATA_URLS)
    orig_motilal = list(MOTILAL_URLS)

    if args.from_period != '2023-01' or args.to_period != '2099-12':
        QUANT_URLS[:]   = _filter_urls(orig_quant)
        UTI_URLS[:]     = _filter_urls(orig_uti)
        TATA_URLS[:]    = _filter_urls(orig_tata)
        MOTILAL_URLS[:] = _filter_urls(orig_motilal)
        print(f"Period filter: {args.from_period} – {args.to_period}")
        print(f"  Quant: {len(QUANT_URLS)}, UTI: {len(UTI_URLS)}, "
              f"Tata: {len(TATA_URLS)}, Motilal: {len(MOTILAL_URLS)} periods")

    print("Building AMFI lookup...")
    name_lk, isin_lk = build_amfi_lookup()

    if 'meta' in args.amc or do_all:
        extend_fund_meta(name_lk)

    if 'quant' in args.amc or do_all:
        ingest_quant(name_lk, isin_lk, dry_run=args.dry_run)

    if 'uti' in args.amc or do_all:
        ingest_uti(name_lk, isin_lk, dry_run=args.dry_run)

    if 'tata' in args.amc or do_all:
        ingest_tata(name_lk, isin_lk, dry_run=args.dry_run)

    if 'motilal' in args.amc or do_all:
        ingest_motilal(name_lk, isin_lk, dry_run=args.dry_run)

    if 'bandhan' in args.amc or do_all:
        ingest_bandhan(name_lk, isin_lk, dry_run=args.dry_run)

    print("\nDone.")
