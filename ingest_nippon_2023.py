#!/usr/bin/env python3
"""
ingest_nippon_2023.py
=====================
Backfill Nippon India 2023 monthly holdings (Jan–Dec 2023, minus Jun)
directly from nipponindiaim.com CDN URLs into the holdings parquet cache.

Nippon 2023 files are single multi-sheet XLSX workbooks:
  - INDEX sheet: col A = sheet code, col B = INDIRECT formula (unusable)
  - Fund sheets: row 0 col 1 = full scheme name
                 row 3 col 1 = "ISIN", col 2 = "Name of the Instrument", col 6 = "% to NAV"
                 row 6+ = holding rows (equity ISINs in col 1, pct_nav decimal in col 6)

pct_nav is stored as a decimal (0.0478 = 4.78%) → multiply by 100.

Usage:
    python3 ingest_nippon_2023.py --mf-data ./mf_data
    python3 ingest_nippon_2023.py --mf-data ./mf_data --force
    python3 ingest_nippon_2023.py --mf-data ./mf_data --month 2023-11
"""
from __future__ import annotations

import argparse
import io
import logging
import subprocess
import sys
from pathlib import Path

import openpyxl
import pandas as pd

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s  %(levelname)-8s  %(message)s",
    datefmt="%H:%M:%S",
)
log = logging.getLogger("nippon2023")

UA = (
    "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) "
    "AppleWebKit/537.36 (KHTML, like Gecko) Chrome/124.0.0.0 Safari/537.36"
)

# ── URLs (from AdvisorKhoj scrape of Nippon-India-Mutual-Fund) ───────────────
# June 2023 is absent from AdvisorKhoj — Nippon didn't post it there
NIPPON_2023_URLS: dict[str, str] = {
    "2023-01": "https://mf.nipponindiaim.com/InvestorServices/FactsheetsDocuments/MONTHLY-PORTFOLIO-JAN-23.xls",
    "2023-02": "https://mf.nipponindiaim.com/InvestorServices/FactsheetsDocuments/MONTHLY-PORTFOLIO-FEB-23.xls",
    "2023-03": "https://mf.nipponindiaim.com/InvestorServices/FactsheetsDocuments/MONTHLY-PORTFOLIO-MAR-23.xls",
    "2023-04": "https://mf.nipponindiaim.com/InvestorServices/FactsheetsDocuments/MONTHLY-PORTFOLIO-APR-23.xls",
    "2023-05": "https://mf.nipponindiaim.com/InvestorServices/FactsheetsDocuments/MONTHLY-PORTFOLIO-MAY-2023.xls",
    # 2023-06: not available on AdvisorKhoj
    "2023-07": "https://mf.nipponindiaim.com/InvestorServices/FactsheetsDocuments/MONTHLY-PORTFOLIO-JULY-2023.xls",
    "2023-08": "https://mf.nipponindiaim.com/InvestorServices/FactsheetsDocuments/MONTHLY-PORTFOLIO-AUGUST-2023.xls",
    "2023-09": "https://mf.nipponindiaim.com/InvestorServices/FactsheetsDocuments/NIMF-MONTHLY-PORTFOLIO-Sep-23.xls",
    "2023-10": "https://mf.nipponindiaim.com/InvestorServices/FactsheetsDocuments/MONTHLY-PORTFOLIO-OCTOBER-2023.xls",
    "2023-11": "https://mf.nipponindiaim.com/InvestorServices/FactsheetsDocuments/MONTHLY-PORTFOLIO-NOV-23.xls",
    "2023-12": "https://mf.nipponindiaim.com/InvestorServices/FactsheetsDocuments/MONTHLY-PORTFOLIO-DEC-23.xls",
}

COLS = ["isin", "stock_name", "pct_nav", "scheme_name", "_sheet", "amc", "scheme_code", "as_of_date"]

# Equity-focused scheme keywords to keep (skip debt/gilt/liquid/overnight funds)
SKIP_KEYWORDS = [
    "liquid", "gilt", "overnight", "money market", "ultra short",
    "low duration", "short term", "credit risk", "corporate bond",
    "floating rate", "dynamic bond", "banking and psu", "arbitrage",
    "hybrid bond", "income fund", "debt", "fixed maturity",
]


def _download(url: str) -> bytes | None:
    cmd = [
        "curl", "-s", "-L", "--max-time", "60",
        "-A", UA,
        "-H", "Accept-Language: en-IN,en;q=0.9",
        "--compressed",
        url,
    ]
    r = subprocess.run(cmd, capture_output=True, timeout=65)
    if r.returncode != 0 or len(r.stdout) < 1000:
        return None
    # Nippon serves xlsx bytes at .xls URLs — check magic
    data = r.stdout
    if data[:4] not in (b'PK\x03\x04', b'\xd0\xcf\x11\xe0'):
        return None
    return data


def _is_equity_scheme(name: str) -> bool:
    nl = name.lower()
    return not any(kw in nl for kw in SKIP_KEYWORDS)


def _parse_nippon_xlsx(data: bytes, month_str: str) -> pd.DataFrame:
    """
    Parse a Nippon monthly portfolio XLSX (multi-sheet, INDEX + fund sheets).

    Fund sheet layout:
      row 0: (internal_code, FULL SCHEME NAME, ...)
      row 1: (None, "Monthly Portfolio Statement as on …", ...)
      row 3: (None, "ISIN", "Name of the Instrument", "Industry/Rating",
              "Quantity", "Market Value", "% to NAV", ...)
      row 6+: data rows — col 1 = ISIN, col 2 = stock name, col 6 = pct_nav (decimal)

    We skip the INDEX sheet and read scheme names directly from each fund sheet row 0.
    """
    # Try xlsx first (most Nippon files are xlsx-in-disguise)
    if data[:4] == b'PK\x03\x04':
        try:
            wb = openpyxl.load_workbook(io.BytesIO(data), data_only=True, read_only=True)
            return _parse_wb_openpyxl(wb, month_str)
        except Exception as e:
            log.warning(f"  openpyxl failed ({e}), trying xlrd")

    # Fall back to xlrd for true OLE2 .xls
    try:
        import xlrd
        wb_xls = xlrd.open_workbook(file_contents=data)
        return _parse_wb_xlrd(wb_xls, month_str)
    except Exception as e:
        log.warning(f"  xlrd also failed: {e}")
        return pd.DataFrame()


def _is_equity_isin(isin: str) -> bool:
    """Equity ISINs: 12 chars, start with IN, 3rd char is E (equity) or F (ETF)."""
    if not isinstance(isin, str) or len(isin) != 12:
        return False
    if not isin.startswith("IN"):
        return False
    # Debt ISINs: IN followed by digits or 'E' with bond series
    # Equity ISINs: typically INE + 6 alphanums + 3 digits
    # Accept any 12-char IN-starting ISIN that isn't purely numeric after IN
    return True  # let pct_nav filter do the heavy lifting


def _parse_wb_openpyxl(wb: openpyxl.Workbook, month_str: str) -> pd.DataFrame:
    records = []
    for sn in wb.sheetnames:
        if sn.upper() == "INDEX":
            continue
        ws = wb[sn]
        rows = list(ws.iter_rows(values_only=True))
        if len(rows) < 5:
            continue

        # Scheme name from row 0, col 1 (not col 0 which is internal code)
        raw_name = rows[0][1] if len(rows[0]) > 1 else None
        if not raw_name or str(raw_name).startswith("="):
            # Fallback: try col 0
            raw_name = rows[0][0] if rows[0] else sn
        scheme_name = str(raw_name).strip()[:150]

        if not _is_equity_scheme(scheme_name):
            continue

        # Holdings start at row 5 or 6 (skip header rows 0-4)
        for row in rows[5:]:
            if not row or len(row) < 7:
                continue
            isin = row[1]  # ISIN in col 1
            if not isinstance(isin, str):
                continue
            isin = isin.strip()
            if not isin.startswith("IN") or len(isin) != 12:
                continue

            stock_name = str(row[2]).strip() if row[2] else ""
            if not stock_name or stock_name.lower().startswith(("sub total", "total", "grand")):
                continue

            try:
                pct_decimal = float(row[6])
            except (TypeError, ValueError):
                continue

            # pct_nav stored as decimal: 0.0478 = 4.78%
            if not (0.0001 < pct_decimal <= 1.05):
                continue
            pct_val = round(pct_decimal * 100, 4)

            records.append({
                "isin": isin,
                "stock_name": stock_name,
                "pct_nav": pct_val,
                "scheme_name": scheme_name,
                "_sheet": sn,
                "amc": "Nippon",
                "scheme_code": pd.NA,
                "as_of_date": month_str,
            })

    return pd.DataFrame(records) if records else pd.DataFrame()


def _parse_wb_xlrd(wb, month_str: str) -> pd.DataFrame:
    import xlrd
    records = []
    for sn in wb.sheet_names():
        if sn.upper() == "INDEX":
            continue
        ws = wb.sheet_by_name(sn)
        if ws.nrows < 5:
            continue

        raw_name = ws.cell_value(0, 1) if ws.ncols > 1 else ws.cell_value(0, 0)
        scheme_name = str(raw_name).strip()[:150]

        if not _is_equity_scheme(scheme_name):
            continue

        for r in range(5, ws.nrows):
            if ws.ncols < 7:
                continue
            isin = str(ws.cell_value(r, 1)).strip()
            if not isin.startswith("IN") or len(isin) != 12:
                continue

            stock_name = str(ws.cell_value(r, 2)).strip()
            if not stock_name or stock_name.lower().startswith(("sub total", "total")):
                continue

            try:
                pct_decimal = float(ws.cell_value(r, 6))
            except (TypeError, ValueError):
                continue

            if not (0.0001 < pct_decimal <= 1.05):
                continue
            pct_val = round(pct_decimal * 100, 4)

            records.append({
                "isin": isin,
                "stock_name": stock_name,
                "pct_nav": pct_val,
                "scheme_name": scheme_name,
                "_sheet": sn,
                "amc": "Nippon",
                "scheme_code": pd.NA,
                "as_of_date": month_str,
            })

    return pd.DataFrame(records) if records else pd.DataFrame()


def _attach_codes(df: pd.DataFrame, scheme_df: pd.DataFrame) -> pd.DataFrame:
    try:
        from rapidfuzz import process, fuzz
    except ImportError:
        return df

    known = scheme_df["scheme_name"].tolist()
    code_map = dict(zip(scheme_df["scheme_name"], scheme_df["scheme_code"]))

    results = {}
    for sname in df["scheme_name"].unique():
        match, score, _ = process.extractOne(sname, known, scorer=fuzz.token_sort_ratio)
        results[sname] = code_map.get(match, pd.NA) if score >= 72 else pd.NA

    df["scheme_code"] = df["scheme_name"].map(results)
    return df


def _merge_month(month_str: str, new_df: pd.DataFrame, hold_dir: Path) -> None:
    out_path = hold_dir / f"{month_str}.parquet"
    if out_path.exists():
        existing = pd.read_parquet(out_path)
        existing = existing[existing["amc"] != "Nippon"]
    else:
        existing = pd.DataFrame(columns=COLS)

    for c in COLS:
        if c not in existing.columns:
            existing[c] = pd.NA
        if c not in new_df.columns:
            new_df[c] = pd.NA

    combined = pd.concat([existing[COLS], new_df[COLS]], ignore_index=True)
    combined.to_parquet(out_path, index=False)
    n_matched = combined["scheme_code"].notna().sum()
    log.info(f"  {month_str}: saved {len(combined):,} total rows  "
             f"{n_matched/len(combined)*100:.0f}% matched")


def main():
    p = argparse.ArgumentParser()
    p.add_argument("--mf-data", default="./mf_data")
    p.add_argument("--force", action="store_true")
    p.add_argument("--month", default=None)
    args = p.parse_args()

    mf_data = Path(args.mf_data)
    hold_dir = mf_data / "holdings"
    hold_dir.mkdir(parents=True, exist_ok=True)

    scheme_path = mf_data / "scheme_list.parquet"
    scheme_df = pd.read_parquet(scheme_path) if scheme_path.exists() else None

    months = list(NIPPON_2023_URLS.keys())
    if args.month:
        months = [m for m in months if m == args.month]

    total_ok = 0
    for month_str in sorted(months):
        out_path = hold_dir / f"{month_str}.parquet"

        if not args.force and out_path.exists():
            existing = pd.read_parquet(out_path)
            if "Nippon" in existing.get("amc", pd.Series([])).values:
                nrows = len(existing[existing["amc"] == "Nippon"])
                log.info(f"  {month_str}: Nippon already present ({nrows} rows) — skipping")
                continue

        url = NIPPON_2023_URLS[month_str]
        log.info(f"\n{month_str}: downloading …")
        data = _download(url)
        if not data:
            log.warning(f"  {month_str}: download failed")
            continue

        df = _parse_nippon_xlsx(data, month_str)
        if df.empty:
            log.warning(f"  {month_str}: no equity records parsed")
            continue

        log.info(f"  {month_str}: {len(df)} equity rows, {df['scheme_name'].nunique()} schemes")

        if scheme_df is not None:
            df = _attach_codes(df, scheme_df)
            df["scheme_code"] = pd.to_numeric(df["scheme_code"], errors="coerce").astype("Int64")

        _merge_month(month_str, df, hold_dir)
        total_ok += 1

    log.info(f"\nDone: {total_ok}/{len(months)} months ingested")

    # Summary
    import os
    files = sorted(f for f in os.listdir(hold_dir) if f.endswith(".parquet"))
    all_rows = []
    for f in files:
        df_f = pd.read_parquet(hold_dir / f)
        month = f.replace(".parquet", "")
        if "amc" in df_f.columns:
            for amc in df_f["amc"].dropna().unique():
                all_rows.append({"month": month, "amc": amc})
    cov = pd.DataFrame(all_rows)
    print("\n=== Holdings coverage ===")
    print(cov.groupby("amc")["month"].agg(["min", "max", "count"])
          .sort_values("count", ascending=False).to_string())


if __name__ == "__main__":
    main()
