#!/usr/bin/env python3
"""
ingest_axis_2023.py
===================
Backfill Axis Mutual Fund 2023 monthly holdings (Jan–Dec 2023)
from axismf.com CDN links (via AdvisorKhoj) into the holdings parquet cache.

Axis 2023 files are single multi-sheet XLSX workbooks:
  - Index sheet: row 0 = header, row 1+ = (Sr No., Short Name, Scheme Name)
  - Fund sheets:
      row 0: (short_code, full_scheme_name, ...)
      row 3: (None, Name of Instrument, ISIN, Industry, Qty, Mkt Value, % to Net Assets)
      row 6+: (internal_code, stock_name, ISIN, industry, qty, mkt_val, pct_nav_decimal)

ISIN is col 2 (not col 1 like Nippon!), pct_nav is col 6 as decimal × 100.

Usage:
    python3 ingest_axis_2023.py --mf-data ./mf_data
    python3 ingest_axis_2023.py --mf-data ./mf_data --force
    python3 ingest_axis_2023.py --mf-data ./mf_data --month 2023-06
"""
from __future__ import annotations

import argparse
import io
import logging
import subprocess
import sys
from pathlib import Path

import openpyxl
import pandas as pd

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s  %(levelname)-8s  %(message)s",
    datefmt="%H:%M:%S",
)
log = logging.getLogger("axis2023")

UA = (
    "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) "
    "AppleWebKit/537.36 (KHTML, like Gecko) Chrome/124.0.0.0 Safari/537.36"
)

COLS = ["isin", "stock_name", "pct_nav", "scheme_name", "_sheet", "amc", "scheme_code", "as_of_date"]

SKIP_KEYWORDS = [
    "liquid", "gilt", "overnight", "money market", "ultra short", "low duration",
    "short term", "credit risk", "corporate bond", "floating rate", "dynamic bond",
    "banking and psu", "arbitrage", "fixed term", "capital builder", "conservative",
    "debt fund", "income fund", "treasury", "constant maturity",
]


def _get_axis_urls() -> dict[str, str]:
    """Scrape current Axis 2023 URLs from AdvisorKhoj."""
    import sys
    sys.path.insert(0, str(Path(__file__).parent))
    try:
        import logging as _l; _l.getLogger("backfill_holdings").setLevel(_l.WARNING)
        from backfill_holdings import _ak_url_map
        url_map = _ak_url_map("Axis-Mutual-Fund")
        return {k: v for k, v in url_map.items() if "2023" in k}
    except Exception as e:
        log.warning(f"AdvisorKhoj scrape failed: {e} — using hardcoded URLs")
        return HARDCODED_URLS


# Hardcoded fallback (scraped 2026-05-25) — in case AdvisorKhoj changes
HARDCODED_URLS: dict[str, str] = {
    "2023-01": "https://www.axismf.com/cms/sites/default/files/Statutory/Monthly%20Portfolio-31.01.2023.xls",
    "2023-02": "https://www.axismf.com/cms/sites/default/files/Statutory/Monthly%20Portfolio-28.02.2023.xls",
    "2023-03": "https://www.axismf.com/cms/sites/default/files/Statutory/Monthly%20Portfolio%20-%20Consolidated%20-%2031%20March%202023.xls",
    "2023-05": "https://www.axismf.com/cms/sites/default/files/Statutory/Monthly%20Portfolio%20as%20on%2031%20May%202023.xls",
    "2023-06": "https://www.axismf.com/cms/sites/default/files/Statutory/Monthly%20Portfolio%20as%20on%20June%2030%202023.xls",
    "2023-07": "https://www.axismf.com/cms/sites/default/files/Statutory/Monthly%20Portfolio%20as%20on%2031%20July%202023.xls",
    "2023-08": "https://www.axismf.com/cms/sites/default/files/Statutory/Monthly%20Portfolio%20as%20on%20Aug%2031%20%202023.xls",
    "2023-09": "https://www.axismf.com/cms/sites/default/files/Statutory/Monthly%20Portfolio%20as%20on%2030.09.2023.xls",
    "2023-10": "https://www.axismf.com/cms/sites/default/files/Statutory/Monthly%20Portfolio%20-%2031%20Oct%202023.xls",
    "2023-11": "https://www.axismf.com/cms/sites/default/files/Statutory/Monthly%20Portfolio%20as%20on%20Nov%2030%20%202023.xls",
    "2023-12": "https://www.axismf.com/cms/sites/default/files/Statutory/Monthly%20Portfolio%20as%20on%20Dec%2031%20%202023.xlsx",
}


def _download(url: str) -> bytes | None:
    cmd = [
        "curl", "-s", "-L", "--max-time", "60",
        "-A", UA,
        "-H", "Accept-Language: en-IN,en;q=0.9",
        "--compressed", url,
    ]
    r = subprocess.run(cmd, capture_output=True, timeout=65)
    if r.returncode != 0 or len(r.stdout) < 1000:
        return None
    data = r.stdout
    if data[:4] not in (b'PK\x03\x04', b'\xd0\xcf\x11\xe0'):
        return None
    return data


def _is_equity_scheme(name: str) -> bool:
    nl = name.lower()
    return not any(kw in nl for kw in SKIP_KEYWORDS)


def _parse_axis_xlsx(data: bytes, month_str: str) -> pd.DataFrame:
    """
    Parse a single Axis multi-sheet monthly portfolio file (XLSX or XLS OLE2).

    Index sheet: col 1 = short_code, col 2 = full scheme name
    Fund sheet layout:
      row 0: (short_code, full_scheme_name, ...)
      row 3: headers — col 1 = Name, col 2 = ISIN, col 6 = % to Net Assets
      row 6+: data — col 1 = stock_name, col 2 = ISIN, col 6 = pct_nav (decimal)
    """
    # OLE2 .xls path (Oct/Nov 2023 are true BIFF files)
    if data[:4] == b'\xd0\xcf\x11\xe0':
        return _parse_axis_xlrd(data, month_str)

    try:
        wb = openpyxl.load_workbook(io.BytesIO(data), data_only=True, read_only=True)
    except Exception as e:
        log.warning(f"  openpyxl failed: {e}")
        return pd.DataFrame()

    # Build code → scheme_name from Index sheet (literal values, not formulas)
    code_to_name: dict[str, str] = {}
    idx_sn = next((s for s in wb.sheetnames if s.lower() == "index"), None)
    if idx_sn:
        idx_ws = wb[idx_sn]
        for row in idx_ws.iter_rows(values_only=True):
            # row: (Sr No., Short Name, Scheme Name, ...)
            if len(row) >= 3 and row[1] and row[2]:
                code = str(row[1]).strip()
                name = str(row[2]).strip()
                if code and name and code.lower() != "short name":
                    code_to_name[code] = name[:150]

    records = []
    for sn in wb.sheetnames:
        if sn.lower() == "index":
            continue

        # Prefer Index lookup; fallback to row 0 col 1 of the fund sheet
        scheme_name = code_to_name.get(sn)
        if not scheme_name:
            ws = wb[sn]
            first_rows = list(ws.iter_rows(values_only=True, max_row=1))
            if first_rows and len(first_rows[0]) > 1 and first_rows[0][1]:
                scheme_name = str(first_rows[0][1]).strip()[:150]
            else:
                scheme_name = sn

        if not _is_equity_scheme(scheme_name):
            continue

        ws = wb[sn]
        rows = list(ws.iter_rows(values_only=True))

        # Data rows start after the header block (rows 0–5)
        for row in rows[6:]:
            if not row or len(row) < 7:
                continue

            # ISIN in col 2
            isin = row[2]
            if not isinstance(isin, str):
                continue
            isin = isin.strip()
            if len(isin) != 12 or not isin.startswith("IN"):
                continue

            # Stock name in col 1
            stock_name = str(row[1]).strip() if row[1] else ""
            if not stock_name:
                continue
            sl = stock_name.lower()
            if sl.startswith(("sub total", "total", "grand", "equity", "debt",
                               "(a)", "(b)", "listed", "privately")):
                continue

            # pct_nav in col 6 as decimal (0.0908 = 9.08%)
            try:
                pct_decimal = float(row[6])
            except (TypeError, ValueError):
                continue
            if not (0.0001 < pct_decimal <= 1.05):
                continue
            pct_val = round(pct_decimal * 100, 4)

            records.append({
                "isin": isin,
                "stock_name": stock_name,
                "pct_nav": pct_val,
                "scheme_name": scheme_name,
                "_sheet": sn,
                "amc": "Axis",
                "scheme_code": pd.NA,
                "as_of_date": month_str,
            })

    return pd.DataFrame(records) if records else pd.DataFrame()


def _parse_axis_xlrd(data: bytes, month_str: str) -> pd.DataFrame:
    """xlrd path for OLE2 .xls Axis files (Oct/Nov 2023)."""
    import xlrd
    try:
        wb = xlrd.open_workbook(file_contents=data)
    except Exception as e:
        log.warning(f"  xlrd failed: {e}")
        return pd.DataFrame()

    # Build code→name from Index
    code_to_name: dict[str, str] = {}
    if "Index" in wb.sheet_names():
        ws_i = wb.sheet_by_name("Index")
        for r in range(1, ws_i.nrows):  # skip header row 0
            if ws_i.ncols >= 3:
                code = str(ws_i.cell_value(r, 1)).strip()
                name = str(ws_i.cell_value(r, 2)).strip()
                if code and name:
                    code_to_name[code] = name[:150]

    records = []
    for sn in wb.sheet_names():
        if sn.lower() == "index":
            continue

        scheme_name = code_to_name.get(sn)
        if not scheme_name:
            ws = wb.sheet_by_name(sn)
            scheme_name = str(ws.cell_value(0, 1)).strip()[:150] if ws.ncols > 1 else sn

        if not _is_equity_scheme(scheme_name):
            continue

        ws = wb.sheet_by_name(sn)
        for r in range(6, ws.nrows):
            if ws.ncols < 7:
                continue
            isin = str(ws.cell_value(r, 2)).strip()
            if len(isin) != 12 or not isin.startswith("IN"):
                continue
            stock_name = str(ws.cell_value(r, 1)).strip()
            if not stock_name:
                continue
            sl = stock_name.lower()
            if sl.startswith(("sub total", "total", "grand", "equity", "debt",
                               "(a)", "(b)", "listed", "privately")):
                continue
            try:
                pct_decimal = float(ws.cell_value(r, 6))
            except (TypeError, ValueError):
                continue
            if not (0.0001 < pct_decimal <= 1.05):
                continue
            records.append({
                "isin": isin,
                "stock_name": stock_name,
                "pct_nav": round(pct_decimal * 100, 4),
                "scheme_name": scheme_name,
                "_sheet": sn,
                "amc": "Axis",
                "scheme_code": pd.NA,
                "as_of_date": month_str,
            })

    return pd.DataFrame(records) if records else pd.DataFrame()


def _attach_codes(df: pd.DataFrame, scheme_df: pd.DataFrame) -> pd.DataFrame:
    try:
        from rapidfuzz import process, fuzz
    except ImportError:
        return df

    known = scheme_df["scheme_name"].tolist()
    code_map = dict(zip(scheme_df["scheme_name"], scheme_df["scheme_code"]))
    results = {}
    for sname in df["scheme_name"].unique():
        match, score, _ = process.extractOne(sname, known, scorer=fuzz.token_sort_ratio)
        results[sname] = code_map.get(match, pd.NA) if score >= 72 else pd.NA
    df["scheme_code"] = df["scheme_name"].map(results)
    return df


def _merge_month(month_str: str, new_df: pd.DataFrame, hold_dir: Path) -> None:
    out_path = hold_dir / f"{month_str}.parquet"
    if out_path.exists():
        existing = pd.read_parquet(out_path)
        existing = existing[existing["amc"] != "Axis"]
    else:
        existing = pd.DataFrame(columns=COLS)

    for c in COLS:
        if c not in existing.columns:
            existing[c] = pd.NA
        if c not in new_df.columns:
            new_df[c] = pd.NA

    combined = pd.concat([existing[COLS], new_df[COLS]], ignore_index=True)
    combined.to_parquet(out_path, index=False)
    n_matched = combined["scheme_code"].notna().sum()
    log.info(f"  {month_str}: {len(new_df)} Axis rows → {len(combined):,} total  "
             f"{n_matched/len(combined)*100:.0f}% matched")


def main():
    p = argparse.ArgumentParser()
    p.add_argument("--mf-data", default="./mf_data")
    p.add_argument("--force", action="store_true")
    p.add_argument("--month", default=None)
    args = p.parse_args()

    mf_data = Path(args.mf_data)
    hold_dir = mf_data / "holdings"
    hold_dir.mkdir(parents=True, exist_ok=True)

    scheme_path = mf_data / "scheme_list.parquet"
    scheme_df = pd.read_parquet(scheme_path) if scheme_path.exists() else None

    url_map = _get_axis_urls()
    log.info(f"Axis 2023 URLs available: {sorted(url_map.keys())}")

    months = sorted(url_map.keys())
    if args.month:
        months = [m for m in months if m == args.month]

    total_ok = 0
    for month_str in months:
        out_path = hold_dir / f"{month_str}.parquet"

        if not args.force and out_path.exists():
            existing = pd.read_parquet(out_path)
            if "Axis" in existing.get("amc", pd.Series([])).values:
                ax_rows = len(existing[existing["amc"] == "Axis"])
                log.info(f"  {month_str}: Axis already present ({ax_rows} rows) — skipping")
                continue

        url = url_map[month_str]
        log.info(f"\n{month_str}: {url.split('/')[-1]}")
        data = _download(url)
        if not data:
            log.warning(f"  {month_str}: download failed")
            continue

        df = _parse_axis_xlsx(data, month_str)
        if df.empty:
            log.warning(f"  {month_str}: no equity records parsed")
            continue

        log.info(f"  {month_str}: {len(df)} equity rows, {df['scheme_name'].nunique()} schemes")

        if scheme_df is not None:
            df = _attach_codes(df, scheme_df)
            df["scheme_code"] = pd.to_numeric(df["scheme_code"], errors="coerce").astype("Int64")

        _merge_month(month_str, df, hold_dir)
        total_ok += 1

    log.info(f"\nDone: {total_ok}/{len(months)} months ingested")

    import os
    files = sorted(f for f in os.listdir(hold_dir) if f.endswith(".parquet"))
    all_rows = []
    for f in files:
        df_f = pd.read_parquet(hold_dir / f)
        month = f.replace(".parquet", "")
        if "amc" in df_f.columns:
            for amc in df_f["amc"].dropna().unique():
                all_rows.append({"month": month, "amc": amc})
    cov = pd.DataFrame(all_rows)
    print("\n=== Holdings coverage ===")
    print(cov.groupby("amc")["month"].agg(["min", "max", "count"])
          .sort_values("count", ascending=False).to_string())


if __name__ == "__main__":
    main()
