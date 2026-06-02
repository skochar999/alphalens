"""
Parse SBI monthly all-schemes xlsx files and merge into holdings/{YYYY-MM}.parquet
Format: col[2]=Name, col[3]=ISIN, col[4]=Industry, col[5]=Qty, col[6]=MktVal, col[7]=% AUM
pct_nav col is percentage (10.24), must divide by 100 for decimal fraction.
"""
import re, os
import pandas as pd
import openpyxl
from pathlib import Path

ISIN_RE = re.compile(r'^IN[A-Z0-9]{10}$')

RAW_DIR  = Path("/sessions/admiring-nifty-dijkstra/mnt/outputs/mf_data/holdings_raw/SBI")
HOLD_DIR = Path("/sessions/admiring-nifty-dijkstra/mnt/outputs/mf_data/holdings")

# sheet_name → (AMFI scheme_code, scheme_name)
# Only equity/hybrid funds we care about
SHEET_MAP = {
    "SBLUECHIP": (119598, "SBI Large Cap FUND-DIRECT PLAN -GROWTH"),
    "SEHF":      (119609, "SBI EQUITY HYBRID FUND - DIRECT PLAN - Growth"),
    "SMIDCAP":   (119716, "SBI MIDCAP FUND - DIRECT PLAN - GROWTH"),
    "SFLEXI":    (119718, "SBI Flexicap Fund - DIRECT PLAN - Growth Option"),
    "SLMF":      (119721, "SBI LARGE & MIDCAP FUND -DIRECT PLAN -Growth"),
    "SLTEF":     (119723, "SBI ELSS Tax Saver FUND - DIRECT PLAN -GROWTH"),
    "SFEF":      (119727, "SBI FOCUSED FUND - DIRECT PLAN -GROWTH"),
    "SHOF":      (119783, "SBI HEALTHCARE OPPORTUNITIES FUND - DIRECT PLAN -GROWTH"),
    "SCF":       (119835, "SBI CONTRA FUND - DIRECT PLAN - GROWTH"),
    "SCOF":      (120575, "SBI CONSUMPTION OPPORTUNITIES FUND - DIRECT PLAN - GROWTH"),
    "STOF":      (120578, "SBI TECHNOLOGY OPPORTUNITIES FUND - DIRECT PLAN - GROWTH"),
    "SSCF":      (125497, "SBI Small Cap Fund - Direct Plan - Growth"),
    "SESF":      (134643, "SBI Equity Savings Fund - Direct Plan - Growth"),
    "SLTAF-III": (136007, "SBI Long Term Advantage Fund - Series III - Direct Plan - Growth"),
    "SLTAF-IV":  (140487, "SBI Long Term Advantage Fund - Series IV - Direct Plan - Growth"),
    "SLTAF-V":   (142138, "SBI Long Term Advantage Fund - Series V - Direct Plan - Growth"),
    "SLTAF-VI":  (143178, "SBI Long Term Advantage Fund - Series VI - Direct Plan - Growth"),
    "SEMVF":     (146643, "SBI Equity Minimum Variance Fund - Direct Plan - Growth"),
    "SBAF":      (149134, "SBI Balanced Advantage Fund - Direct Plan - Growth"),
    "SMCF":      (149882, "SBI Multicap Fund- Direct Plan- Growth option"),
    "SEOF":      (152417, "SBI Energy Opportunities Fund - Direct Plan - Growth"),
    "SRBF-AP":   (None,   "SBI Retirement Benefit Fund - Aggressive Plan - Direct"),
    "SRBF-AHP":  (None,   "SBI Retirement Benefit Fund - Aggressive Hybrid - Direct"),
    "SPSU":      (None,   "SBI PSU Fund - Direct Plan - Growth"),
    "SDYF":      (None,   "SBI Dividend Yield Fund - Direct Plan - Growth"),
    "SQF":       (None,   "SBI Quant Fund - Direct Plan - Growth"),
}

DEBT_KEYWORDS = ['DEBT','MONEY MARKET','CASH & CASH','NET ASSETS','GRAND TOTAL',
                 'CERTIFICATE','COMMERCIAL PAPER','TREASURY','GOVERNMENT SECURIT',
                 'REPO ','CBLO','TOTAL NET ASSETS','TOTAL ASSETS',
                 'REVERSE REPO','FIXED DEPOSIT','PREFERENCE SHARE']

def parse_sheet(ws, scheme_code, scheme_name):
    """Parse one equity fund sheet → list of holding dicts"""
    holdings = []
    in_equity = False

    for row in ws.iter_rows(values_only=True):
        if not row or len(row) < 8:
            continue

        c2 = str(row[2]).strip() if row[2] is not None else ""
        c3 = str(row[3]).strip() if row[3] is not None else ""

        # Detect equity section start
        if 'EQUITY' in c2.upper() and ('RELATED' in c2.upper() or 'EQUITY' in c2.upper()):
            if 'DEBT' not in c2.upper() and 'SAVING' not in c2.upper():
                in_equity = True
                continue

        if 'LISTED' in c2.upper() and 'AWAITING' in c2.upper():
            continue

        # Detect end of equity section
        if in_equity and c2:
            upper = c2.upper()
            if any(k in upper for k in DEBT_KEYWORDS):
                in_equity = False
                continue
            if re.match(r'^[B-Zb-z]\)', c2.strip()) or re.match(r'^\([B-Z]\)', c2.strip()):
                in_equity = False
                continue
            # "TOTAL OF EQUITY" or "TOTAL EQUITY" marks end
            if upper.startswith('TOTAL') and 'EQUITY' in upper:
                in_equity = False
                continue

        if not in_equity:
            continue

        # ISIN in col[3]
        if not ISIN_RE.match(c3):
            continue

        try:
            pct_raw = float(row[7]) if row[7] is not None else None
        except (ValueError, TypeError):
            continue

        if pct_raw is None or pct_raw <= 0:
            continue

        # SBI stores as percentage (10.24 = 10.24% NAV) → convert to fraction
        pct_nav = pct_raw / 100.0

        holdings.append({
            "isin":        c3,
            "stock_name":  c2,
            "pct_nav":     pct_nav,
            "sector":      str(row[4]).strip() if row[4] else "",
            "scheme_code": scheme_code,
            "scheme_name": scheme_name,
            "amc":         "SBI",
        })
    return holdings


def process_all():
    """Parse all SBI xlsx files → dict[date_str → list[dicts]]"""
    results = {}
    xlsx_files = sorted(RAW_DIR.glob("*.xlsx"))
    print(f"Processing {len(xlsx_files)} SBI xlsx files…")

    for path in xlsx_files:
        dt = path.stem  # YYYY-MM
        if not re.match(r'^\d{4}-\d{2}$', dt):
            continue
        try:
            wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
        except Exception as e:
            print(f"  ERR loading {path.name}: {e}")
            continue

        month_rows = []
        available_sheets = set(wb.sheetnames)
        for sheet_name, (sc, sn) in SHEET_MAP.items():
            if sheet_name not in available_sheets:
                continue
            ws = wb[sheet_name]
            rows = parse_sheet(ws, sc, sn)
            if rows:
                month_rows.extend(rows)

        wb.close()
        if month_rows:
            results[dt] = month_rows
            # Summary for a few key funds
            by_fund = {}
            for r in month_rows:
                by_fund.setdefault(r['scheme_name'][:35], []).append(r['pct_nav'])
            key_funds = ["SBI Large Cap","SBI Small Cap","SBI Contra","SBI Multi"]
            summary = []
            for f, navs in by_fund.items():
                if any(k.lower() in f.lower() for k in key_funds):
                    summary.append(f"{f[:25]}:{len(navs)}h/{sum(navs):.0%}")
            print(f"  {dt}: {len(month_rows)} rows across {len(by_fund)} funds  {' | '.join(summary[:4])}")

    return results


def merge_into_holdings(new_data):
    """Merge SBI rows into holdings parquets, replacing existing SBI rows."""
    for date_str, rows in sorted(new_data.items()):
        parq = HOLD_DIR / f"{date_str}.parquet"
        new_df = pd.DataFrame(rows)
        new_df["as_of_date"] = date_str
        new_df["_sheet"]     = "SBI"
        new_df["year"]       = float(date_str[:4])
        new_df["month"]      = float(date_str[5:7])
        new_df["scheme_code"]= pd.array(new_df["scheme_code"], dtype="Float64")
        keep = ["isin","stock_name","pct_nav","scheme_name","_sheet","amc",
                "scheme_code","as_of_date","year","month"]
        new_df = new_df[[c for c in keep if c in new_df.columns]]

        if parq.exists():
            existing = pd.read_parquet(parq)
            existing = existing[existing["amc"] != "SBI"]
            combined = pd.concat([existing, new_df], ignore_index=True)
        else:
            combined = new_df
        combined.to_parquet(parq, index=False)
    print(f"Merged SBI data into {len(new_data)} parquets")


if __name__ == "__main__":
    new_data = process_all()
    total = sum(len(v) for v in new_data.values())
    print(f"\nTotal: {total} rows across {len(new_data)} months")
    merge_into_holdings(new_data)
    print("Done.")
