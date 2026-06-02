#!/usr/bin/env python3
"""
debug_icici_absl.py
===================
Examine the raw structure of ICICI Pru and ABSL Excel files to diagnose
why the parsers return wrong scheme_names (ICICI) or 0 rows (ABSL).
"""
import io, subprocess, re, zipfile, sys
import openpyxl

UA = "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/124.0.0.0 Safari/537.36"

def curl(url, timeout=90):
    r = subprocess.run(
        ["curl", "-s", "-L", "--max-time", str(timeout), "-A", UA,
         "-H", "Accept-Language: en-IN,en;q=0.9", "--compressed", url],
        capture_output=True, timeout=timeout + 5
    )
    return r.stdout

# ─── ICICI ───────────────────────────────────────────────────────────────────
print("=" * 68)
print("[ICICI] Downloading ZIP …")
url = "https://www.icicipruamc.com/blob/downloads/Files/Monthly%20Portfolio%20Disclosures/2026/Apr/Monthly-Portfolio-Disclosure-April-2026.zip"
data = curl(url)
print(f"  Downloaded {len(data):,} bytes  magic={data[:4]}")

if data[:4] == b'PK\x03\x04':
    with zipfile.ZipFile(io.BytesIO(data)) as zf:
        names = [n for n in zf.namelist() if n.lower().endswith(('.xlsx','.xls')) and not n.startswith('__')]
        print(f"  {len(names)} Excel files in ZIP")
        print("  First 5 filenames:")
        for n in names[:5]:
            print(f"    {n}")

        # Examine first file in detail
        print("\n  === Detailed look at first file ===")
        raw = zf.read(names[0])
        wb = openpyxl.load_workbook(io.BytesIO(raw))
        print(f"  Sheets: {wb.sheetnames}")
        ws = wb[wb.sheetnames[0]]
        rows = list(ws.iter_rows(values_only=True))
        print(f"  Total rows: {len(rows)}")
        print("  First 15 rows:")
        for i, row in enumerate(rows[:15]):
            # show non-None cells
            cells = [(j, v) for j, v in enumerate(row) if v is not None]
            if cells:
                print(f"    row[{i:2d}]: {cells}")

        # Also check second file
        if len(names) > 1:
            print("\n  === Second file structure (rows 0-10) ===")
            print(f"  Filename: {names[1]}")
            raw2 = zf.read(names[1])
            wb2 = openpyxl.load_workbook(io.BytesIO(raw2))
            ws2 = wb2[wb2.sheetnames[0]]
            rows2 = list(ws2.iter_rows(values_only=True))
            for i, row in enumerate(rows2[:10]):
                cells = [(j, v) for j, v in enumerate(row) if v is not None]
                if cells:
                    print(f"    row[{i:2d}]: {cells}")

# ─── ABSL ─────────────────────────────────────────────────────────────────────
print("\n" + "=" * 68)
print("[ABSL] Trying download patterns …")

import calendar, datetime
now = datetime.date(2026, 4, 30)
y, m = now.year, now.month
last = str(calendar.monthrange(y, m)[1])
mname = now.strftime("%B")
mabbr = now.strftime("%b")
mlower = mname.lower()
mabbr_lower = mabbr.lower()

candidates = [
    f"monthly-disclosure-{mlower}-{last}-{y}.zip",
    f"monthly-disclosure-{mabbr_lower}-{last}-{y}.zip",
    f"sebi_monthly_portfolio-{last}-{mabbr_lower}-{y}.zip",
    f"monthly-portfolio-{mabbr_lower}-{y}.zip",
    f"monthly-disclosure-{last}-{mabbr_lower}-{y}.zip",
    f"portfolio-disclosure-{mabbr_lower}-{y}.zip",
]
base = f"https://mutualfund.adityabirlacapital.com/-/media/bsl/files/resources/monthly-portfolio/{y}/"

absl_data = None
for fname in candidates:
    url = base + fname
    print(f"  Trying: {fname} … ", end="", flush=True)
    d = curl(url, timeout=60)
    print(f"{len(d):,} bytes  magic={d[:4]}")
    if d[:4] == b'PK\x03\x04':
        absl_data = d
        print(f"  ✅ Found working URL: {url}")
        break

if absl_data:
    with zipfile.ZipFile(io.BytesIO(absl_data)) as zf:
        names = [n for n in zf.namelist() if not n.startswith('__')]
        print(f"\n  ZIP contains {len(names)} entries:")
        for n in names:
            print(f"    {n}")

        xlsx_names = [n for n in names if n.lower().endswith(('.xlsx','.xls'))]
        if xlsx_names:
            print(f"\n  === Examining first Excel: {xlsx_names[0]} ===")
            raw = zf.read(xlsx_names[0])
            print(f"  Size: {len(raw):,} bytes  magic={raw[:4]}")
            try:
                wb = openpyxl.load_workbook(io.BytesIO(raw))
                print(f"  Sheets: {wb.sheetnames}")
                for sn in wb.sheetnames[:3]:
                    ws = wb[sn]
                    rows = list(ws.iter_rows(values_only=True))
                    print(f"\n  Sheet '{sn}' — {len(rows)} rows × {ws.max_column} cols")
                    print(f"  First 20 rows:")
                    for i, row in enumerate(rows[:20]):
                        cells = [(j, v) for j, v in enumerate(row) if v is not None]
                        if cells:
                            print(f"    row[{i:2d}]: {cells[:6]}")  # first 6 non-None
            except Exception as e:
                print(f"  openpyxl failed: {e}")
                # Try xlrd
                try:
                    import xlrd
                    wb2 = xlrd.open_workbook(file_contents=raw)
                    print(f"  xlrd sheets: {wb2.sheet_names()}")
                    ws2 = wb2.sheet_by_index(0)
                    for i in range(min(20, ws2.nrows)):
                        row = [ws2.cell_value(i, j) for j in range(ws2.ncols)]
                        cells = [(j, v) for j, v in enumerate(row) if v]
                        if cells:
                            print(f"    row[{i:2d}]: {cells[:6]}")
                except Exception as e2:
                    print(f"  xlrd also failed: {e2}")
else:
    print("\n  ❌ All ABSL patterns failed — trying AdvisorKhoj for the URL …")
    ak_url = "https://www.advisorkhoj.com/form-download-centre/Mutual/Aditya-Birla-Sun-Life-Mutual-Fund/Monthly-Portfolio-Disclosures"
    html = curl(ak_url, timeout=30).decode('utf-8', errors='replace')
    urls = re.findall(r'https?://[^\s"\'<>\\]+(?:monthly|portfolio)[^\s"\'<>\\]+\.(?:zip|xlsx|xls)', html, re.I)
    absl_urls = [u for u in urls if 'adityabirla' in u.lower() or 'absl' in u.lower() or 'bsl' in u.lower()]
    print(f"  Found {len(absl_urls)} ABSL URLs on AdvisorKhoj:")
    for u in absl_urls[:5]:
        print(f"    {u}")

print("\nDone.")
