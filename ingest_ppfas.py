#!/usr/bin/env python3
"""
ingest_ppfas.py
===============
Download and ingest PPFAS (Parag Parikh) monthly portfolio disclosures
for all 36 months (May 2023 – Apr 2026) into the holdings cache.

URL pattern (verified):
  May–Jun 2023  : .xls
  Jul 2023      : .xls (both work; prefer .xls)
  Aug–Dec 2023  : .xlsx
  Jan–Dec 2024  : .xlsx
  Jan–Feb 2025  : .xls
  Mar–Apr 2025  : .xlsx
  May 2025–     : .xls

PPFAS file format:
  Sheet "Index"  : col 0=sr_no, col 1=short_code, col 2=scheme_name
  Fund sheets    : row 2=date, row 3=headers
                   col 1=stock_name, col 2=ISIN, col 3=industry
                   col 6=pct_nav  (DECIMAL: 0.0811 = 8.11%)

Usage:
    python3 ingest_ppfas.py --mf-data ./mf_data
    python3 ingest_ppfas.py --mf-data ./mf_data --force
    python3 ingest_ppfas.py --mf-data ./mf_data --month 2024-03
"""

from __future__ import annotations

import argparse
import calendar
import io
import logging
import re
import subprocess
import sys
from datetime import date, timedelta
from pathlib import Path
from typing import Optional

import pandas as pd

logging.basicConfig(level=logging.INFO, format="%(asctime)s  %(levelname)-8s  %(message)s",
                    datefmt="%H:%M:%S")
log = logging.getLogger("ppfas")

AMC_LABEL = "PPFAS"

# ─────────────────────────────────────────────────────────────────────────────
# PPFAS URL builder — returns (url, ext)
# ─────────────────────────────────────────────────────────────────────────────

def _ppfas_ext(year: int, month: int) -> str:
    """Return the correct file extension for the given month."""
    d = date(year, month, 1)
    # .xls periods (determined empirically):
    #   May–Jul 2023, Jan–Feb 2025, May 2025+
    if d >= date(2025, 5, 1):
        return ".xls"
    if d in [date(2025, 1, 1), date(2025, 2, 1)]:
        return ".xls"
    if date(2023, 5, 1) <= d <= date(2023, 7, 1):
        return ".xls"
    # Everything else: .xlsx
    return ".xlsx"


def _ppfas_url(year: int, month: int) -> str:
    last_day = calendar.monthrange(year, month)[1]
    month_name = date(year, month, 1).strftime("%B")  # e.g. "January"
    ext = _ppfas_ext(year, month)
    return (
        f"https://amc.ppfas.com/downloads/portfolio-disclosure/{year}/"
        f"PPFAS_Monthly_Portfolio_Report_{month_name}_{last_day}_{year}{ext}"
    )


def _all_ppfas_months(start: date, end: date) -> list[tuple[int, int]]:
    """Return list of (year, month) tuples in range [start, end]."""
    result = []
    d = start.replace(day=1)
    end_d = end.replace(day=1)
    while d <= end_d:
        result.append((d.year, d.month))
        # Advance one month
        if d.month == 12:
            d = date(d.year + 1, 1, 1)
        else:
            d = date(d.year, d.month + 1, 1)
    return result


# ─────────────────────────────────────────────────────────────────────────────
# HTTP downloader
# ─────────────────────────────────────────────────────────────────────────────

UA = ("Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) "
      "AppleWebKit/537.36 (KHTML, like Gecko) Chrome/124.0.0.0 Safari/537.36")


def _download(url: str, timeout: int = 60) -> Optional[bytes]:
    cmd = ["curl", "-s", "-L", "--max-time", str(timeout), "-A", UA,
           "-H", "Accept: application/octet-stream,*/*", url]
    try:
        r = subprocess.run(cmd, capture_output=True, timeout=timeout + 5)
        if r.returncode == 0 and len(r.stdout) > 1000:
            return r.stdout
    except Exception as e:
        log.warning(f"    Download error: {e}")
    return None


# ─────────────────────────────────────────────────────────────────────────────
# PPFAS parser
# ─────────────────────────────────────────────────────────────────────────────

ISIN_RE = re.compile(r'^IN[A-Z0-9]{10}$', re.IGNORECASE)
DATE_RE = re.compile(
    r'(?:as on|portfolio as of|statement as on)\s+([A-Za-z]+)[,\s]+(\d{4})',
    re.IGNORECASE
)
MONTH_MAP = {
    "january": 1, "february": 2, "march": 3, "april": 4,
    "may": 5, "june": 6, "july": 7, "august": 8,
    "september": 9, "october": 10, "november": 11, "december": 12,
    "jan": 1, "feb": 2, "mar": 3, "apr": 4,
    "jun": 6, "jul": 7, "aug": 8, "sep": 9,
    "oct": 10, "nov": 11, "dec": 12,
}


def _safe_str(v) -> str:
    if v is None:
        return ""
    return str(v).strip()


def _open_workbook(data: bytes, ext: str):
    """Open workbook with appropriate library. Returns (wb, lib) tuple."""
    if ext == ".xlsx":
        import openpyxl
        return openpyxl.load_workbook(io.BytesIO(data), read_only=True, data_only=True), "openpyxl"
    else:
        import xlrd
        return xlrd.open_workbook(file_contents=data), "xlrd"


def _sheet_names(wb, lib: str) -> list[str]:
    if lib == "openpyxl":
        return list(wb.sheetnames)
    else:
        return wb.sheet_names()


def _read_sheet_rows(wb, lib: str, sheet_name: str) -> list[dict]:
    """Read a sheet into list of {col_index: value} dicts."""
    rows = []
    try:
        if lib == "openpyxl":
            ws = wb[sheet_name]
            for row in ws.iter_rows(values_only=True):
                row_dict = {i: row[i] for i in range(len(row))}
                rows.append(row_dict)
        else:  # xlrd
            ws = wb.sheet_by_name(sheet_name)
            for r in range(ws.nrows):
                row_dict = {c: ws.cell_value(r, c) for c in range(ws.ncols)}
                rows.append(row_dict)
    except Exception as e:
        log.debug(f"  Error reading sheet {sheet_name!r}: {e}")
    return rows


def _detect_ext(data: bytes) -> str:
    """Auto-detect whether bytes are xlsx (PK/ZIP) or xls (OLE2/D0CF) or unknown."""
    if data[:4] == b'PK\x03\x04':
        return ".xlsx"
    if data[:8] == b'\xd0\xcf\x11\xe0\xa1\xb1\x1a\xe1':
        return ".xls"
    return ".xls"  # default guess


def _parse_ppfas_file(data: bytes, ext: str, year: int, month: int) -> list[dict]:
    """
    Parse a PPFAS monthly portfolio XLS/XLSX file.
    Auto-detects format from magic bytes regardless of the URL extension.
    Returns list of holding dicts.
    """
    records: list[dict] = []

    # Auto-detect actual format from file signature (overrides URL extension)
    actual_ext = _detect_ext(data)
    if actual_ext != ext:
        log.debug(f"  Format auto-corrected: URL={ext}, actual={actual_ext}")
        ext = actual_ext

    try:
        wb, lib = _open_workbook(data, ext)
    except Exception as e:
        log.error(f"  Cannot open workbook: {e}")
        return records

    sheet_names = _sheet_names(wb, lib)
    log.info(f"  Sheets: {sheet_names}")

    # Build fund_code → scheme_name from Index sheet
    fund_map: dict[str, str] = {}
    if "Index" in sheet_names:
        rows = _read_sheet_rows(wb, lib, "Index")
        for row in rows[1:]:  # skip header
            code = _safe_str(row.get(1))
            name = _safe_str(row.get(2))
            if code and name:
                fund_map[code] = name
    log.info(f"  Fund map: {fund_map}")

    # Process each fund sheet
    for sheet in sheet_names:
        if sheet == "Index":
            continue

        rows = _read_sheet_rows(wb, lib, sheet)
        if not rows:
            continue

        # Extract fund name from row 0 col 1
        fund_name_raw = _safe_str(rows[0].get(1, "")) if rows else ""

        # Extract date from row 2 col 1
        date_raw = _safe_str(rows[2].get(1, "")) if len(rows) > 2 else ""
        m_date = DATE_RE.search(date_raw)
        if m_date:
            file_month = MONTH_MAP.get(m_date.group(1).lower(), 0)
            file_year = int(m_date.group(2))
            if file_month != month or file_year != year:
                log.warning(f"  {sheet}: date mismatch: expected {year}-{month:02d}, "
                             f"got {file_year}-{file_month:02d}. Using URL date.")
        # Use the URL-derived year/month (more reliable)

        # Determine scheme name
        scheme_name = fund_map.get(sheet, fund_name_raw or sheet)

        # Find header row (col 2 starts with "ISIN")
        header_row_idx = 3  # Default
        for i, row in enumerate(rows):
            cell2 = _safe_str(row.get(2, "")).lower()
            if "isin" in cell2:
                header_row_idx = i
                break

        # Parse holdings rows (after header)
        for row in rows[header_row_idx + 1:]:
            isin = _safe_str(row.get(2, ""))
            if not ISIN_RE.match(isin):
                continue

            stock_name = _safe_str(row.get(1, ""))
            pct_raw = row.get(6)
            if pct_raw is None or pct_raw == "":
                continue
            try:
                pct_decimal = float(pct_raw)
            except (ValueError, TypeError):
                continue

            # PPFAS stores as decimal (0.0811 = 8.11%)
            # Sanity check: reject if > 1.05 (would be >105%, impossible)
            if not (0 < pct_decimal <= 1.05):
                continue

            pct_val = round(pct_decimal * 100, 4)

            records.append({
                "amc": AMC_LABEL,
                "scheme_name": scheme_name,
                "isin": isin,
                "stock_name": stock_name,
                "pct_nav": pct_val,
                "year": year,
                "month": month,
            })

    log.info(f"  Parsed {len(records)} holdings from {len(sheet_names)-1} fund sheets")
    return records


# ─────────────────────────────────────────────────────────────────────────────
# Scheme code attachment
# ─────────────────────────────────────────────────────────────────────────────

# Hardcoded PPFAS scheme codes from AMFI (Direct-Growth plans only — used for scoring)
PPFAS_SCHEME_CODES = {
    "Parag Parikh Flexi Cap Fund":            122639,  # Direct-Growth
    "Parag Parikh Liquid Fund":                143269,  # Direct-Growth
    "Parag Parikh ELSS Tax Saver Fund":        147481,  # Direct-Growth
    "Parag Parikh Conservative Hybrid Fund":   148958,  # Direct-Growth
    # New funds (codes from mfapi)
    "Parag Parikh Arbitrage Fund":             0,       # TBD
    "Parag Parikh Dynamic Asset Allocation Fund": 0,    # TBD
}

# Full scheme list from AMFI (all plans) for comprehensive matching
PPFAS_ALL_SCHEMES = [
    (153965, "Parag Parikh Flexi Cap Fund - Regular Plan - IDCW"),
    (153964, "Parag Parikh Flexi Cap Fund - Direct Plan - IDCW"),
    (122640, "Parag Parikh Flexi Cap Fund - Regular Plan - Growth"),
    (122639, "Parag Parikh Flexi Cap Fund - Direct Plan - Growth"),
    (143261, "Parag Parikh Liquid Fund - Regular Plan - Monthly IDCW"),
    (143262, "Parag Parikh Liquid Fund - Direct Plan - Monthly IDCW"),
    (143264, "Parag Parikh Liquid Fund - Regular Plan - Daily Reinvestment of IDCW"),
    (143263, "Parag Parikh Liquid Fund - Direct Plan - Daily Reinvestment of IDCW"),
    (143266, "Parag Parikh Liquid Fund - Regular Plan - Weekly Reinvestment of IDCW"),
    (143265, "Parag Parikh Liquid Fund - Direct Plan - Weekly Reinvestment of IDCW"),
    (143260, "Parag Parikh Liquid Fund - Regular Plan - Growth"),
    (143269, "Parag Parikh Liquid Fund - Direct Plan - Growth"),
    (147481, "Parag Parikh ELSS Tax Saver Fund - Direct Growth"),
    (147482, "Parag Parikh ELSS Tax Saver Fund - Regular Growth"),
    (148958, "Parag Parikh Conservative Hybrid Fund - Direct Plan - Growth"),
]


def _load_scheme_master(mf_data_dir: Path) -> pd.DataFrame:
    """Load AMFI scheme master for PPFAS, falling back to hardcoded list."""
    # Try local scheme_list (has codes we've already ingested)
    for fname in ["scheme_master.parquet", "scheme_list.parquet"]:
        try:
            sm = pd.read_parquet(mf_data_dir / fname)
            ppfas = sm[sm.apply(
                lambda r: r.astype(str).str.contains("Parag|PPFAS", case=False).any(), axis=1
            )].copy()
            if not ppfas.empty:
                log.info(f"  Loaded {len(ppfas)} PPFAS schemes from {fname}")
                # Standardize column names
                if "amc_name" not in ppfas.columns and "amc" in ppfas.columns:
                    ppfas = ppfas.rename(columns={"amc": "amc_name"})
                return ppfas
        except Exception:
            pass

    # Fallback: use hardcoded PPFAS scheme list
    log.info("  Using hardcoded PPFAS scheme list (36 schemes)")
    rows = [{"scheme_code": code, "scheme_name": name, "amc_name": "PPFAS Mutual Fund"}
            for code, name in PPFAS_ALL_SCHEMES]
    return pd.DataFrame(rows)


def _clean_name(name: str) -> str:
    """Strip plan/option suffixes for fuzzy matching."""
    name = re.sub(r'\(.*?\)', '', name)
    name = re.sub(
        r'\b(direct|regular|growth|idcw|dividend|bonus|reinvest|plan|option|fund|flexi)\b',
        '', name, flags=re.IGNORECASE
    )
    return re.sub(r'\s+', ' ', name).strip().lower()


def _attach_scheme_codes(df: pd.DataFrame, sm: pd.DataFrame) -> pd.DataFrame:
    """Add scheme_code column by fuzzy matching scheme_name to AMFI master."""
    if sm.empty or df.empty:
        df["scheme_code"] = None
        return df

    import difflib
    clean_sm = sm.copy()
    clean_sm["_clean"] = clean_sm["scheme_name"].apply(_clean_name)
    name_to_code = dict(zip(clean_sm["_clean"], clean_sm["scheme_code"]))
    clean_choices = list(name_to_code.keys())

    def match_name(raw: str) -> Optional[int]:
        c = _clean_name(raw)
        # Exact
        if c in name_to_code:
            return name_to_code[c]
        # Fuzzy
        matches = difflib.get_close_matches(c, clean_choices, n=1, cutoff=0.6)
        if matches:
            return name_to_code[matches[0]]
        return None

    unique_names = df["scheme_name"].unique()
    match_map = {n: match_name(n) for n in unique_names}
    matched = sum(1 for v in match_map.values() if v is not None)
    log.info(f"  Scheme matching: {matched}/{len(unique_names)} matched")
    for name, code in match_map.items():
        if code is None:
            log.debug(f"    No match: {name!r}")

    df = df.copy()
    df["scheme_code"] = df["scheme_name"].map(match_map)
    return df


# ─────────────────────────────────────────────────────────────────────────────
# Cache merge
# ─────────────────────────────────────────────────────────────────────────────

def _merge_into_cache(df: pd.DataFrame, holdings_dir: Path, force: bool) -> bool:
    """Merge new records into the monthly parquet cache."""
    if df.empty:
        return False
    months = df[["year", "month"]].drop_duplicates()
    changed = False
    for _, row in months.iterrows():
        y, m = int(row["year"]), int(row["month"])
        month_str = f"{y}-{m:02d}"
        pq = holdings_dir / f"{month_str}.parquet"

        new_rows = df[(df["year"] == y) & (df["month"] == m)].copy()

        if pq.exists():
            existing = pd.read_parquet(pq)
            already = (existing["amc"] == AMC_LABEL).any()
            if already and not force:
                log.info(f"  {month_str}: {AMC_LABEL} already in cache, skipping (use --force to overwrite)")
                continue
            # Drop old AMC rows, append new
            existing = existing[existing["amc"] != AMC_LABEL]
            merged = pd.concat([existing, new_rows], ignore_index=True)
        else:
            merged = new_rows

        merged.to_parquet(pq, index=False)
        log.info(f"  {month_str}: saved {len(new_rows)} rows → {pq.name}")
        changed = True
    return changed


# ─────────────────────────────────────────────────────────────────────────────
# Main
# ─────────────────────────────────────────────────────────────────────────────

def main():
    ap = argparse.ArgumentParser(description="Ingest PPFAS monthly portfolio disclosures")
    ap.add_argument("--mf-data", default="./mf_data", help="Path to mf_data directory")
    ap.add_argument("--force", action="store_true", help="Overwrite existing months")
    ap.add_argument("--month", default=None, help="Single month to process, e.g. 2024-03")
    ap.add_argument("--start", default="2023-05", help="Start month (default: 2023-05)")
    ap.add_argument("--end", default=None, help="End month (default: today's month)")
    args = ap.parse_args()

    mf_data_dir = Path(args.mf_data)
    holdings_dir = mf_data_dir / "holdings"
    holdings_dir.mkdir(parents=True, exist_ok=True)

    # Determine month range
    if args.month:
        y, m = map(int, args.month.split("-"))
        months_to_process = [(y, m)]
    else:
        start_y, start_m = map(int, args.start.split("-"))
        if args.end:
            end_y, end_m = map(int, args.end.split("-"))
        else:
            today = date.today()
            # Use previous month (current month may not be published yet)
            prev = today.replace(day=1) - timedelta(days=1)
            end_y, end_m = prev.year, prev.month
        months_to_process = _all_ppfas_months(
            date(start_y, start_m, 1), date(end_y, end_m, 1)
        )

    log.info(f"PPFAS ingest: {len(months_to_process)} months to process")
    sm = _load_scheme_master(mf_data_dir)

    all_records: list[dict] = []
    errors: list[str] = []

    for year, month in months_to_process:
        month_str = f"{year}-{month:02d}"

        # Check if already cached (unless force)
        pq = holdings_dir / f"{month_str}.parquet"
        if not args.force and pq.exists():
            try:
                existing = pd.read_parquet(pq)
                if (existing["amc"] == AMC_LABEL).any():
                    log.info(f"{month_str}: already cached, skipping")
                    continue
            except Exception:
                pass

        # Try both extensions — some months have mismatched extensions on server
        primary_ext = _ppfas_ext(year, month)
        fallback_ext = ".xlsx" if primary_ext == ".xls" else ".xls"

        url = _ppfas_url(year, month)
        log.info(f"{month_str}: downloading {url}")
        data = _download(url)
        used_ext = primary_ext

        records = []
        if data is not None:
            records = _parse_ppfas_file(data, primary_ext, year, month)

        # If parse failed (wrong extension) or download failed, try fallback
        if not records:
            fallback_url = url.replace(primary_ext, fallback_ext)
            log.info(f"{month_str}: trying fallback extension {fallback_ext}")
            data2 = _download(fallback_url)
            if data2 is not None:
                records2 = _parse_ppfas_file(data2, fallback_ext, year, month)
                if records2:
                    records = records2
                    used_ext = fallback_ext
                    log.info(f"{month_str}: fallback {fallback_ext} succeeded")

        if data is None and not records:
            log.error(f"{month_str}: download failed — skipping")
            errors.append(month_str)
            continue
        if not records:
            log.warning(f"{month_str}: no records parsed from either extension")
            errors.append(month_str)
            continue

        df = pd.DataFrame(records)

        # Attach scheme codes
        if not sm.empty:
            df = _attach_scheme_codes(df, sm)

        # Merge into cache
        _merge_into_cache(df, holdings_dir, args.force)
        all_records.extend(records)

    # Summary
    log.info("=" * 60)
    log.info(f"PPFAS ingest complete:")
    log.info(f"  Months processed : {len(months_to_process)}")
    log.info(f"  Errors/skipped   : {len(errors)} {errors if errors else ''}")
    log.info(f"  Total records    : {len(all_records)}")
    if all_records:
        df_all = pd.DataFrame(all_records)
        funds = df_all["scheme_name"].unique()
        log.info(f"  Unique funds     : {len(funds)}: {list(funds)}")


if __name__ == "__main__":
    main()
