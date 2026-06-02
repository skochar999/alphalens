#!/usr/bin/env python3
"""
amc_holdings_scraper.py
=======================
Downloads SEBI-mandated monthly portfolio disclosures for all 10 AMCs
in the IEC-1 FundLens dataset and returns a unified DataFrame:

    scheme_code | isin | stock_name | pct_nav | as_of_date | amc

Usage
-----
    python3 amc_holdings_scraper.py              # downloads April 2026 (latest)
    python3 amc_holdings_scraper.py --month 2026-03   # specific month
    python3 amc_holdings_scraper.py --amc SBI         # single AMC only

The output is written to:
    amc_portfolio_<YYYY-MM>.parquet   (used by fetch_holdings_and_score.py)
    amc_portfolio_<YYYY-MM>.csv       (human-readable)

AMC coverage (155/172 funds = 90%):
  ✅ SBI         30 funds  – consolidated xlsx (Sitefinity CMS)
  ✅ ICICI_Pru   25 funds  – zip → per-scheme xlsx (blob CDN)
  ✅ HDFC        19 funds  – per-scheme xlsx (S3 CDN, URLs scraped from page)
  ✅ Aditya_Birla 16 funds – zip → xlsx (Sitecore media)
  ✅ Nippon      16 funds  – consolidated xlsx (SharePoint)
  ✅ Axis        14 funds  – consolidated xlsx (Drupal CDN)
  ✅ DSP         12 funds  – zip → xlsx (Statamic, URL scraped from page)
  ✅ Franklin    12 funds  – consolidated xlsx (UUID URL scraped from AdvisorKhoj)
  ⚠️  Kotak      17 funds  – blocked by Radware; add URL manually
  ⚠️  Mirae      11 funds  – Sitefinity wildcard; add URL manually
"""

from __future__ import annotations

import argparse
import calendar
import datetime
import io
import logging
import os
import re
import subprocess
import sys
import urllib.parse
import zipfile
from pathlib import Path
from typing import Optional

import openpyxl
import pandas as pd
from difflib import SequenceMatcher

logging.basicConfig(level=logging.INFO, format="%(levelname)s  %(message)s")
log = logging.getLogger(__name__)

UA = (
    "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) "
    "AppleWebKit/537.36 (KHTML, like Gecko) Chrome/124.0.0.0 Safari/537.36"
)

# ─────────────────────────────────────────────────────────────────────────────
# Helpers
# ─────────────────────────────────────────────────────────────────────────────

def _curl(url: str, timeout: int = 45, binary: bool = False) -> bytes | str | None:
    """Fetch URL via system curl.  Returns bytes if binary else str, or None."""
    cmd = [
        "curl", "-s", "-L", "--max-time", str(timeout),
        "-A", UA,
        "-H", "Accept-Language: en-IN,en;q=0.9",
        "--compressed",
        url,
    ]
    try:
        r = subprocess.run(cmd, capture_output=True, timeout=timeout + 5)
        if r.returncode != 0 or len(r.stdout) < 50:
            return None
        return r.stdout if binary else r.stdout.decode("utf-8", errors="replace")
    except Exception as e:
        log.debug(f"curl failed for {url}: {e}")
        return None


def _month_parts(month_str: str):
    """'2026-04'  →  (2026, 4, '30', 'April', 'Apr', 'april', '26')"""
    dt = datetime.datetime.strptime(month_str, "%Y-%m")
    y, m = dt.year, dt.month
    last_day = calendar.monthrange(y, m)[1]
    month_name   = dt.strftime("%B")         # April
    month_abbr   = dt.strftime("%b")         # Apr
    year2        = str(y)[-2:]               # 26
    return y, m, str(last_day), month_name, month_abbr, month_name.lower(), year2


def _ordinal(n: int) -> str:
    """1 → '1st', 28 → '28th', 30 → '30th', 31 → '31st'"""
    suffix = {1: "st", 2: "nd", 3: "rd"}.get(n if n < 20 else n % 10, "th")
    return f"{n}{suffix}"


def _month_num_str(m: int) -> str:
    return f"{m:02d}"

# ─────────────────────────────────────────────────────────────────────────────
# Generic SEBI-format Excel parser
# ─────────────────────────────────────────────────────────────────────────────

ISIN_RE = re.compile(r'^[A-Z]{2}[A-Z0-9]{10}$')


def _is_isin(val) -> bool:
    if not isinstance(val, str):
        return False
    return bool(ISIN_RE.match(val.strip()))


def _parse_sheet(ws) -> pd.DataFrame:
    """
    Parse one worksheet from a SEBI-format portfolio Excel.
    Auto-detects header row, ISIN column, name column, % NAV column.
    Returns DataFrame with columns: isin, stock_name, pct_nav, scheme_name.
    """
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return pd.DataFrame()

    # ── 1. Detect scheme name (usually rows 1–3, the cell with longest text) ──
    scheme_name = ""
    for i, row in enumerate(rows[:5]):
        for v in row:
            if isinstance(v, str) and 20 < len(v) < 120:
                if any(kw in v.lower() for kw in ("fund", "scheme", "direct", "growth")):
                    scheme_name = v.strip()
                    break
        if scheme_name:
            break

    # ── 2. Find header row (contains 'isin' and a weight column) ─────────────
    NAV_KEYWORDS = ("% to nav", "% to aum", "% to net assets", "% net assets",
                    "%tonav", "%toaum", "% of nav", "weightage")
    header_row_idx = None
    for i, row in enumerate(rows):
        cells = [str(v).lower().strip() if v else "" for v in row]
        has_isin   = any("isin" in c for c in cells)
        has_weight = any(any(kw in c for kw in NAV_KEYWORDS) for c in cells)
        if has_isin and has_weight:
            header_row_idx = i
            break
    # Fallback: row with just 'isin' (weight col detected from data later)
    if header_row_idx is None:
        for i, row in enumerate(rows):
            cells = [str(v).lower().strip() if v else "" for v in row]
            if any(c == "isin" for c in cells):
                header_row_idx = i
                break

    if header_row_idx is None:
        return pd.DataFrame()

    # Normalise headers: collapse whitespace/newlines, strip
    def _norm_header(v) -> str:
        if v is None:
            return ""
        return re.sub(r"\s+", " ", str(v).lower()).strip()

    headers = [_norm_header(v) for v in rows[header_row_idx]]

    # ── 3. Identify columns ───────────────────────────────────────────────────
    isin_col = next((i for i, h in enumerate(headers) if h == "isin"), None)
    if isin_col is None:
        isin_col = next((i for i, h in enumerate(headers) if "isin" in h), None)
    nav_col = next(
        (i for i, h in enumerate(headers)
         if any(kw in h for kw in NAV_KEYWORDS)),
        None
    )
    name_col = next((i for i, h in enumerate(headers)
                     if "name" in h and "instrument" in h), None)
    if name_col is None:
        name_col = next((i for i, h in enumerate(headers)
                         if "name" in h and i != isin_col), None)

    if isin_col is None or nav_col is None:
        # Last resort: auto-detect ISIN column from first data row
        for row in rows[header_row_idx+1:header_row_idx+6]:
            for ci, v in enumerate(row):
                if _is_isin(v):
                    isin_col = ci
                    break
            if isin_col is not None:
                break
        if isin_col is None or nav_col is None:
            return pd.DataFrame()

    # ── 4. Extract rows ───────────────────────────────────────────────────────
    records = []
    for row in rows[header_row_idx + 1:]:
        if len(row) <= max(isin_col, nav_col):
            continue
        isin_val = row[isin_col]
        if not _is_isin(isin_val):
            continue
        nav_val  = row[nav_col]
        if nav_val is None:
            continue
        try:
            pct = float(nav_val)
        except (ValueError, TypeError):
            continue
        if pct <= 0 or pct > 100:
            continue
        name_val = ""
        if name_col is not None and name_col < len(row):
            name_val = str(row[name_col] or "").strip()
        records.append({
            "isin":        isin_val.strip(),
            "stock_name":  name_val,
            "pct_nav":     pct,
            "scheme_name": scheme_name,
        })
    df = pd.DataFrame(records)
    if df.empty:
        return df
    # Auto-scale: if all pct_nav ≤ 1.5, values are decimal fractions (e.g. 0.0421 = 4.21%)
    if df["pct_nav"].max() <= 1.5:
        df["pct_nav"] = (df["pct_nav"] * 100).round(4)
    return df


def _parse_workbook(wb: openpyxl.Workbook, skip_sheets=("index",)) -> pd.DataFrame:
    """Parse all relevant sheets in a workbook, concatenate results."""
    dfs = []
    for sn in wb.sheetnames:
        if sn.lower() in skip_sheets:
            continue
        ws = wb[sn]
        df = _parse_sheet(ws)
        if not df.empty:
            df["_sheet"] = sn
            dfs.append(df)
    return pd.concat(dfs, ignore_index=True) if dfs else pd.DataFrame()


def _load_workbook_bytes(data: bytes) -> Optional[openpyxl.Workbook]:
    """Load from raw bytes (handles .xls extension that is actually xlsx)."""
    try:
        return openpyxl.load_workbook(io.BytesIO(data))
    except Exception as e:
        log.debug(f"openpyxl failed: {e}")
        return None


def _parse_zip(data: bytes, month_str: str,
               scheme_name_from_file: bool = False) -> pd.DataFrame:
    """Extract xlsx files from a zip and parse each.

    If scheme_name_from_file=True the filename stem (e.g. 'ICICI Prudential
    Bluechip Fund') is used as scheme_name instead of what _parse_sheet detects
    from the worksheet content.  Use this for ICICI's per-scheme zip layout.
    """
    dfs = []
    try:
        with zipfile.ZipFile(io.BytesIO(data)) as zf:
            xlsx_names = [n for n in zf.namelist()
                          if n.lower().endswith((".xlsx", ".xls"))
                          and not n.startswith("__MACOSX")]
            log.info(f"  ZIP contains {len(xlsx_names)} Excel files")
            for name in xlsx_names:
                raw = zf.read(name)
                wb  = _load_workbook_bytes(raw)
                if wb is None:
                    continue
                df = _parse_workbook(wb)
                if not df.empty:
                    if scheme_name_from_file:
                        # Filename IS the scheme name for ICICI per-scheme zips
                        stem = Path(name).stem  # e.g. "ICICI Prudential Bluechip Fund"
                        df["scheme_name"] = stem
                    dfs.append(df)
    except Exception as e:
        log.warning(f"  ZIP parse error: {e}")
    return pd.concat(dfs, ignore_index=True) if dfs else pd.DataFrame()

# ─────────────────────────────────────────────────────────────────────────────
# Per-AMC downloaders
# ─────────────────────────────────────────────────────────────────────────────

def _download_sbi(month_str: str) -> pd.DataFrame:
    """
    SBI MF — consolidated xlsx (Sitefinity CMS)
    Pattern: all-schemes-monthly-portfolio---as-on-{ordinal}-{month}-{year}.xlsx
    """
    y, m, last, mname, mabbr, mlower, y2 = _month_parts(month_str)
    ord_day = _ordinal(int(last))
    fname = f"all-schemes-monthly-portfolio---as-on-{ord_day}-{mlower}-{y}.xlsx"
    url = f"https://www.sbimf.com/docs/default-source/scheme-portfolios/{fname}"
    log.info(f"[SBI] {url}")
    data = _curl(url, binary=True)
    if not data:
        log.warning("[SBI] Download failed")
        return pd.DataFrame()
    wb = _load_workbook_bytes(data)
    if wb is None:
        log.warning("[SBI] Not a valid Excel file")
        return pd.DataFrame()
    df = _parse_workbook(wb)
    n_schemes = df["scheme_name"].nunique() if not df.empty else 0
    log.info(f"[SBI] {len(df)} holding rows across {n_schemes} schemes")
    return df


def _download_icici(month_str: str) -> pd.DataFrame:
    """
    ICICI Prudential — zip of per-scheme xlsx (blob CDN)
    Pattern: /blob/downloads/Files/Monthly Portfolio Disclosures/{YYYY}/{Mon}/Monthly-Portfolio-Disclosure-{Month}-{YYYY}.zip
    """
    y, m, last, mname, mabbr, mlower, y2 = _month_parts(month_str)
    fname = f"Monthly-Portfolio-Disclosure-{mname}-{y}.zip"
    url   = (
        f"https://www.icicipruamc.com/blob/downloads/Files/"
        f"Monthly%20Portfolio%20Disclosures/{y}/{mabbr}/{fname}"
    )
    log.info(f"[ICICI_Pru] {url}")
    data = _curl(url, binary=True, timeout=90)
    if not data or data[:4] != b'PK\x03\x04':
        log.warning("[ICICI_Pru] Download failed or not a ZIP")
        return pd.DataFrame()
    # scheme_name_from_file=True: each file is named after its fund
    # (e.g. "ICICI Prudential Bluechip Fund.xlsx"), which is more reliable
    # than the in-sheet row 0 which says "ICICI Prudential Mutual Fund" for all.
    df = _parse_zip(data, month_str, scheme_name_from_file=True)
    log.info(f"[ICICI_Pru] {len(df)} holding rows across {df['scheme_name'].nunique() if not df.empty else 0} schemes")
    return df


def _download_hdfc(month_str: str) -> pd.DataFrame:
    """
    HDFC — per-scheme xlsx files from S3; URLs scraped from statutory disclosure page.
    Page: https://www.hdfcfund.com/statutory-disclosure/portfolio/monthly-portfolio
    Files: https://files.hdfcfund.com/s3fs-public/{YYYY-MM}/Monthly%20{scheme}%20-%20{DD}%20{Month}%20{YYYY}.xlsx
    """
    y, m, last, mname, mabbr, mlower, y2 = _month_parts(month_str)
    page_url = "https://www.hdfcfund.com/statutory-disclosure/portfolio/monthly-portfolio"
    log.info(f"[HDFC] Fetching page: {page_url}")
    html = _curl(page_url)
    if not html:
        log.warning("[HDFC] Page fetch failed")
        return pd.DataFrame()

    # HDFC publishes month-end data in the NEXT calendar month's S3 folder
    # e.g. "30 April 2026" files live in s3fs-public/2026-05/
    # → match on month name in filename, not folder path
    all_hdfc_urls = list(dict.fromkeys(re.findall(
        r'https://files\.hdfcfund\.com/s3fs-public/\d{4}-\d{2}/[^\s"\'<>]+\.xlsx',
        html, re.I
    )))
    # Filter to files containing target month+year in the URL
    file_urls = [u for u in all_hdfc_urls
                 if mname.lower() in urllib.parse.unquote(u).lower()
                 and str(y) in u]
    # Fallback: if no month match, take all from the page (assume it shows current month)
    if not file_urls:
        file_urls = all_hdfc_urls
    log.info(f"[HDFC] Found {len(file_urls)} scheme files for {mname} {y}")

    dfs = []
    for url in file_urls:
        data = _curl(url, binary=True, timeout=30)
        if not data:
            continue
        wb = _load_workbook_bytes(data)
        if wb is None:
            continue
        df = _parse_workbook(wb)
        if not df.empty:
            dfs.append(df)

    result = pd.concat(dfs, ignore_index=True) if dfs else pd.DataFrame()
    log.info(f"[HDFC] {len(result)} holding rows across {len(dfs)} schemes")
    return result


def _download_kotak(month_str: str) -> pd.DataFrame:
    """
    Kotak MF — blocked by Radware bot protection.

    TO FIX: From your browser, navigate to:
      https://www.kotakmf.com/Information/statutory-disclosure
    Find the "Monthly Portfolio Disclosure" link for the relevant month,
    copy the direct URL and paste it below as `url`.

    Expected CDN pattern (approximate):
      https://www.kotakmf.com/assets/portfolio/monthly-portfolio-{month}-{year}.xlsx
    """
    y, m, last, mname, mabbr, mlower, y2 = _month_parts(month_str)

    # ── Attempt: try the Kotak statutory disclosure page ─────────────────────
    page_url = "https://www.kotakmf.com/Information/statutory-disclosure"
    html = _curl(page_url)
    if html:
        file_urls = re.findall(
            r'https?://[^\s"\'<>\\]+\.(?:xlsx|xls|zip)',
            html, re.I
        )
        if file_urls:
            log.info(f"[Kotak] Found {len(file_urls)} file URLs on page")
            dfs = []
            for url in file_urls[:5]:
                if mlower in url.lower() or str(y) in url:
                    data = _curl(url, binary=True)
                    if data:
                        if data[:4] == b'PK\x03\x04':
                            df = _parse_zip(data, month_str) if data[:2] == b'PK' else pd.DataFrame()
                        else:
                            wb = _load_workbook_bytes(data)
                            df = _parse_workbook(wb) if wb else pd.DataFrame()
                        if not df.empty:
                            dfs.append(df)
            if dfs:
                return pd.concat(dfs, ignore_index=True)

    log.warning(
        "[Kotak] Auto-download failed (Radware bot protection). "
        "Please download the monthly portfolio Excel manually from "
        "https://www.kotakmf.com/Information/statutory-disclosure "
        "and place it as: kotak_portfolio_manual.xlsx"
    )
    # Fallback: check for manually placed file
    manual = Path("kotak_portfolio_manual.xlsx")
    if manual.exists():
        log.info("[Kotak] Using manual file")
        wb = openpyxl.load_workbook(manual)
        return _parse_workbook(wb)
    return pd.DataFrame()


def _parse_absl_xls(raw: bytes) -> pd.DataFrame:
    """
    Parse an ABSL consolidated portfolio .xls (OLE2/BIFF format — requires xlrd).

    Structure per sheet:
      row 0:  [fund_code, full_fund_name]
      row 1:  description
      row 2:  "Portfolio Statement as on {date}"
      row 3:  headers — Name, ISIN, Industry, Quantity, Market Value, % to Net Assets, …
      row 4+: category rows (no ISIN) then holding rows (valid ISIN)

    pct_nav values are decimal (0.0949 = 9.49%), auto-scaled below.
    Index sheet maps short codes to full names (used as fallback).
    """
    try:
        import xlrd
    except ImportError:
        log.error("[ABSL] xlrd not installed — run: pip install xlrd")
        return pd.DataFrame()
    try:
        wb = xlrd.open_workbook(file_contents=raw)
    except Exception as e:
        log.warning(f"[ABSL] xlrd open failed: {e}")
        return pd.DataFrame()

    NAV_KEYWORDS = ("% to nav", "% to aum", "% to net assets", "% net assets",
                    "%tonav", "%toaum", "% of nav", "weightage")

    # Build code→name from Index sheet
    code_to_name: dict[str, str] = {}
    idx = next((s for s in wb.sheet_names() if s.lower() == "index"), None)
    if idx:
        ws_i = wb.sheet_by_name(idx)
        for ri in range(ws_i.nrows):
            row = [ws_i.cell_value(ri, j) for j in range(ws_i.ncols)]
            if len(row) >= 2 and row[0] and isinstance(row[0], str) and 1 <= len(row[0].strip()) <= 12:
                code_to_name[row[0].strip()] = str(row[1]).strip()

    dfs = []
    for sn in wb.sheet_names():
        if sn.lower() == "index":
            continue
        ws = wb.sheet_by_name(sn)
        if ws.nrows < 5:
            continue

        # Scheme name: prefer Index lookup, fall back to row-0 col-1
        scheme_name = code_to_name.get(sn, "")
        if not scheme_name and ws.nrows > 0:
            scheme_name = str(ws.cell_value(0, 1)).strip()

        # Find header row (contains 'isin' + a weight keyword)
        header_row_idx = None
        for ri in range(min(10, ws.nrows)):
            row_vals = [re.sub(r"\s+", " ", str(ws.cell_value(ri, j)).lower()).strip()
                        for j in range(ws.ncols)]
            has_isin   = any("isin" in c for c in row_vals)
            has_weight = any(any(kw in c for kw in NAV_KEYWORDS) for c in row_vals)
            if has_isin and has_weight:
                header_row_idx = ri
                break
        if header_row_idx is None:
            continue

        headers = [re.sub(r"\s+", " ", str(ws.cell_value(header_row_idx, j)).lower()).strip()
                   for j in range(ws.ncols)]

        isin_col = next((i for i, h in enumerate(headers) if h == "isin"), None)
        if isin_col is None:
            isin_col = next((i for i, h in enumerate(headers) if "isin" in h), None)
        nav_col  = next((i for i, h in enumerate(headers)
                         if any(kw in h for kw in NAV_KEYWORDS)), None)
        name_col = next((i for i, h in enumerate(headers)
                         if "name" in h and i != isin_col), None)

        if isin_col is None or nav_col is None:
            continue

        records = []
        for ri in range(header_row_idx + 1, ws.nrows):
            isin_val = str(ws.cell_value(ri, isin_col)).strip()
            if not ISIN_RE.match(isin_val):
                continue
            nav_raw = ws.cell_value(ri, nav_col)
            try:
                pct = float(nav_raw)
            except (TypeError, ValueError):
                continue
            if pct <= 0 or pct > 100:
                continue
            name_val = ""
            if name_col is not None and name_col < ws.ncols:
                name_val = str(ws.cell_value(ri, name_col)).strip()
            records.append({
                "isin":        isin_val,
                "stock_name":  name_val,
                "pct_nav":     pct,
                "scheme_name": scheme_name,
            })

        if records:
            df = pd.DataFrame(records)
            if df["pct_nav"].max() <= 1.5:
                df["pct_nav"] = (df["pct_nav"] * 100).round(4)
            dfs.append(df)

    result = pd.concat(dfs, ignore_index=True) if dfs else pd.DataFrame()
    return result


def _download_absl(month_str: str) -> pd.DataFrame:
    """
    Aditya Birla Sun Life — zip → .xls (OLE2 BIFF, Sitecore media CDN).

    ABSL changes its filename pattern every month (e.g.
      Apr-2026: monthly-disclosure-april-30-2026.zip
      Mar-2026: monthly-portfolio-mar-2026.zip
      Feb-2026: sebi_monthly_portfolio-28-feb-2026.zip).
    Strategy: scrape AdvisorKhoj (like Franklin) to get the exact current URL,
    then fall back to pattern-guessing.  The zip contains one consolidated
    multi-sheet .xls workbook in old OLE2/BIFF format (parsed with xlrd).
    """
    y, m, last, mname, mabbr, mlower, y2 = _month_parts(month_str)
    mabbr_lower = mabbr.lower()

    # ── Step 1: scrape AdvisorKhoj for the exact URL ─────────────────────────
    ak_url = (
        "https://www.advisorkhoj.com/form-download-centre/Mutual/"
        "Aditya-Birla-Sun-Life-Mutual-Fund/Monthly-Portfolio-Disclosures"
    )
    log.info(f"[ABSL] Scraping AdvisorKhoj for URL …")
    html = _curl(ak_url, timeout=30)
    target = None
    if html:
        # All ABSL CDN links look like: mutualfund.adityabirlacapital.com/...zip or .xls
        all_urls = list(dict.fromkeys(re.findall(
            r'https://mutualfund\.adityabirlacapital\.com'
            r'/-/media/bsl/files/resources/monthly-portfolio/[^\s"\'<>\\]+\.(?:zip|xls|xlsx)',
            html, re.I
        )))
        # Match our target month
        for url in all_urls:
            fname = url.split("/")[-1].lower()
            if (mlower in fname or mabbr_lower in fname) and str(y) in fname:
                target = url
                break
        if target is None and all_urls:
            target = all_urls[0]
            log.warning(f"[ABSL] Month match not found — using most recent: {target.split('/')[-1]}")

    # ── Step 2: pattern fallback if AdvisorKhoj scrape failed ────────────────
    if target is None:
        log.warning("[ABSL] AdvisorKhoj scrape failed; trying known URL patterns …")
        base = (f"https://mutualfund.adityabirlacapital.com/-/media/bsl/files/"
                f"resources/monthly-portfolio/{y}/")
        for fname in [
            f"monthly-disclosure-{mlower}-{last}-{y}.zip",
            f"monthly-disclosure-{mabbr_lower}-{last}-{y}.zip",
            f"sebi_monthly_portfolio-{last}-{mabbr_lower}-{y}.zip",
            f"monthly-portfolio-{mabbr_lower}-{y}.zip",
        ]:
            log.info(f"[ABSL] Trying {fname}")
            data = _curl(base + fname, binary=True, timeout=60)
            if data and data[:4] == b'PK\x03\x04':
                target = base + fname
                break

    if target is None:
        log.warning("[ABSL] All URL strategies failed")
        return pd.DataFrame()

    # ── Step 3: download the zip ──────────────────────────────────────────────
    log.info(f"[ABSL] Downloading {target.split('/')[-1]}")
    data = _curl(target, binary=True, timeout=90)
    if not data or data[:4] != b'PK\x03\x04':
        log.warning("[ABSL] Download failed or not a ZIP")
        return pd.DataFrame()

    log.info(f"[ABSL] Downloaded zip {len(data):,} bytes")
    try:
        with zipfile.ZipFile(io.BytesIO(data)) as zf:
            xls_names = [n for n in zf.namelist()
                         if n.lower().endswith((".xls", ".xlsx"))
                         and not n.startswith("__MACOSX")]
            if not xls_names:
                log.warning("[ABSL] No Excel files in ZIP")
                return pd.DataFrame()
            raw_xls = zf.read(xls_names[0])
            log.info(f"[ABSL] Extracted {xls_names[0]} ({len(raw_xls):,} bytes)")
    except Exception as e:
        log.warning(f"[ABSL] ZIP unpack error: {e}")
        return pd.DataFrame()

    # OLE2 (.xls BIFF) format → xlrd; xlsx (PK magic) → openpyxl
    if raw_xls[:4] == b'\xd0\xcf\x11\xe0':
        df = _parse_absl_xls(raw_xls)
    else:
        wb = _load_workbook_bytes(raw_xls)
        df = _parse_workbook(wb) if wb else pd.DataFrame()

    log.info(f"[ABSL] {len(df)} holding rows across {df['scheme_name'].nunique() if not df.empty else 0} schemes")
    return df


def _download_nippon(month_str: str) -> pd.DataFrame:
    """
    Nippon India — consolidated xlsx (SharePoint, .xls extension but xlsx format)
    Pattern: NIMF-MONTHLY-PORTFOLIO-{DD}-{Month}-{YY}.xls
    """
    y, m, last, mname, mabbr, mlower, y2 = _month_parts(month_str)
    fname = f"NIMF-MONTHLY-PORTFOLIO-{last}-{mname}-{y2}.xls"
    url = f"https://mf.nipponindiaim.com/InvestorServices/FactsheetsDocuments/{fname}"
    log.info(f"[Nippon] {url}")
    data = _curl(url, binary=True, timeout=60)
    if not data:
        log.warning("[Nippon] Download failed")
        return pd.DataFrame()
    wb = _load_workbook_bytes(data)
    if wb is None:
        log.warning("[Nippon] Not a valid Excel file")
        return pd.DataFrame()
    # Nippon: Index sheet maps 2-letter codes to scheme names
    idx_sheet = next((s for s in wb.sheetnames if s.lower() == "index"), None)
    code_to_name: dict[str, str] = {}
    if idx_sheet:
        ws_idx = wb[idx_sheet]
        for row in ws_idx.iter_rows(values_only=True):
            if row[0] and isinstance(row[0], str) and len(row[0]) <= 5:
                code_to_name[row[0]] = str(row[1] or "")[:80]

    dfs = []
    for sn in wb.sheetnames:
        if sn.lower() == "index":
            continue
        ws = wb[sn]
        df = _parse_sheet(ws)
        if not df.empty:
            # Override scheme_name from index if available
            if sn in code_to_name and code_to_name[sn]:
                df["scheme_name"] = code_to_name[sn]
            df["_sheet"] = sn
            dfs.append(df)

    result = pd.concat(dfs, ignore_index=True) if dfs else pd.DataFrame()
    log.info(f"[Nippon] {len(result)} holding rows across {len(dfs)} schemes")
    return result


def _download_axis(month_str: str) -> pd.DataFrame:
    """
    Axis — consolidated xlsx (Drupal CDN)
    Pattern: Monthly%20Portfolio-{DD}%20{MM}%20{YY}.xlsx
    Primary CDN: transact.axismf.com (recent) / www.axismf.com (older)
    """
    y, m, last, mname, mabbr, mlower, y2 = _month_parts(month_str)
    mm  = _month_num_str(m)
    # URL-encoded filename: "Monthly Portfolio-30 04 26.xlsx"
    fname = f"Monthly%20Portfolio-{last}%20{mm}%20{y2}.xlsx"
    for base in [
        "https://transact.axismf.com/cms/sites/default/files/Statutory/",
        "https://www.axismf.com/cms/sites/default/files/Statutory/",
    ]:
        url = base + fname
        log.info(f"[Axis] Trying {url}")
        data = _curl(url, binary=True, timeout=60)
        if data and data[:4] == b'PK\x03\x04':
            wb = _load_workbook_bytes(data)
            if wb is None:
                continue
            df = _parse_workbook(wb)
            log.info(f"[Axis] {len(df)} holding rows")
            return df
    log.warning("[Axis] All download patterns failed")
    return pd.DataFrame()


def _download_dsp(month_str: str) -> pd.DataFrame:
    """
    DSP — zip containing xlsx files; URL has a content-hash that changes monthly.
    Strategy: scrape mandatory-disclosures/portfolio-disclosures to find current URL.
    File pattern: monthend-portfolios_{day}-{month}-{year}.zip
    """
    y, m, last, mname, mabbr, mlower, y2 = _month_parts(month_str)
    page_url = "https://www.dspim.com/mandatory-disclosures/portfolio-disclosures"
    log.info(f"[DSP] Scraping {page_url}")
    html = _curl(page_url, timeout=30)
    if not html:
        log.warning("[DSP] Page fetch failed")
        return pd.DataFrame()

    # Find monthend-portfolio zip URLs; pick the one matching our month
    all_zips = list(dict.fromkeys(re.findall(
        r'https://www\.dspim\.com/media/pages/mandatory-disclosures/portfolio-disclosures/'
        r'[^\s"\'<>\\]+/monthend-portfolio[^\s"\'<>\\]+\.zip',
        html, re.I
    )))

    # Match by month/year in filename
    target = None
    for url in all_zips:
        fname = url.split("/")[-1].lower()
        if mlower in fname and str(y) in fname:
            target = url
            break
        if mabbr.lower() in fname and str(y) in fname:
            target = url
            break

    if target is None and all_zips:
        # Fall back to first result (most recent)
        target = all_zips[0]
        log.warning(f"[DSP] Month match not found — using most recent: {target.split('/')[-1]}")

    if target is None:
        log.warning("[DSP] No zip URL found on page")
        return pd.DataFrame()

    log.info(f"[DSP] Downloading {target.split('/')[-1]}")
    data = _curl(target, binary=True, timeout=90)
    if not data or data[:4] != b'PK\x03\x04':
        log.warning("[DSP] Download failed or not a ZIP")
        return pd.DataFrame()

    df = _parse_zip(data, month_str)
    log.info(f"[DSP] {len(df)} holding rows")
    return df


def _download_franklin(month_str: str) -> pd.DataFrame:
    """
    Franklin Templeton — consolidated xlsx with UUID in URL.
    Strategy: scrape AdvisorKhoj to find the current URL.
    """
    y, m, last, mname, mabbr, mlower, y2 = _month_parts(month_str)
    page_url = (
        "https://www.advisorkhoj.com/form-download-centre/Mutual/"
        "Franklin-Templeton-Mutual-Fund/Monthly-Portfolio-Disclosures"
    )
    log.info(f"[Franklin] Scraping {page_url}")
    html = _curl(page_url, timeout=30)
    if not html:
        log.warning("[Franklin] Page fetch failed")
        return pd.DataFrame()

    # Find all franklintempletonindia.com xlsx URLs
    all_urls = list(dict.fromkeys(re.findall(
        r'https://www\.franklintempletonindia\.com/download/[^\s"\'<>\\]+\.xlsx',
        html, re.I
    )))

    # Pick URL matching our month
    target = None
    for url in all_urls:
        fname = url.split("/")[-1].lower()
        if mlower in fname and str(y) in fname:
            target = url
            break
        if mabbr.lower() in fname and str(y) in fname:
            target = url
            break

    if target is None and all_urls:
        target = all_urls[0]
        log.warning(f"[Franklin] Month match not found — using most recent: {target.split('/')[-1]}")

    if target is None:
        log.warning("[Franklin] No xlsx URL found on AdvisorKhoj page")
        return pd.DataFrame()

    log.info(f"[Franklin] Downloading {target.split('/')[-1]}")
    data = _curl(target, binary=True, timeout=60)
    if not data:
        log.warning("[Franklin] Download failed")
        return pd.DataFrame()
    wb = _load_workbook_bytes(data)
    if wb is None:
        log.warning("[Franklin] Not a valid Excel file")
        return pd.DataFrame()
    df = _parse_workbook(wb)
    log.info(f"[Franklin] {len(df)} holding rows across {df['scheme_name'].nunique() if not df.empty else 0} schemes")
    return df


def _download_mirae(month_str: str) -> pd.DataFrame:
    """
    Mirae Asset — Sitefinity CMS returns HTML for any .xlsx URL (false 200s).

    TO FIX: From your browser, navigate to:
      https://www.miraeassetmf.co.in/downloads/portfolio
    Find the "Monthly Portfolio Disclosure" Excel link for the relevant month,
    right-click → Copy link address, and paste it below as `url`.

    Fallback: manually downloaded file named mirae_portfolio_manual.xlsx
    """
    log.warning(
        "[Mirae] Auto-download not yet implemented (Sitefinity wildcard routing). "
        "Please download the monthly portfolio Excel manually from "
        "https://www.miraeassetmf.co.in/downloads/portfolio "
        "and place it as: mirae_portfolio_manual.xlsx"
    )
    manual = Path("mirae_portfolio_manual.xlsx")
    if manual.exists():
        log.info("[Mirae] Using manual file")
        wb = openpyxl.load_workbook(manual)
        return _parse_workbook(wb)
    return pd.DataFrame()


# ─────────────────────────────────────────────────────────────────────────────
# Dispatch table
# ─────────────────────────────────────────────────────────────────────────────

AMC_DOWNLOADERS = {
    "SBI":          _download_sbi,
    "ICICI_Pru":    _download_icici,
    "HDFC":         _download_hdfc,
    "Kotak":        _download_kotak,
    "Aditya_Birla": _download_absl,
    "Nippon":       _download_nippon,
    "Axis":         _download_axis,
    "DSP":          _download_dsp,
    "Franklin":     _download_franklin,
    "Mirae":        _download_mirae,
}

# ─────────────────────────────────────────────────────────────────────────────
# Scheme-name → scheme_code fuzzy matcher
# ─────────────────────────────────────────────────────────────────────────────

def _normalise(s: str) -> str:
    """Strip common fund suffixes and normalise for fuzzy matching."""
    s = s.lower()
    for pat in [
        r'\(.*?\)',          # parentheses
        r'direct plan',
        r'regular plan',
        r'growth',
        r'idcw',
        r'dividend',
        r'monthly',
        r'quarterly',
        r'annual',
        r'half.?yearly',
        r'- plan',
        r' - ',
        r'fund$',
    ]:
        s = re.sub(pat, ' ', s)
    return re.sub(r'\s+', ' ', s).strip()


def build_scheme_matcher(scheme_df: pd.DataFrame):
    """
    Build a fuzzy matcher from a DataFrame with columns [scheme_code, scheme_name, amc].
    Returns a function: (amc_name, fund_name) → scheme_code or None
    """
    records = []
    for _, row in scheme_df.iterrows():
        records.append({
            "scheme_code":  row["scheme_code"],
            "scheme_name":  row["scheme_name"],
            "amc":          row["amc"],
            "_norm":        _normalise(row["scheme_name"]),
        })

    def match(amc: str, fund_name: str, threshold: float = 0.65) -> Optional[str]:
        norm = _normalise(fund_name)
        # Filter to same AMC first
        candidates = [r for r in records if r["amc"] == amc]
        if not candidates:
            candidates = records  # fallback: search all
        best_score, best_code = 0.0, None
        for r in candidates:
            score = SequenceMatcher(None, norm, r["_norm"]).ratio()
            if score > best_score:
                best_score, best_code = score, r["scheme_code"]
        if best_score >= threshold:
            return best_code
        return None

    return match

# ─────────────────────────────────────────────────────────────────────────────
# Main pipeline
# ─────────────────────────────────────────────────────────────────────────────

def download_all_holdings(month_str: str, amc_filter: Optional[str] = None) -> pd.DataFrame:
    """
    Download portfolio holdings for all AMCs (or just one if amc_filter is set).
    Returns unified DataFrame with columns:
        scheme_name | isin | stock_name | pct_nav | amc
    """
    amcs = ([amc_filter] if amc_filter else list(AMC_DOWNLOADERS.keys()))
    all_dfs = []
    for amc in amcs:
        if amc not in AMC_DOWNLOADERS:
            log.warning(f"Unknown AMC: {amc}")
            continue
        log.info(f"── Downloading {amc} ──────────────────────────────────")
        try:
            df = AMC_DOWNLOADERS[amc](month_str)
        except Exception as e:
            log.error(f"[{amc}] Unexpected error: {e}")
            df = pd.DataFrame()
        if not df.empty:
            df["amc"] = amc
            all_dfs.append(df)
        else:
            log.warning(f"[{amc}] No data obtained")

    if not all_dfs:
        return pd.DataFrame()
    combined = pd.concat(all_dfs, ignore_index=True)
    combined["as_of_date"] = month_str
    return combined


def attach_scheme_codes(holdings: pd.DataFrame, scheme_df: pd.DataFrame) -> pd.DataFrame:
    """
    Fuzzy-match scheme_name + amc → scheme_code.
    Adds 'scheme_code' column; rows with no match get NaN.
    """
    matcher = build_scheme_matcher(scheme_df)
    match_cache: dict[tuple, Optional[str]] = {}

    def _get_code(row):
        key = (row["amc"], row["scheme_name"])
        if key not in match_cache:
            match_cache[key] = matcher(row["amc"], row["scheme_name"])
        return match_cache[key]

    holdings = holdings.copy()
    holdings["scheme_code"] = holdings.apply(_get_code, axis=1)

    n_matched   = holdings["scheme_code"].notna().sum()
    n_total     = len(holdings)
    n_schemes_matched = holdings.loc[holdings["scheme_code"].notna(), "scheme_code"].nunique()
    log.info(
        f"Scheme matching: {n_matched}/{n_total} holding rows matched "
        f"({n_schemes_matched} unique schemes)"
    )
    # Report unmatched scheme names
    unmatched = (
        holdings[holdings["scheme_code"].isna()][["amc", "scheme_name"]]
        .drop_duplicates()
        .head(20)
    )
    if not unmatched.empty:
        log.warning(f"Unmatched schemes (top 20):\n{unmatched.to_string(index=False)}")

    return holdings


# ─────────────────────────────────────────────────────────────────────────────
# CLI entry point
# ─────────────────────────────────────────────────────────────────────────────

def main():
    parser = argparse.ArgumentParser(description="Download AMC monthly portfolio holdings")
    parser.add_argument("--month",   default=None,
                        help="YYYY-MM (default: previous calendar month)")
    parser.add_argument("--amc",     default=None,
                        help="Single AMC name (e.g. SBI, HDFC, Nippon)")
    parser.add_argument("--out-dir", default=".",
                        help="Output directory (default: current dir)")
    parser.add_argument("--scheme-list", default="scheme_list.parquet",
                        help="Parquet file with scheme_code, scheme_name, amc columns")
    args = parser.parse_args()

    # Default month = previous calendar month
    if args.month is None:
        today = datetime.date.today()
        first = today.replace(day=1)
        prev  = first - datetime.timedelta(days=1)
        args.month = prev.strftime("%Y-%m")

    log.info(f"Target month: {args.month}")

    # ── Download holdings ─────────────────────────────────────────────────────
    holdings = download_all_holdings(args.month, amc_filter=args.amc)
    if holdings.empty:
        log.error("No holdings downloaded — exiting")
        sys.exit(1)

    log.info(f"Total: {len(holdings):,} holding rows, "
             f"{holdings['scheme_name'].nunique()} schemes, "
             f"{holdings['amc'].value_counts().to_dict()}")

    # ── Attach scheme codes ───────────────────────────────────────────────────
    scheme_path = Path(args.scheme_list)
    if scheme_path.exists():
        scheme_df = pd.read_parquet(scheme_path)
        holdings  = attach_scheme_codes(holdings, scheme_df)
    else:
        log.warning(f"scheme_list not found at {scheme_path} — skipping code mapping")
        holdings["scheme_code"] = None

    # ── Save output ───────────────────────────────────────────────────────────
    out_dir = Path(args.out_dir)
    out_dir.mkdir(parents=True, exist_ok=True)

    suffix = args.month.replace("-", "")
    parquet_path = out_dir / f"amc_portfolio_{suffix}.parquet"
    csv_path     = out_dir / f"amc_portfolio_{suffix}.csv"

    holdings.to_parquet(parquet_path, index=False)
    holdings.to_csv(csv_path, index=False)

    log.info(f"Saved: {parquet_path}")
    log.info(f"Saved: {csv_path}")

    # Quick summary
    print("\n=== Holdings Summary ===")
    print(holdings.groupby("amc").agg(
        schemes  = ("scheme_name", "nunique"),
        holdings = ("isin",        "count"),
        matched  = ("scheme_code", lambda x: x.notna().sum()),
    ).to_string())


if __name__ == "__main__":
    main()
